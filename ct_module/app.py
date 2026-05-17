"""
Weighbridge & Cycle Time Report Processing System  — v3
All critical fixes:
  1. WB: Merge both files → filter → ACCEPTED split FIRST → then formulas
  2. CT: Merge both files → filter → anomaly split FIRST (conditions only) → then formulas
  3. No Unnamed / junk columns anywhere
  4. Failure = strict 5-consecutive Exceeded=1 (AND, not SUM)
"""

from flask import Flask, request, jsonify, make_response
import io, json, traceback, os, sys

_HERE = os.path.dirname(os.path.abspath(__file__))
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__, static_folder='static', static_url_path='/static')

@app.after_request
def add_cors(resp):
    resp.headers['Access-Control-Allow-Origin']  = '*'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp


# ── Register /generate blueprint ──────────────────────────────────────────
try:
    from generate_routes import generate_bp
    app.register_blueprint(generate_bp)
    print('[generate] Blueprint registered at /generate')
except Exception as _ge:
    print(f'[generate] Blueprint skipped: {_ge}')



# ─────────────────────────────────────────────────────────────────────────────
#  Styling
# ─────────────────────────────────────────────────────────────────────────────

def _side(c='D0D0D0'): return Side(style='thin', color=c)
def _border(): s = _side(); return Border(left=s, right=s, top=s, bottom=s)
def _fill(h): return PatternFill('solid', fgColor=h)
def _font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='Calibri')
def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(ws, row, n, bg='1F3864', fg='FFFFFF', height=28):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = _fill(bg), _font(bold=True, color=fg, size=10)
        cell.alignment, cell.border = _align(h='center', wrap=True), _border()
    ws.row_dimensions[row].height = height

def style_data(ws, row, n, alt=False):
    bg = 'EBF0F8' if alt else 'FFFFFF'
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = _fill(bg), _font(size=9)
        cell.alignment, cell.border = _align(), _border()
    ws.row_dimensions[row].height = 18

def style_banner(ws, row, n, text, bg='1F3864', fg='FFFFFF'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(n, 1))
    cell = ws.cell(row=row, column=1)
    cell.value, cell.fill = text, _fill(bg)
    cell.font, cell.alignment = _font(bold=True, color=fg, size=11), _align(h='center')
    ws.row_dimensions[row].height = 26

def auto_width(ws, mn=8, mx=30):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value)) for c in col if c.value not in (None, '')]
        ws.column_dimensions[letter].width = min(max(max(lengths, default=0) + 2, mn), mx)


# ─────────────────────────────────────────────────────────────────────────────
#  Column helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s):
    return s.lower().replace(' ', '').replace('_', '').replace('/', '').replace(':', '')

def find_col(columns, *keywords):
    """Return first column whose normalised name contains ALL keywords."""
    for col in columns:
        n = _norm(col)
        if all(_norm(k) in n for k in keywords):
            return col
    return None

# Pre-existing computed column names to strip from raw input
WB_COMPUTED = {'formathour', 'formathours', 'formathr', 'peakcalculation', 'peakcalc'}
CT_COMPUTED = {
    'formathour', 'formathours', 'formathr', 'peakcalculation', 'peakcalc',
    'avgcycletime', 'averagecycletime', 'peakhours', 'exceeded', 'failure', 'totalcount',
}

def drop_computed(df, cset):
    """Drop columns whose normalised name is in cset."""
    drop = [c for c in df.columns if _norm(c) in cset]
    return df.drop(columns=drop, errors='ignore')

def drop_unnamed(df):
    """Remove any 'Unnamed: N' or truly empty-name columns."""
    keep = [c for c in df.columns
            if not str(c).startswith('Unnamed') and str(c).strip() != '']
    return df[keep]

def write_val(cell, v):
    if isinstance(v, pd.Timestamp):
        cell.value = v.strftime('%d-%m-%Y %H:%M')
    elif not isinstance(v, str) and pd.isna(v):
        cell.value = ''
    else:
        cell.value = v

def parse_dt_col(df, col):
    df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  Merge helper
# ─────────────────────────────────────────────────────────────────────────────

def merge_files(df1, df2=None):
    """
    Merge two DataFrames vertically.
    - Drop unnamed / junk columns from each before merging.
    - Drop pre-existing computed columns so we re-calculate cleanly.
    - Align columns: use df1 columns as reference; df2 gets same cols (fill missing with NaN).
    """
    df1 = drop_unnamed(df1.copy())

    if df2 is not None:
        df2 = drop_unnamed(df2.copy())
        # Align to df1's columns
        for c in df1.columns:
            if c not in df2.columns:
                df2[c] = pd.NA
        df2 = df2[df1.columns]
        merged = pd.concat([df1, df2], ignore_index=True)
    else:
        merged = df1

    return merged


# ─────────────────────────────────────────────────────────────────────────────
#  1. WEIGHBRIDGE TOTAL
#     Order: Merge → Filter → Split ACCEPTED → Strip computed → Write formulas
# ─────────────────────────────────────────────────────────────────────────────

def process_wb(df1, df2, start_dt, end_dt):
    """
    Returns (df_accepted, df_rejected, wb_in_col)

    df_accepted : rows with ACCEPTED != NO, computed cols stripped, ready for formulas
    df_rejected : rows with ACCEPTED == NO, full raw columns (no formula cols)
    """
    # Step 1 – Merge
    df = merge_files(df1, df2)

    # Step 2 – Locate WB IN TIME and filter
    wb_in_col = find_col(df.columns, 'wb', 'in', 'time')
    if not wb_in_col:
        raise ValueError(f"'WB IN TIME' column not found. Available: {list(df.columns)}")

    df = parse_dt_col(df, wb_in_col)
    df = df[(df[wb_in_col] >= start_dt) & (df[wb_in_col] <= end_dt)].reset_index(drop=True)

    # Step 3 – VALIDATE ACCEPTED IMMEDIATELY (before any column changes)
    acc_col = find_col(df.columns, 'accepted')
    if acc_col:
        is_rej = df[acc_col].astype(str).str.strip().str.upper() == 'NO'
        df_rejected = df[is_rej].copy().reset_index(drop=True)    # full raw row
        df_accepted = df[~is_rej].copy().reset_index(drop=True)   # accepted only
    else:
        df_rejected = pd.DataFrame(columns=df.columns)
        df_accepted = df.copy()

    # Step 4 – Strip pre-existing computed cols from ACCEPTED only
    df_accepted = drop_computed(df_accepted, WB_COMPUTED)

    return df_accepted, df_rejected, wb_in_col


def write_wb_sheets(wb_out, df_acc, df_rej, wb_in_col):
    """
    WB Total    → accepted rows + Format Hours + Peak Calculation formulas
    WB Rejected → raw rejected rows, no formula columns added
    """

    # ── WB Total ─────────────────────────────────────────────────────────
    orig    = list(df_acc.columns)
    wb_pos  = orig.index(wb_in_col)           # 0-based
    wb_L    = get_column_letter(wb_pos + 1)   # Excel letter

    # Insert Format Hours + Peak Calculation right after WB IN TIME
    out_cols = orig[:wb_pos + 1] + ['Format Hours', 'Peak Calculation'] + orig[wb_pos + 1:]
    N = len(out_cols)

    ws = wb_out.create_sheet('WB Total')
    for ci, name in enumerate(out_cols, 1):
        ws.cell(row=1, column=ci, value=name)
    style_header(ws, 1, N)
    ws.freeze_panes = 'A2'

    for ri, (_, row) in enumerate(df_acc.iterrows(), 2):
        style_data(ws, ri, N, alt=(ri % 2 == 0))
        for ci, col in enumerate(out_cols, 1):
            cell = ws.cell(row=ri, column=ci)
            if col == 'Format Hours':
                cell.value = (
                    f'=TEXT({wb_L}{ri},"hh")' if ri == 2 else
                    f'=IF(TEXT({wb_L}{ri},"hh")=TEXT({wb_L}{ri-1},"hh"),"",TEXT({wb_L}{ri},"hh"))'
                )
            elif col == 'Peak Calculation':
                cell.value = f'=VALUE(TEXT({wb_L}{ri},"hhmm"))'
            elif col in row.index:
                write_val(cell, row[col])
    auto_width(ws)

    # ── WB Rejected Loads ────────────────────────────────────────────────
    ws_r = wb_out.create_sheet('WB Rejected Loads')
    rcols = list(df_rej.columns)
    NR = max(len(rcols), 3)
    style_banner(ws_r, 1, NR, 'Weighbridges Rejected & Reclassified Loads', bg='7B0000')
    for ci, name in enumerate(rcols, 1):
        ws_r.cell(row=2, column=ci, value=name)
    style_header(ws_r, 2, len(rcols), bg='C00000')
    ws_r.freeze_panes = 'A3'
    for ri, (_, row) in enumerate(df_rej.iterrows(), 3):
        for ci, col in enumerate(rcols, 1):
            write_val(ws_r.cell(row=ri, column=ci), row[col])
        style_data(ws_r, ri, len(rcols), alt=(ri % 2 == 1))
    auto_width(ws_r)


# ─────────────────────────────────────────────────────────────────────────────
#  2. CYCLE TIME / ONLINE TOTAL
#     Order: Merge → Filter → Anomaly split → Strip computed → Write formulas
# ─────────────────────────────────────────────────────────────────────────────

ANOMALY_DELAY = 0    # minutes — default threshold for queue/lag anomaly conditions

def process_ct(df1, df2, start_dt, end_dt, main_threshold,
               queue_minutes=None, lag_minutes=None):
    """
    Anomaly conditions (applied BEFORE any formula columns are added):

      Condition 1: WB In Time − Date Time Arrival > queue_minutes  (default: ANOMALY_DELAY)
      Condition 2: Date Time Exit − WB Out Time   > lag_minutes    (default: ANOMALY_DELAY)
      Condition 3: Duration column from uploaded Excel > main_threshold min

    ANY one TRUE → row goes to Anomaly & Not Considered.
    Remaining rows → Online Total (formulas applied after split).

    Returns (df_online, df_anomaly, arr_col, dur_col)
    """
    _queue = queue_minutes if queue_minutes is not None else ANOMALY_DELAY
    _lag   = lag_minutes   if lag_minutes   is not None else ANOMALY_DELAY
    # Step 1 – Merge + drop junk columns
    df = merge_files(df1, df2)

    # Step 2 – Locate key columns
    arr_col    = find_col(df.columns, 'datetime', 'arrival') or find_col(df.columns, 'date', 'arrival')
    exit_col   = find_col(df.columns, 'datetime', 'exit')   or find_col(df.columns, 'date', 'exit')
    wb_in_col  = find_col(df.columns, 'wb', 'in', 'time')
    wb_out_col = find_col(df.columns, 'wb', 'out', 'time')

    if not arr_col:
        raise ValueError(f"'Date Time Arrival' not found. Available: {list(df.columns)}")
    if not exit_col:
        raise ValueError(f"'Date Time Exit' not found. Available: {list(df.columns)}")

    # Step 3 – Parse datetimes
    for col in [arr_col, exit_col, wb_in_col, wb_out_col]:
        if col:
            df = parse_dt_col(df, col)

    # Step 4 – Filter by Date Time Arrival range
    df = df[(df[arr_col] >= start_dt) & (df[arr_col] <= end_dt)].reset_index(drop=True)

    # Step 5 – Anomaly conditions
    # Cond1: WB In Time – Date Time Arrival > queue threshold
    # SKIPPED ENTIRELY when _queue == 0 (zero means "ignore this check")
    if _queue > 0 and wb_in_col:
        cond1 = (df[wb_in_col] - df[arr_col]).dt.total_seconds() / 60 > _queue
    else:
        cond1 = pd.Series(False, index=df.index)

    # Cond2: Date Time Exit – WB Out Time > lag threshold
    # SKIPPED ENTIRELY when _lag == 0 (zero means "ignore this check")
    if _lag > 0 and wb_out_col:
        cond2 = (df[exit_col] - df[wb_out_col]).dt.total_seconds() / 60 > _lag
    else:
        cond2 = pd.Series(False, index=df.index)

    # Cond3: use Duration column from uploaded Excel directly — no datetime subtraction.
    # Falls back to datetime calculation only if Duration column is absent.
    main_dur = (df[exit_col] - df[arr_col]).dt.total_seconds() / 60
    dur_col_existing_pre = find_col(df.columns, 'duration')
    if dur_col_existing_pre:
        dur_vals_pre = pd.to_numeric(df[dur_col_existing_pre], errors='coerce').fillna(0)
        cond3 = dur_vals_pre > main_threshold
    else:
        cond3 = main_dur > main_threshold

    is_anomaly = cond1 | cond2 | cond3

    # Step 6 – Write Duration column (recalculated, replaces any existing)
    dur_col_existing = find_col(df.columns, 'duration')
    dur_vals = main_dur.round(0).astype('Int64', errors='ignore')
    if dur_col_existing:
        df[dur_col_existing] = dur_vals
        dur_col = dur_col_existing
    else:
        insert_pos = df.columns.get_loc(exit_col) + 1
        df.insert(insert_pos, 'Duration', dur_vals)
        dur_col = 'Duration'

    # Step 7 – SPLIT (before any computed columns are added)
    df_anomaly = df[is_anomaly].copy().reset_index(drop=True)
    df_online  = df[~is_anomaly].copy().reset_index(drop=True)

    # Step 8 – Strip pre-existing computed cols from ONLINE only
    df_online = drop_computed(df_online, CT_COMPUTED)

    return df_online, df_anomaly, arr_col, dur_col


def write_ct_sheets(wb_out, df_on, df_an, arr_col, dur_col):
    """
    Online Total  → clean data + all formula columns (after anomaly removal)
    Anomaly sheet → raw copied rows, NO extra columns, NO unnamed cols, S/No. prepended
    """

    orig    = list(df_on.columns)
    arr_pos = orig.index(arr_col)
    dur_pos = orig.index(dur_col) if dur_col in orig else len(orig) - 1

    METRICS = ['Avg Cycle Time', 'Peak Hours', 'Exceeded', 'Failure', 'Total Count']

    # Final column order:
    # [cols up to arr] + [FmtHr, PkCalc] + [cols after arr up to+incl dur] + [metrics]
    out_cols = (
        orig[:arr_pos + 1]
        + ['Format Hours', 'Peak Calculation']
        + orig[arr_pos + 1:dur_pos + 1]
        + METRICS
    )
    # Note: columns after Duration (if any) are excluded — they are junk/unnamed

    def L(name): return get_column_letter(out_cols.index(name) + 1)

    arr_L  = L(arr_col)
    pc_L   = L('Peak Calculation')
    dur_L  = L(dur_col)
    avg_L  = L('Avg Cycle Time')
    phrs_L = L('Peak Hours')
    exc_L  = L('Exceeded')
    fail_L = L('Failure')
    N      = len(out_cols)

    # ── Online Total ──────────────────────────────────────────────────────
    ws = wb_out.create_sheet('Online Total')
    for ci, name in enumerate(out_cols, 1):
        ws.cell(row=1, column=ci, value=name)
    style_header(ws, 1, N)
    ws.freeze_panes = 'A2'

    for ri, (_, row) in enumerate(df_on.iterrows(), 2):
        style_data(ws, ri, N, alt=(ri % 2 == 0))
        dr = ri - 1   # 1-based data row number

        for ci, col in enumerate(out_cols, 1):
            cell = ws.cell(row=ri, column=ci)

            if col == 'Format Hours':
                cell.value = (
                    f'=TEXT({arr_L}{ri},"hh")' if ri == 2 else
                    f'=IF(TEXT({arr_L}{ri},"hh")=TEXT({arr_L}{ri-1},"hh"),"",TEXT({arr_L}{ri},"hh"))'
                )

            elif col == 'Peak Calculation':
                cell.value = f'=VALUE(TEXT({arr_L}{ri},"hhmm"))'

            elif col == 'Avg Cycle Time':
                # Blank for first 9 data rows; rolling 10 from data row 10 onwards
                if dr < 10:
                    cell.value = ''
                else:
                    cell.value = f'=AVERAGE({dur_L}{ri-9}:{dur_L}{ri})'

            elif col == 'Peak Hours':
                # 45 if 09:00–<11:30 or 15:00–<17:30, else 30
                cell.value = (
                    f'=IF(OR(AND({pc_L}{ri}>=900,{pc_L}{ri}<1130),'
                    f'AND({pc_L}{ri}>=1500,{pc_L}{ri}<1730)),45,30)'
                )

            elif col == 'Exceeded':
                # Blank until Avg available (dr < 10)
                if dr < 10:
                    cell.value = ''
                else:
                    # 0 = OK (Peak Hours >= Avg), 1 = Exceeded
                    cell.value = (
                        f'=IF({avg_L}{ri}="","",'
                        f'IF({phrs_L}{ri}>={avg_L}{ri},0,1))'
                    )

            elif col == 'Failure':
                # Needs 5 consecutive Exceeded=1
                # Avg available from dr=10 → first 5 Exceeded at dr=10..14 → first Failure at dr=14
                if dr < 14:
                    cell.value = ''
                else:
                    e = exc_L
                    # STRICT: AND each of the 5 cells equals 1 individually (not SUM)
                    # This correctly resets on any 0
                    cell.value = (
                        f'=IF(AND('
                        f'{e}{ri}<>"",'
                        f'{e}{ri-1}<>"",'
                        f'{e}{ri-2}<>"",'
                        f'{e}{ri-3}<>"",'
                        f'{e}{ri-4}<>"",'
                        f'{e}{ri}=1,'
                        f'{e}{ri-1}=1,'
                        f'{e}{ri-2}=1,'
                        f'{e}{ri-3}=1,'
                        f'{e}{ri-4}=1),'
                        f'"fail","")'
                    )

            elif col == 'Total Count':
                # Only in row 2; counts all "fail" in Failure column
                cell.value = f'=COUNTIF({fail_L}:{fail_L},"fail")' if ri == 2 else ''

            elif col in row.index:
                write_val(cell, row[col])

    auto_width(ws)

    # ── Anomaly & Not Considered ──────────────────────────────────────────
    # Clean columns only: no Unnamed, no computed/formula cols — plain raw data
    an_clean = drop_unnamed(drop_computed(df_an.copy(), CT_COMPUTED))
    an_src   = list(an_clean.columns)
    an_cols  = ['S/No.'] + an_src
    NA       = len(an_cols)

    ws_a = wb_out.create_sheet('Anomaly & Not Considered')
    style_banner(ws_a, 1, NA, 'ANOMALY ADJUSTMENT & THOSE NOT CONSIDERED', bg='833C00')
    for ci, name in enumerate(an_cols, 1):
        ws_a.cell(row=2, column=ci, value=name)
    style_header(ws_a, 2, NA, bg='BF5F00')
    ws_a.freeze_panes = 'A3'

    for ri, (_, row) in enumerate(an_clean.iterrows(), 3):
        ws_a.cell(row=ri, column=1, value=ri - 2)   # S/No.
        for ci, col in enumerate(an_src, 2):
            write_val(ws_a.cell(row=ri, column=ci), row[col])
        style_data(ws_a, ri, NA, alt=(ri % 2 == 1))

    auto_width(ws_a)


# ─────────────────────────────────────────────────────────────────────────────
#  Flask routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/api/process', methods=['POST', 'OPTIONS'])
def report_process():
    try:
        rt        = request.form.get('report_type', 'wb_total')
        start_str = request.form.get('start_datetime', '')
        end_str   = request.form.get('end_datetime', '')
        threshold      = float(request.form.get('anomaly_threshold', 100))
        thresh_queue   = float(request.form.get('thresh_queue',    ANOMALY_DELAY))
        thresh_lag     = float(request.form.get('thresh_lag',      ANOMALY_DELAY))
        wb_sh1    = request.form.get('wb_sheet', '0')
        wb_sh2    = request.form.get('wb_sheet2', '0')
        ct_sh1    = request.form.get('ct_sheet', '0')
        ct_sh2    = request.form.get('ct_sheet2', '0')

        if not start_str or not end_str:
            return jsonify({'error': 'Start and End datetime are required.'}), 400

        start_dt = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        end_dt   = datetime.strptime(end_str,   '%Y-%m-%dT%H:%M')
        if start_dt >= end_dt:
            return jsonify({'error': 'Start must be before End.'}), 400

        file1 = request.files.get('file1')
        file2 = request.files.get('file2')
        if not file1:
            return jsonify({'error': 'File 1 is required.'}), 400

        def sid(s):
            try: return int(s)
            except: return s or 0

        # Read both uploaded files (df2 = None if not supplied)
        df1_wb = pd.read_excel(file1, sheet_name=sid(wb_sh1))
        file1.seek(0)                                               # reset for second read
        df1_ct = pd.read_excel(file1, sheet_name=sid(ct_sh1))

        df2_wb = df2_ct = None
        if file2:
            df2_wb = pd.read_excel(file2, sheet_name=sid(wb_sh2))
            file2.seek(0)
            df2_ct = pd.read_excel(file2, sheet_name=sid(ct_sh2))

        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        stats  = {}

        if rt == 'wb_total':
            df_acc, df_rej, wc = process_wb(df1_wb, df2_wb, start_dt, end_dt)
            write_wb_sheets(wb_out, df_acc, df_rej, wc)
            stats = {'WB Accepted': len(df_acc), 'WB Rejected': len(df_rej),
                     'Total Filtered': len(df_acc) + len(df_rej)}
            fname = f"WB_Total_{start_dt.strftime('%d%m%Y')}.xlsx"

        elif rt == 'cycle_time':
            df_on, df_an, arr, dur = process_ct(df1_ct, df2_ct, start_dt, end_dt, threshold,
                                                  queue_minutes=thresh_queue, lag_minutes=thresh_lag)
            write_ct_sheets(wb_out, df_on, df_an, arr, dur)
            stats = {'Online Total': len(df_on), 'Anomaly Rows': len(df_an),
                     'Total Filtered': len(df_on) + len(df_an)}
            fname = f"Cycle_Time_{start_dt.strftime('%d%m%Y')}.xlsx"

        elif rt == 'both':
            df_acc, df_rej, wc = process_wb(df1_wb, df2_wb, start_dt, end_dt)
            write_wb_sheets(wb_out, df_acc, df_rej, wc)
            df_on, df_an, arr, dur = process_ct(df1_ct, df2_ct, start_dt, end_dt, threshold,
                                                  queue_minutes=thresh_queue, lag_minutes=thresh_lag)
            write_ct_sheets(wb_out, df_on, df_an, arr, dur)
            stats = {'WB Accepted': len(df_acc), 'WB Rejected': len(df_rej),
                     'CT Online': len(df_on), 'CT Anomaly': len(df_an)}
            fname = f"Combined_{start_dt.strftime('%d%m%Y')}.xlsx"
        else:
            return jsonify({'error': f'Unknown report_type: {rt}'}), 400

        buf = io.BytesIO()
        wb_out.save(buf); buf.seek(0)
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp.headers['X-Stats']    = json.dumps(stats)
        resp.headers['X-Filename'] = fname
        return resp

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500






# ════════════════════════════════════════════════════════════════════
#  APSG REPORT ROUTES  (/report/*)
# ════════════════════════════════════════════════════════════════════

@app.route('/version')
def rpt_version():
    return jsonify({'version': 5, 'ok': True})

# ── In-memory store ───────────────────────────────────────────────────────────
_rpt_store = {
    'df_ct': None, 'df_anomaly': None,
    'df_wb_total': None, 'df_wb_rejected': None,
    'failure_list': [], 'exceedances': 0,
    'applicable_hours': 0, 'net_weight': 0.0,
    'start_dt': None, 'end_dt': None, 'reason': '',
    'main_xlsx': None, 'summary_xlsx': None, 'ppt_bytes': None,
}

# ── Lazy-import heavy modules so startup errors are caught cleanly ─────────────
def _get_engine():
    try:
        import report_engine as eng
        return eng
    except Exception as e:
        raise RuntimeError(f"Failed to import report_engine: {e}\n{traceback.format_exc()}")

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/report/process', methods=['POST', 'OPTIONS'])
def rpt_process():
    if request.method == 'OPTIONS':
        return '', 200
    global _rpt_store
    try:
        eng = _get_engine()
        ct_files    = request.files.getlist('ct_files')
        online_file = request.files.get('online_file')
        from_s      = request.form.get('from_dt', '')
        to_s        = request.form.get('to_dt', '')
        reason      = request.form.get('reason', '')

        if not ct_files or ct_files[0].filename == '':
            return jsonify({'error': 'No Cycle Time files uploaded'}), 400
        if not online_file or online_file.filename == '':
            return jsonify({'error': 'No Online Data file uploaded'}), 400

        try:
            start_dt = datetime.fromisoformat(from_s) if from_s else None
            end_dt   = datetime.fromisoformat(to_s)   if to_s   else None
        except ValueError:
            start_dt = end_dt = None

        ct_ios = [io.BytesIO(f.read()) for f in ct_files]
        on_io  = io.BytesIO(online_file.read())

        # Read the 3 UI-configurable anomaly thresholds (with safe defaults)
        thresh_queue    = float(request.form.get('thresh_queue',    0))
        thresh_lag      = float(request.form.get('thresh_lag',      0))
        thresh_duration = float(request.form.get('thresh_duration', 120))
        min_duration    = float(request.form.get('min_duration',    5))    # lower-limit (default 5 min)

        df_ct, df_an, fl, exc, ah = eng.prepare_ct_data(
            ct_ios, start_dt, end_dt,
            queue_minutes=thresh_queue,
            lag_minutes=thresh_lag,
            duration_threshold=thresh_duration,
            min_duration_threshold=min_duration,
        )
        df_wb, df_rj, nw = eng.prepare_online_data(on_io, start_dt, end_dt)

        _rpt_store.update({
            'df_ct': df_ct, 'df_anomaly': df_an,
            'df_wb_total': df_wb, 'df_wb_rejected': df_rj,
            'failure_list': fl, 'exceedances': exc,
            'applicable_hours': ah, 'net_weight': nw,
            'start_dt': start_dt, 'end_dt': end_dt, 'reason': reason,
            'main_xlsx': None, 'summary_xlsx': None, 'ppt_bytes': None,
        })
        return jsonify({'ok': True, 'ct_records': len(df_ct), 'wb_records': len(df_wb),
                        'exceedances': exc, 'applicable_hours': ah, 'net_weight': nw})
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/report/build/main', methods=['POST', 'OPTIONS'])
def rpt_build_main():
    if request.method == 'OPTIONS':
        return '', 200
    global _rpt_store
    try:
        if _rpt_store['df_ct'] is None:
            return jsonify({'error': 'Run /report/process first'}), 400
        eng  = _get_engine()
        xlsx = eng.build_main_report(
            _rpt_store['df_ct'], _rpt_store['df_anomaly'],
            _rpt_store['df_wb_total'], _rpt_store['df_wb_rejected'],
            _rpt_store['failure_list'], _rpt_store['start_dt'], _rpt_store['end_dt'])
        _rpt_store['main_xlsx'] = xlsx
        return jsonify({'ok': True, 'size_kb': round(len(xlsx)/1024)})
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/report/build/summary', methods=['POST', 'OPTIONS'])
def rpt_build_summary():
    if request.method == 'OPTIONS':
        return '', 200
    global _rpt_store
    try:
        if _rpt_store['df_wb_total'] is None:
            return jsonify({'error': 'Run /report/process first'}), 400
        eng       = _get_engine()
        demo_path = os.path.join(_HERE, 'Demo.xlsx')
        xlsx      = eng.build_summary_report(
            _rpt_store['df_wb_total'], demo_path,
            start_dt=_rpt_store.get('start_dt'),
            end_dt=_rpt_store.get('end_dt'),
        )
        _rpt_store['summary_xlsx'] = xlsx
        return jsonify({'ok': True, 'size_kb': round(len(xlsx)/1024)})
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/report/build/ppt', methods=['POST', 'OPTIONS'])
def rpt_build_ppt():
    if request.method == 'OPTIONS':
        return '', 200
    global _rpt_store
    try:
        if _rpt_store['df_ct'] is None:
            return jsonify({'error': 'Run /report/process first'}), 400
        if _rpt_store['summary_xlsx'] is None:
            return jsonify({'error': 'Build summary first'}), 400
        eng = _get_engine()
        ppt = eng.build_ppt_report(
            _rpt_store['df_ct'], _rpt_store['df_wb_total'],
            reason=_rpt_store['reason'], exceedances=_rpt_store['exceedances'],
            applicable_hours=_rpt_store['applicable_hours'],
            summary_xlsx_bytes=_rpt_store['summary_xlsx'],
            ppt_template_path=os.path.join(_HERE, 'demo.pptx'))
        _rpt_store['ppt_bytes'] = ppt
        return jsonify({'ok': True, 'size_kb': round(len(ppt)/1024)})
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/report/download/<kind>', methods=['GET'])
def rpt_download(kind):
    global _rpt_store
    _sd = _rpt_store.get('start_dt')
    _ed = _rpt_store.get('end_dt')
    # Build date suffix: single day → 'YYYYMMDD'; range → 'YYYYMMDD&YYYYMMDD'
    if _sd is None:
        _suffix = ''
    elif _ed is None or _sd.date() == _ed.date():
        _suffix = _sd.strftime('%Y%m%d')
    else:
        _suffix = f"{_sd.strftime('%Y%m%d')}&{_ed.strftime('%Y%m%d')}"

    _ct_name  = f"Cycle Time - {_suffix}" if _suffix else "Cycle Time"
    _hrq_name = f"APSG-Hourly Truck Quantity {_suffix}" if _suffix else "APSG-Hourly Truck Quantity"

    MAP = {
        'main':    ('main_xlsx',
                    f'{_ct_name}.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        'summary': ('summary_xlsx',
                    f'{_hrq_name}.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        'ppt':     ('ppt_bytes',
                    f'{_ct_name}.pptx',
                    'application/vnd.openxmlformats-officedocument.presentationml.presentation'),
    }
    if kind not in MAP:
        return jsonify({'error': 'Unknown type'}), 404
    key, fname, mime = MAP[kind]
    data = _rpt_store.get(key)
    if not data:
        return jsonify({'error': f'{kind} not generated yet'}), 404
    resp = make_response(data)
    resp.headers['Content-Type'] = mime
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


if __name__ == '__main__':
    import threading, webbrowser, time
    threading.Thread(target=lambda: (time.sleep(1.5), webbrowser.open('http://localhost:5050')), daemon=True).start()
    app.run(host='0.0.0.0', port=5050, debug=False)


# ─────────────────────────────────────────────────────────────────────────────
#  PHASE 1 — Upload, Merge, Filter
#  Handles CT + WB files independently; filters both by WB In Time
# ─────────────────────────────────────────────────────────────────────────────

# In-memory store for the current session's merged+filtered datasets
_phase1_store = {'ct': None, 'wb': None}


@app.route('/phase1')
def phase1_ui():
    return app.send_static_file('phase1.html')


def _sid(s):
    """Convert sheet selector string to int index or name string."""
    try:   return int(s)
    except: return s or 0


def _read_and_merge(file1, file2, sheet1, sheet2):
    """
    Read one or two uploaded Excel files and merge them vertically.
    - Drops Unnamed columns from each file before merging.
    - Aligns columns: uses file1's columns as reference.
    Returns a single combined DataFrame.
    """
    df1 = pd.read_excel(file1, sheet_name=_sid(sheet1))
    df1 = drop_unnamed(df1)

    files_merged = 1
    if file2:
        df2 = pd.read_excel(file2, sheet_name=_sid(sheet2))
        df2 = drop_unnamed(df2)
        # Align to df1's column set
        for c in df1.columns:
            if c not in df2.columns:
                df2[c] = pd.NA
        df2 = df2[df1.columns]
        df1 = pd.concat([df1, df2], ignore_index=True)
        files_merged = 2

    return df1, files_merged


def _find_wb_in_col(columns):
    """Locate the WB In Time column (case/space insensitive)."""
    col = find_col(columns, 'wb', 'in', 'time')
    if col is None:
        col = find_col(columns, 'wbintime')
    return col


def _filter_by_wb_in_time(df, start_dt, end_dt):
    """
    Filter DataFrame to rows where WB In Time is within [start_dt, end_dt].
    Returns (filtered_df, wb_in_col_name)
    Raises ValueError if WB In Time column cannot be found.
    """
    wb_in_col = _find_wb_in_col(df.columns)
    if wb_in_col is None:
        raise ValueError(
            f"'WB In Time' column not found. "
            f"Available columns: {list(df.columns)}"
        )
    df[wb_in_col] = pd.to_datetime(df[wb_in_col], dayfirst=True, errors='coerce')
    mask = (df[wb_in_col] >= start_dt) & (df[wb_in_col] <= end_dt)
    return df[mask].reset_index(drop=True), wb_in_col


def _wb_in_range_str(df, col):
    """Return human-readable 'min → max' string for the WB In Time column."""
    valid = df[col].dropna()
    if valid.empty:
        return '—'
    return f"{valid.min().strftime('%d/%m %H:%M')} → {valid.max().strftime('%d/%m %H:%M')}"


def _build_preview(df, max_rows=10):
    """Build a JSON-serialisable preview dict from a DataFrame."""
    preview_df = df.head(max_rows).copy()
    # Convert Timestamps to strings
    for col in preview_df.columns:
        if pd.api.types.is_datetime64_any_dtype(preview_df[col]):
            preview_df[col] = preview_df[col].dt.strftime('%d-%m-%Y %H:%M')
    return {
        'columns':    list(preview_df.columns),
        'rows':       preview_df.where(pd.notnull(preview_df), None).values.tolist(),
        'total_rows': len(df),
    }


def _df_to_xlsx_bytes(df):
    """Serialise DataFrame to xlsx bytes with basic styling."""
    from openpyxl import Workbook as WB
    wb_xls = WB()
    ws_xls = wb_xls.active

    # Write header
    for ci, col in enumerate(df.columns, 1):
        cell = ws_xls.cell(row=1, column=ci, value=col)
        cell.fill   = _fill('1F3864')
        cell.font   = _font(bold=True, color='FFFFFF', size=10)
        cell.alignment = _align(h='center', wrap=True)
        cell.border = _border()
    ws_xls.row_dimensions[1].height = 26

    # Write data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        alt = (ri % 2 == 0)
        for ci, col in enumerate(df.columns, 1):
            cell = ws_xls.cell(row=ri, column=ci)
            v = row[col]
            if isinstance(v, pd.Timestamp):
                cell.value = v.strftime('%d-%m-%Y %H:%M')
            elif not isinstance(v, str) and pd.isna(v):
                cell.value = ''
            else:
                cell.value = v
            cell.fill      = _fill('EBF0F8' if alt else 'FFFFFF')
            cell.font      = _font(size=9)
            cell.alignment = _align()
            cell.border    = _border()
        ws_xls.row_dimensions[ri].height = 17

    auto_width(ws_xls)
    ws_xls.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb_xls.save(buf)
    buf.seek(0)
    return buf.read()


@app.route('/phase1/process', methods=['POST', 'OPTIONS'])
def phase1_process():
    """
    Phase 1 endpoint:
      1. Read + merge CT files (File1 + File2)
      2. Read + merge WB files (File1 + File2)
      3. Filter both merged datasets by WB In Time within [start_dt, end_dt]
      4. Store results in _phase1_store
      5. Return metadata + preview as JSON in response headers
    """
    global _phase1_store
    try:
        ct_file1 = request.files.get('ct_file1')
        ct_file2 = request.files.get('ct_file2')
        wb_file1 = request.files.get('wb_file1')
        wb_file2 = request.files.get('wb_file2')

        ct_sheet1 = request.form.get('ct_sheet1', '0')
        ct_sheet2 = request.form.get('ct_sheet2', '0')
        wb_sheet1 = request.form.get('wb_sheet1', '0')
        wb_sheet2 = request.form.get('wb_sheet2', '0')

        start_str = request.form.get('start_dt', '')
        end_str   = request.form.get('end_dt',   '')

        if not ct_file1:
            return jsonify({'error': 'Cycle Time File 1 is required.'}), 400
        if not wb_file1:
            return jsonify({'error': 'Weighbridge File 1 is required.'}), 400
        if not start_str or not end_str:
            return jsonify({'error': 'Start and End date/time are required.'}), 400

        start_dt = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        end_dt   = datetime.strptime(end_str,   '%Y-%m-%dT%H:%M')
        if start_dt >= end_dt:
            return jsonify({'error': 'Start must be before End.'}), 400

        # ── Step 1: Merge CT files ────────────────────────────────────────
        df_ct, ct_files_merged = _read_and_merge(ct_file1, ct_file2, ct_sheet1, ct_sheet2)

        # ── Step 2: Merge WB files ────────────────────────────────────────
        df_wb, wb_files_merged = _read_and_merge(wb_file1, wb_file2, wb_sheet1, wb_sheet2)

        # ── Step 3: Filter both by WB In Time ────────────────────────────
        df_ct_filtered, ct_wb_in_col = _filter_by_wb_in_time(df_ct, start_dt, end_dt)
        df_wb_filtered, wb_wb_in_col = _filter_by_wb_in_time(df_wb, start_dt, end_dt)

        # ── Step 4: Store in session ──────────────────────────────────────
        _phase1_store['ct']       = df_ct_filtered
        _phase1_store['wb']       = df_wb_filtered
        _phase1_store['start_dt'] = start_dt   # keep for Phase 2 chart title
        _phase1_store['end_dt']   = end_dt

        # ── Step 5: Build metadata + preview ─────────────────────────────
        meta = {
            'start_dt':          start_dt.strftime('%Y-%m-%dT%H:%M'),
            'end_dt':            end_dt.strftime('%Y-%m-%dT%H:%M'),
            'ct_rows':           len(df_ct_filtered),
            'ct_cols':           len(df_ct_filtered.columns),
            'ct_files_merged':   ct_files_merged,
            'ct_wb_in_range':    _wb_in_range_str(df_ct_filtered, ct_wb_in_col),
            'ct_preview':        _build_preview(df_ct_filtered),
            'wb_rows':           len(df_wb_filtered),
            'wb_cols':           len(df_wb_filtered.columns),
            'wb_files_merged':   wb_files_merged,
            'wb_wb_in_range':    _wb_in_range_str(df_wb_filtered, wb_wb_in_col),
            'wb_preview':        _build_preview(df_wb_filtered),
        }

        # Return a tiny placeholder body; real data downloaded via /phase1/download/
        resp = make_response(json.dumps({'ok': True}))
        resp.headers['Content-Type']  = 'application/json'
        resp.headers['X-Meta']        = json.dumps(meta)
        return resp

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@app.route('/phase1/download/<dtype>')
def phase1_download(dtype):
    """
    Serve the merged+filtered CT or WB dataset as an xlsx download.
    dtype: 'ct' or 'wb'
    """
    global _phase1_store
    df = _phase1_store.get(dtype)
    if df is None:
        return jsonify({'error': 'No data available. Please run Phase 1 first.'}), 404

    try:
        xlsx_bytes = _df_to_xlsx_bytes(df)
        resp = make_response(xlsx_bytes)
        resp.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        label = 'CT_Merged' if dtype == 'ct' else 'WB_Merged'
        resp.headers['Content-Disposition'] = f'attachment; filename="{label}.xlsx"'
        return resp
    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  PHASE 2 — Cycle Time Validation, Anomaly Handling & Online Data Preparation
# ─────────────────────────────────────────────────────────────────────────────

# In-memory store for Phase 2 outputs
_phase2_store = {
    'online':  None,   # clean Online Data df (after anomaly removal + formulas)
    'anomaly': None,   # Anomaly & Not Considered df
    'total_fail': 0,   # total "fail" count
}


# ── Pure-Python Failure Counter ───────────────────────────────────────────────

def _compute_failure_series(exceeded_series):
    """
    Correct consecutive-5 failure logic:
      - Maintain a running counter of consecutive Exceeded = 1
      - When counter reaches 5: mark this row 'fail', reset counter to 0
      - If Exceeded != 1: reset counter to 0, no fail
    Returns a list of strings: 'fail' or ''
    """
    results = []
    counter = 0
    for val in exceeded_series:
        try:
            v = int(float(val)) if pd.notna(val) else 0
        except (ValueError, TypeError):
            v = 0

        if v == 1:
            counter += 1
            if counter == 5:
                results.append('fail')
                counter = 0          # reset after marking fail
            else:
                results.append('')
        else:
            counter = 0
            results.append('')
    return results


# ── Anomaly Detection ─────────────────────────────────────────────────────────

ANOMALY_QUEUE_MINUTES = 0    # WB In Time – Date Time Arrival
ANOMALY_LAG_MINUTES   = 0    # WB Out Time – Date Time Exit (spec: WBOut – DTExit)


def _detect_anomalies(df, duration_threshold=100,
                      queue_minutes=None, lag_minutes=None,
                      min_duration_threshold=5):
    """
    A row is an anomaly if ANY condition is True:
      Condition 1: WB In Time – Date Time Arrival > queue_minutes
      Condition 2: WB Out Time – Date Time Exit   > lag_minutes
      Condition 3: Duration column from uploaded Excel > duration_threshold (upper-limit)
      Condition 4 (lower-limit): Duration == 0  OR  Duration < min_duration_threshold (default 5 min)
                   Always applied — any duration below 5 minutes is always an anomaly by default.

    queue_minutes defaults to ANOMALY_QUEUE_MINUTES (0) if not supplied.
    lag_minutes   defaults to ANOMALY_LAG_MINUTES   (0) if not supplied.

    Returns boolean Series (True = anomaly).
    """
    _queue_min = queue_minutes if queue_minutes is not None else ANOMALY_QUEUE_MINUTES
    _lag_min   = lag_minutes   if lag_minutes   is not None else ANOMALY_LAG_MINUTES
    arr_col    = find_col(df.columns, 'datetime', 'arrival') or find_col(df.columns, 'date', 'arrival')
    exit_col   = find_col(df.columns, 'datetime', 'exit')   or find_col(df.columns, 'date', 'exit')
    wb_in_col  = find_col(df.columns, 'wb', 'in', 'time')
    wb_out_col = find_col(df.columns, 'wb', 'out', 'time')
    dur_col    = find_col(df.columns, 'duration')

    if not arr_col or not exit_col:
        raise ValueError("Cannot find 'Date Time Arrival' or 'Date Time Exit' columns.")

    for col in [arr_col, exit_col, wb_in_col, wb_out_col]:
        if col:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')

    # Cond1: pre-weighbridge queue too long
    # SKIPPED ENTIRELY when _queue_min == 0 (zero means "ignore this check")
    if _queue_min > 0 and wb_in_col:
        cond1 = (df[wb_in_col] - df[arr_col]).dt.total_seconds() / 60 > _queue_min
    else:
        cond1 = pd.Series(False, index=df.index)

    # Cond2: post-weighbridge lag too long
    # SKIPPED ENTIRELY when _lag_min == 0 (zero means "ignore this check")
    if _lag_min > 0 and wb_out_col:
        cond2 = (df[wb_out_col] - df[exit_col]).dt.total_seconds() / 60 > _lag_min
    else:
        cond2 = pd.Series(False, index=df.index)

    # Cond3: use Duration column from uploaded Excel directly — no datetime subtraction.
    # Falls back to computing from datetimes if no Duration column present.
    if dur_col:
        dur_vals = pd.to_numeric(df[dur_col], errors='coerce').fillna(0)
        cond3 = dur_vals > duration_threshold
    else:
        # Fallback: calculate from datetimes if no Duration column
        dur_calc = (df[exit_col] - df[arr_col]).dt.total_seconds() / 60
        cond3 = dur_calc > duration_threshold

    # Cond4 (lower-limit): Duration == 0  OR  Duration < min_duration_threshold (default 5 min)
    # This is ALWAYS applied regardless of other threshold settings.
    # Any duration below the minimum (5 min by default) is always an anomaly.
    if dur_col:
        dur_vals4 = pd.to_numeric(df[dur_col], errors='coerce').fillna(0)
        cond4 = (dur_vals4 == 0) | (dur_vals4 < min_duration_threshold)
    else:
        dur_calc4 = (df[exit_col] - df[arr_col]).dt.total_seconds() / 60
        cond4 = (dur_calc4 == 0) | (dur_calc4 < min_duration_threshold)

    return cond1 | cond2 | cond3 | cond4


# ── Chart axis text helper ───────────────────────────────────────────────────

def _axis_text_properties(font_size=1000):
    """
    Return a txPr object with the given font size (in 100ths of a point).
    font_size=1100 → 11pt, 1000 → 10pt, 900 → 9pt.
    Used to make axis labels readable in the Cycle Time chart.
    """
    try:
        from openpyxl.drawing.text import (
            RichTextProperties, Paragraph, ParagraphProperties,
            CharacterProperties, RichText
        )
        cp = CharacterProperties(sz=font_size)
        pp = Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)
        txPr = RichText(bodyPr=RichTextProperties(), p=[pp])
        return txPr
    except Exception:
        return None


# ── Shared CT Chart Data Preparation ─────────────────────────────────────────
# This function is the SINGLE SOURCE OF TRUTH for Cycle Time chart data.
# It is called by both:
#   • _write_phase2_xlsx  → online Excel "Cycle Time Existence Chart"
#   • report_engine._render_chart_image  → PPT Slide 1 chart image
# Using the same function guarantees the PPT chart is an exact replica of
# the online chart in every respect: sort order, rolling mean, peak hours.

def _get_ct_chart_data(df_ct):
    """
    Prepare Cycle Time chart data using the EXACT same logic as the online
    'Cycle Time Existence Chart' in _write_phase2_xlsx.

    Column detection (mirrors _write_phase2_xlsx exactly):
      arr_col = find_col('datetime','arrival') OR find_col('date','arrival')
      dur_col = find_col('duration')
      sort    = find_col('date','time','arrival') OR find_col('wb','in','time')

    Steps (mirrors _write_phase2_xlsx exactly):
      1. Sort by Date Time Arrival ascending  (_arr_col_srt logic)
      2. Avg Cycle Time = rolling 10-row mean of Duration
         (Excel: =AVERAGE(DUR[r-9]:DUR[r]), NaN for rows 1-9)
      3. Peak Calculation = hour*100 + minute from Date Time Arrival
         (Excel: =VALUE(TEXT(ARR,"hhmm")))
      4. Peak Hours = 45 if 09:00<=t<11:30 OR 15:00<=t<17:30, else 30
         (Excel: =IF(OR(AND(pcalc>=900,pcalc<1130),AND(pcalc>=1500,pcalc<1730)),45,30))

    Returns dict with 'x_data', 'avg_ct', 'peak_hours', 'dur_max', 'arr_col', 'df'
    or None if required columns are missing / data is empty.
    """
    # Column detection — SAME as _write_phase2_xlsx
    def _lfc(*kw):
        return find_col(df_ct.columns, *kw)

    arr_col = _lfc('datetime', 'arrival') or _lfc('date', 'arrival')
    dur_col = _lfc('duration')

    if not dur_col:
        return None  # Cannot build chart without Duration column

    # Sort column — SAME as _arr_col_srt in _write_phase2_xlsx
    sort_col = (_lfc('date', 'time', 'arrival') or
                _lfc('wb', 'in', 'time') or
                arr_col)

    if not sort_col:
        return None  # No time column available

    # arr_col for X-axis (fallback to sort_col)
    if not arr_col:
        arr_col = sort_col

    df = df_ct.copy()
    df[arr_col]  = pd.to_datetime(df[arr_col],  dayfirst=True, errors='coerce')
    df[dur_col]  = pd.to_numeric(df[dur_col],   errors='coerce')
    if sort_col != arr_col:
        df[sort_col] = pd.to_datetime(df[sort_col], dayfirst=True, errors='coerce')

    # Step 1: Sort by Date Time Arrival — SAME as _write_phase2_xlsx
    df = df.sort_values(sort_col, ascending=True, kind='stable').reset_index(drop=True)
    df = df.dropna(subset=[arr_col, dur_col]).reset_index(drop=True)

    if len(df) == 0:
        return None

    # Step 2: Avg Cycle Time — rolling 10-row mean
    # SAME as Excel: =AVERAGE(DUR_L{r-9}:DUR_L{r})  (NaN for rows 1-9 = blank in Excel)
    avg_ct = df[dur_col].rolling(10).mean()

    # Step 3: Peak Calculation — hour*100+minute from Date Time Arrival
    # SAME as Excel: =VALUE(TEXT(ARR_L{r},"hhmm"))
    pcalc = df[arr_col].dt.hour * 100 + df[arr_col].dt.minute

    # Step 4: Peak Hours threshold
    # SAME as Excel: =IF(OR(AND(pcalc>=900,pcalc<1130),AND(pcalc>=1500,pcalc<1730)),45,30)
    peak_hours = pcalc.apply(
        lambda p: 45 if (900 <= p < 1130) or (1500 <= p < 1730) else 30
    )

    dur_s = df[dur_col].dropna()
    return {
        'x_data':     df[arr_col].values,    # Date Time Arrival timestamps for X-axis
        'avg_ct':     avg_ct.values,          # rolling 10-mean (NaN for first 9 rows)
        'peak_hours': peak_hours.values,      # 45 or 30 per row
        'dur_max':    float(dur_s.max()) if len(dur_s) else 0,
        'arr_col':    arr_col,
        'df':         df,
    }


# ── Online Data Excel Writer ──────────────────────────────────────────────────

def _write_phase2_xlsx(df_online, df_anomaly, total_fail, failure_list=None, report_date=None, start_dt=None, end_dt=None):
    """
    Build xlsx with two sheets:
      1. Online Data  — white bg, Calibri 11, fixed column widths, chart
      2. Anomaly and Not Considered — white bg, Calibri 11, red bold heading, fixed widths
    """
    from datetime import datetime as _dt
    from openpyxl.chart import LineChart, Reference, Series

    # Dynamic chart title from filter date range
    # Format: "Cycle Time Existence Chart on [D1] and [D2] [MON] [YYYY]"
    # e.g. "Cycle Time Existence Chart on 14 and 15 APR 2026"
    if start_dt is not None and end_dt is not None:
        d1   = start_dt.strftime('%d').lstrip('0') or '0'   # day without leading zero
        d2   = end_dt.strftime('%d').lstrip('0') or '0'
        mon  = start_dt.strftime('%b').upper()   # JAN, FEB, MAR ...
        yr   = start_dt.strftime('%Y')
        chart_title = f"Cycle Time Existence Chart on {d1} and {d2} {mon} {yr}"
    elif report_date is not None:
        chart_title = f"Cycle Time Existence Chart on {report_date.strftime('%d %b %Y').upper()}"
    else:
        if report_date is None:
            report_date = _dt.now()
        chart_title = f"Cycle Time Existence Chart on {report_date.strftime('%d %b %Y').upper()}"

    wb_xls = Workbook()
    wb_xls.remove(wb_xls.active)

    # ════════════════════════════════════════════════════════════════════════
    # Shared style constants (Calibri 11, white background)
    # ════════════════════════════════════════════════════════════════════════
    FONT_NAME   = 'Calibri'
    FONT_SIZE   = 11
    WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')
    NO_FILL     = PatternFill(fill_type=None)   # truly no fill
    THIN_BORDER = Border(
        left=Side(style='thin',   color='000000'),
        right=Side(style='thin',  color='000000'),
        top=Side(style='thin',    color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    def _cell_style(cell, bold=False, color='000000', bg='FFFFFF',
                    h_align='left', v_align='center', wrap=False):
        cell.font      = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, color=color)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
        cell.border    = THIN_BORDER

    def _header_cell(cell, text, bg='FFFFFF'):
        """Header: Calibri 11 bold, white background, thin border."""
        cell.value = text
        _cell_style(cell, bold=True, bg=bg, h_align='center', wrap=True)

    def _data_cell(cell, value):
        """Data: Calibri 11, white bg, center-aligned."""
        cell.value = value
        _cell_style(cell, bold=False, bg='FFFFFF', h_align='center')

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1: Online Data
    # ════════════════════════════════════════════════════════════════════════
    ws = wb_xls.create_sheet('Online Data')
    ws.sheet_format.defaultRowHeight = 14.40
    ws.sheet_format.customHeight     = True

    OUT_COLS = [
        'Etoken', 'Veh No.', 'Material',
        'Date Time Arrival', 'Format Hour', 'Peak Calculation',
        'Date Time Exit', 'WB In Time', 'WB Out Time', 'Duration',
        'Avg Cycle Time', 'Peak Hours', 'Exceeded', 'Failure', 'Total Count'
    ]
    N = len(OUT_COLS)

    # ── Fixed column widths (Online Total sheet) — exact spec ────────────
    ONLINE_WIDTHS = {
        'Etoken':             30.78,   # A
        'Veh No.':            19.22,   # B
        'Material':           19.22,   # C
        'Date Time Arrival':  19.22,   # D
        'Format Hour':        15.56,   # E
        'Peak Calculation':   16.33,   # F
        'Date Time Exit':     19.33,   # G
        'WB In Time':         19.33,   # H
        'WB Out Time':        19.33,   # I
        'Duration':           19.22,   # J
        'Avg Cycle Time':     19.22,   # K
        'Peak Hours':         19.22,   # L
        'Exceeded':           19.22,   # M
        'Failure':            19.22,   # N
        'Total Count':        10.78,   # O
    }
    # Display header names (internal name → header text shown in the sheet)
    # Internal names stay unchanged so all .index() lookups continue to work.
    ONLINE_HEADERS = {
        'Etoken':       'EToken',
        'Veh No.':      'Veh No',
        'Peak Hours':   'Peak Hour',
        # All other columns: display name = internal name (already correct)
    }
    for ci, col_name in enumerate(OUT_COLS, 1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = ONLINE_WIDTHS.get(col_name, 19.22)
    ws.row_dimensions[1].height = 14.40

    # ── Map source df columns ────────────────────────────────────────────
    def _fc(*kw): return find_col(df_online.columns, *kw)

    arr_col    = _fc('datetime', 'arrival') or _fc('date', 'arrival')
    exit_col   = _fc('datetime', 'exit')   or _fc('date', 'exit')
    wb_in_col  = _fc('wb', 'in', 'time')
    wb_out_col = _fc('wb', 'out', 'time')
    dur_col    = _fc('duration')
    etoken_col = _fc('etoken') or _fc('token')
    veh_col    = _fc('veh')
    mat_col    = _fc('material')

    # ── Column index constants ───────────────────────────────────────────
    ARR_CI   = OUT_COLS.index('Date Time Arrival') + 1
    FMT_CI   = OUT_COLS.index('Format Hour') + 1
    PCALC_CI = OUT_COLS.index('Peak Calculation') + 1
    DUR_CI   = OUT_COLS.index('Duration') + 1
    AVG_CI   = OUT_COLS.index('Avg Cycle Time') + 1
    PHR_CI   = OUT_COLS.index('Peak Hours') + 1
    EXC_CI   = OUT_COLS.index('Exceeded') + 1
    FAIL_CI  = OUT_COLS.index('Failure') + 1
    TOT_CI   = OUT_COLS.index('Total Count') + 1

    ARR_L   = get_column_letter(ARR_CI)
    PCALC_L = get_column_letter(PCALC_CI)
    DUR_L   = get_column_letter(DUR_CI)
    AVG_L   = get_column_letter(AVG_CI)
    PHR_L   = get_column_letter(PHR_CI)
    EXC_L   = get_column_letter(EXC_CI)
    FAIL_L  = get_column_letter(FAIL_CI)

    # ── Sort df_online by Date Time Arrival (A→Z) BEFORE writing rows ────────
    # This ensures Formula Hour, Peak Calculation, Avg Cycle Time (rolling 10),
    # and the chart categories are all in correct chronological order.
    _arr_col_srt = find_col(df_online.columns, 'date', 'time', 'arrival')                    or find_col(df_online.columns, 'wb', 'in', 'time')
    if _arr_col_srt and _arr_col_srt in df_online.columns:
        df_online = df_online.sort_values(
            _arr_col_srt, ascending=True, kind='stable'
        ).reset_index(drop=True)
        # failure_list must follow same order — re-derive if not externally supplied
        if failure_list is None:
            exc_col_pre    = find_col(df_online.columns, 'exceeded')
            exc_series_pre = df_online[exc_col_pre] if exc_col_pre else pd.Series(dtype=float)
            failure_list   = list(_compute_failure_series(exc_series_pre))

    # ── Header row (row 1) ────────────────────────────────────────────────
    for ci, name in enumerate(OUT_COLS, 1):
        # Column O (Total Count): blank header per spec
        # Other columns: use ONLINE_HEADERS display name if defined, else internal name
        if name == 'Total Count':
            display_name = ''
        else:
            display_name = ONLINE_HEADERS.get(name, name)
        _header_cell(ws.cell(row=1, column=ci), display_name)

    # ── Failure precompute ────────────────────────────────────────────────
    if failure_list is not None:
        failure_values = failure_list
    else:
        exc_col_fb    = find_col(df_online.columns, 'exceeded')
        exc_series_fb = df_online[exc_col_fb] if exc_col_fb else pd.Series(dtype=float)
        failure_values = _compute_failure_series(exc_series_fb)

    # ── Data rows ────────────────────────────────────────────────────────
    def g(col, row):
        if col and col in row.index:
            v = row[col]
            if isinstance(v, pd.Timestamp):
                return v.strftime('%d-%m-%Y %H:%M')
            return '' if (not isinstance(v, str) and pd.isna(v)) else v
        return ''

    for ri, (_, row) in enumerate(df_online.iterrows(), 2):
        dr = ri - 1
        ws.row_dimensions[ri].height = 14.40

        for ci, col_name in enumerate(OUT_COLS, 1):
            cell = ws.cell(row=ri, column=ci)
            _data_cell(cell, '')  # apply base style first

            if col_name == 'Etoken':
                cell.value = g(etoken_col, row)
            elif col_name == 'Veh No.':
                cell.value = g(veh_col, row)
            elif col_name == 'Material':
                cell.value = g(mat_col, row)
            elif col_name == 'Date Time Arrival':
                cell.value = g(arr_col, row)
            elif col_name == 'Format Hour':
                cell.value = (
                    f'=TEXT({ARR_L}{ri},"hh")' if ri == 2 else
                    f'=IF(TEXT({ARR_L}{ri},"hh")=TEXT({ARR_L}{ri-1},"hh"),"",TEXT({ARR_L}{ri},"hh"))'
                )
            elif col_name == 'Peak Calculation':
                cell.value = f'=VALUE(TEXT({ARR_L}{ri},"hhmm"))'
            elif col_name == 'Date Time Exit':
                cell.value = g(exit_col, row)
            elif col_name == 'WB In Time':
                cell.value = g(wb_in_col, row)
            elif col_name == 'WB Out Time':
                cell.value = g(wb_out_col, row)
            elif col_name == 'Duration':
                cell.value = g(dur_col, row)
            elif col_name == 'Avg Cycle Time':
                cell.value = '' if dr < 10 else f'=AVERAGE({DUR_L}{ri-9}:{DUR_L}{ri})'
            elif col_name == 'Peak Hours':
                cell.value = (
                    f'=IF(OR(AND({PCALC_L}{ri}>=900,{PCALC_L}{ri}<1130),'
                    f'AND({PCALC_L}{ri}>=1500,{PCALC_L}{ri}<1730)),45,30)'
                )
            elif col_name == 'Exceeded':
                cell.value = (
                    '' if dr < 10 else
                    f'=IF({AVG_L}{ri}="","",IF({PHR_L}{ri}>={AVG_L}{ri},0,1))'
                )
            elif col_name == 'Failure':
                fv = failure_values[dr - 1]
                cell.value = fv
                # Failure text is black (same as all other data), no special colour
            elif col_name == 'Total Count':
                cell.value = f'=COUNTIF({FAIL_L}:{FAIL_L},"fail")' if ri == 2 else ''

    ws.freeze_panes = None

    # ════════════════════════════════════════════════════════════════════════
    # Chart: Cycle Time Exceedance Chart
    # Exact match to reference file (Cycle_Time-_20260421_22042026.xlsx):
    #   X-axis categories: col E (format hour)  — 'Online Total'!$E$2:$En
    #   Series 1:  col K (Avg Cycle Time)        — 'Online Total'!$K$2:$Kn
    #   Series 2:  col L (Peak Hours)            — 'Online Total'!$L$2:$Ln
    #   No tick skip — Excel auto-reduces labels naturally
    #   Y-axis: 0-110, minor unit 2
    #   X-axis: rotated labels (-60 deg), auto=1
    # ════════════════════════════════════════════════════════════════════════
    try:
        import pandas as _pd

        n_data_rows = len(df_online)
        data_start  = 2
        data_end    = data_start + n_data_rows - 1

        # ── Build chart title ─────────────────────────────────────────────
        try:
            _dt_col = find_col(df_online.columns, 'wb', 'in', 'time')
            if not _dt_col:
                _dt_col = find_col(df_online.columns, 'arrival')
            if _dt_col and _dt_col in df_online.columns:
                _ts = _pd.to_datetime(df_online[_dt_col], dayfirst=True,
                                      errors='coerce').dropna()
            else:
                _ts = _pd.Series(dtype='datetime64[ns]')
            if len(_ts) > 0:
                _d1  = _ts.min().day
                _d2  = _ts.max().day
                _mon = _ts.min().strftime('%b').upper()
                _yr  = _ts.min().year
                _title = (f"Cycle Time Existence Chart on {_d1} {_mon} {_yr}"
                          if _d1 == _d2 else
                          f"Cycle Time Existence Chart on {_d1} and {_d2} {_mon} {_yr}")
            else:
                _title = chart_title
        except Exception:
            _title = chart_title

        # ── Line chart ────────────────────────────────────────────────────────
        chart = LineChart()
        chart.type   = "line"
        # Set chart title as non-bold RichText (size 14, regular weight).
        # Title.tx expects a Text object whose .rich holds the RichText content.
        try:
            from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                                ParagraphProperties, CharacterProperties,
                                                RegularTextRun)
            from openpyxl.chart.text import RichText, Text
            from openpyxl.chart.title import Title
            from openpyxl.chart.layout import Layout, ManualLayout
            _title_body = RichTextProperties(spcFirstLastPara=True, anchor='ctr', anchorCtr=True)
            _title_rpr  = CharacterProperties(sz=1400, b=False, i=False, baseline=0)
            _title_run  = RegularTextRun(t=_title, rPr=_title_rpr)
            _title_para = Paragraph(pPr=ParagraphProperties(defRPr=_title_rpr), r=[_title_run])
            _rich       = RichText(bodyPr=_title_body, p=[_title_para])
            _title_tx   = Text(rich=_rich)   # Title.tx must be Text, not RichText directly
            _title_layout = Layout(manualLayout=ManualLayout(y=0.06, yMode='factor'))
            _title_obj  = Title(tx=_title_tx, overlay=False, layout=_title_layout)
            chart.title = _title_obj
        except Exception:
            chart.title = _title
        chart.style  = 10
        chart.width  = 30.79   # fixed outer container width  (spec: 30.79 cm)
        chart.height = 17.65   # fixed outer container height (spec: 17.65 cm)
        chart.legend.position = "b"
        chart.legend.overlay  = False  # ensure legend does NOT overlap the plot area

        # Series 1: Avg Cycle Time — blue line
        avg_ref = Reference(ws, min_col=AVG_CI, max_col=AVG_CI,
                            min_row=1, max_row=data_end)
        avg_s   = Series(avg_ref, title_from_data=True)
        avg_s.graphicalProperties.line.solidFill = "5B9BD5"
        avg_s.graphicalProperties.line.width     = 28575
        avg_s.smooth = False
        chart.series.append(avg_s)

        # Series 2: Peak Hours — yellow line
        phr_ref = Reference(ws, min_col=PHR_CI, max_col=PHR_CI,
                            min_row=1, max_row=data_end)
        phr_s   = Series(phr_ref, title_from_data=True)
        phr_s.graphicalProperties.line.solidFill = "FFFF00"
        phr_s.graphicalProperties.line.width     = 28575
        phr_s.smooth = False
        chart.series.append(phr_s)

        # X-axis categories: col E (format hour strings)
        cats = Reference(ws, min_col=FMT_CI, max_col=FMT_CI,
                         min_row=data_start, max_row=data_end)
        chart.set_categories(cats)

        # ── Y-axis: AFTER series append so openpyxl serialises correctly ────
        # Rules (per spec):
        #   • y_max  = ceil(actual_max / 10) * 10, minimum 70
        #   • interval = 10 when that gives ≤ 10 gridlines (data ≤ 100 min)
        #   • interval = 20 otherwise, bumping y_max to next multiple of 20 if needed
        #   • always at least 7 gridlines; labels are always clean multiples of 10
        import math as _math
        try:
            _dur_col = find_col(df_online.columns, 'duration')
            _dur_s   = pd.to_numeric(df_online[_dur_col], errors='coerce').dropna()
            _data_max = float(_dur_s.max()) if len(_dur_s) else 0
        except Exception:
            _data_max = 0
        _y_axis_max = max(70, int(_math.ceil(_data_max / 10.0)) * 10)
        if _y_axis_max // 10 <= 10:                    # interval 10, ≤ 10 lines
            _y_major_unit = 10
        else:                                           # interval 20, bump if needed
            _y_major_unit = 20
            if _y_axis_max % 20 != 0:
                _y_axis_max += 10
            while _y_axis_max // 20 < 7:
                _y_axis_max += 20

        chart.y_axis.delete        = False
        chart.y_axis.numFmt        = 'General'
        chart.y_axis.scaling.min   = 0.0
        chart.y_axis.scaling.max   = float(_y_axis_max)
        chart.y_axis.majorUnit     = float(_y_major_unit)
        chart.y_axis.minorUnit     = float(_y_major_unit // 2)
        chart.y_axis.majorTickMark = 'none'
        chart.y_axis.minorTickMark = 'none'

        # ── X-axis: AFTER series append ───────────────────────────────────────
        # tickLblSkip=1 shows all category labels; Format Hour col already has
        # empty strings for repeated hours so only one label appears per hour.
        chart.x_axis.delete         = False
        chart.x_axis.numFmt         = 'General'
        chart.x_axis.majorTickMark  = 'none'
        chart.x_axis.minorTickMark  = 'none'
        chart.x_axis.lblOffset      = 100
        chart.x_axis.noMultiLvlLbl  = 1
        chart.x_axis.tickLblSkip    = 1
        chart.x_axis.tickMarkSkip   = 1

        from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                            ParagraphProperties, CharacterProperties)
        from openpyxl.chart.text import RichText
        try:
            _body = RichTextProperties(rot=-2700000, spcFirstLastPara=True,
                                       vertOverflow='ellipsis', vert='horz',
                                       wrap='square', anchor='ctr', anchorCtr=True)
            _rpr  = CharacterProperties(sz=900, b=False, i=False, baseline=0)
            _para = Paragraph(pPr=ParagraphProperties(defRPr=_rpr))
            _txPr = RichText(bodyPr=_body, p=[_para])
            chart.x_axis.txPr = _txPr
        except Exception:
            chart.x_axis.txPr = _axis_text_properties(font_size=900)

        # Place chart
        ws.add_chart(chart, f"{get_column_letter(N + 2)}2")

    except Exception as chart_err:
        import sys as _sys
        print(f"[WARN] Chart skipped: {chart_err}", file=_sys.stderr)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2: Anomaly and Not Considered
    # ════════════════════════════════════════════════════════════════════════
    ws_an = wb_xls.create_sheet('Anomaly and Not Considered')
    ws_an.sheet_format.defaultRowHeight = 14.40
    ws_an.sheet_format.customHeight     = True

    # FIX 4: Strip computed/internal columns from anomaly sheet (no extra calculated values)
    an_df = drop_unnamed(df_anomaly.copy())
    # Remove ALL internal computed columns (those starting with _ or known derived names)
    computed_drops = [col for col in an_df.columns if col.startswith('_')]
    computed_drops += [col for col in an_df.columns if col in 
                       ('format hour', 'Format Hour', 'Peak Calculation', 
                        'peak calculation', 'Total Count', 'totalcount')]
    an_df = an_df.drop(columns=[c for c in computed_drops if c in an_df.columns], errors='ignore')

    # Exact column order
    ANOMALY_COL_ORDER = [
        'Etoken', 'Veh No.', 'Material',
        'Date Time Arrival', 'Date Time Exit',
        'WB In Time', 'WB Out Time', 'Duration',
        'Avg Cycle Time', 'Peak Hours', 'Exceeded', 'Failure',
    ]
    an_available = list(an_df.columns)
    ordered_cols = []
    for want in ANOMALY_COL_ORDER:
        match = find_col(an_available, *want.lower().split())
        if not match:
            match = next((c for c in an_available
                          if c.strip().lower() == want.lower()), None)
        if match and match not in ordered_cols:
            ordered_cols.append(match)
    for c in an_available:
        if c not in ordered_cols:
            ordered_cols.append(c)

    an_out_cols = ['S/No.'] + ordered_cols + ['Remark']
    NA = len(an_out_cols)

    # ── Fixed column widths (Anomaly sheet) ──────────────────────────────
    ANOMALY_WIDTHS = {
        'S/No.':              8.11,
        'Etoken':            31.56,
        'Veh No.':           16.47,
        'Material':          17.56,
        'Date Time Arrival': 19.22,
        'Date Time Exit':    19.22,
        'WB In Time':        19.22,
        'WB Out Time':       19.22,
        'Duration':          15.67,
        'Avg Cycle Time':    11.78,
        'Peak Hours':         9.22,
        'Exceeded':           8.33,
        'Failure':           11.33,
        'Remark':            67.47,
    }
    for ci, col_name in enumerate(an_out_cols, 1):
        letter = get_column_letter(ci)
        ws_an.column_dimensions[letter].width = ANOMALY_WIDTHS.get(col_name, 19.22)

    # ── Row 1: Main heading — RED, BOLD, SIZE 26, white background ───────
    ws_an.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NA)
    heading_cell = ws_an.cell(row=1, column=1)
    heading_cell.value     = 'ANOMALY ADJUSTMENT AND THOSE NOT CONSIDERED'
    heading_cell.font      = Font(name=FONT_NAME, size=26, bold=True, color='FF0000')
    heading_cell.fill      = PatternFill('solid', fgColor='FFFFFF')
    heading_cell.alignment = Alignment(horizontal='left', vertical='center')
    heading_cell.border    = THIN_BORDER
    ws_an.row_dimensions[1].height = 33.60

    # ── Row 2: Column headers — Calibri 11 bold, white bg ────────────────
    for ci, name in enumerate(an_out_cols, 1):
        _header_cell(ws_an.cell(row=2, column=ci), name)
    ws_an.row_dimensions[2].height = 14.40

    # ── Data rows ─────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(an_df.iterrows(), 3):
        ws_an.row_dimensions[ri].height = 14.40
        # S/No.
        sno_cell = ws_an.cell(row=ri, column=1)
        _data_cell(sno_cell, ri - 2)
        sno_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Data columns
        for ci, col in enumerate(ordered_cols, 2):
            cell = ws_an.cell(row=ri, column=ci)
            v = row[col] if col in row.index else ''
            if isinstance(v, pd.Timestamp):
                v = v.strftime('%d-%m-%Y %H:%M')
            elif not isinstance(v, str) and pd.isna(v):
                v = ''
            _data_cell(cell, v)
        # Remark (blank)
        _data_cell(ws_an.cell(row=ri, column=NA), '')

    buf = io.BytesIO()
    wb_xls.save(buf)
    buf.seek(0)
    return buf.read()



# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/phase2')
def phase2_ui():
    return app.send_static_file('phase2.html')


@app.route('/phase2/process', methods=['POST', 'OPTIONS'])
def phase2_process():
    """
    Phase 2 endpoint:
      Input: CT data from Phase 1 store (already merged + filtered by WB In Time)
      Steps:
        1. Sort by Date Time Arrival (ascending)
        2. Detect anomalies (Cond1 + Cond2)
        3. Split: anomaly rows → Anomaly sheet; clean rows → Online Data
        4. Strip any pre-existing computed cols from clean data
        5. Compute Failure values (Python-side consecutive-5 logic)
        6. Write Online Data + Anomaly sheets to xlsx
        7. Return metadata + download
    """
    global _phase1_store, _phase2_store

    try:
        # Allow re-upload of CT data directly OR use Phase 1 store
        ct_file    = request.files.get('ct_file')
        ct_sheet   = request.form.get('ct_sheet', '0')
        use_stored = request.form.get('use_stored', 'true').lower() == 'true'

        if ct_file:
            # User uploaded fresh CT data for Phase 2
            df_ct = pd.read_excel(ct_file, sheet_name=_sid(ct_sheet))
            df_ct = drop_unnamed(df_ct)
        elif use_stored and _phase1_store.get('ct') is not None:
            # Use data from Phase 1
            df_ct = _phase1_store['ct'].copy()
        else:
            return jsonify({'error': 'No CT data available. Run Phase 1 first or upload a file.'}), 400

        # ── Step 1: Sort by Date Time Arrival (ascending) ─────────────────
        arr_col = find_col(df_ct.columns, 'datetime', 'arrival') or find_col(df_ct.columns, 'date', 'arrival')
        if arr_col:
            df_ct[arr_col] = pd.to_datetime(df_ct[arr_col], dayfirst=True, errors='coerce')
            df_ct = df_ct.sort_values(arr_col, ascending=True, kind="stable").reset_index(drop=True)

        # ── Step 2: Detect anomalies (pass user-configured threshold) ────────
        anomaly_threshold = float(request.form.get('anomaly_threshold', 100))
        min_duration      = float(request.form.get('min_duration', 5))   # lower-limit default 5 min
        is_anomaly = _detect_anomalies(df_ct, duration_threshold=anomaly_threshold,
                                       min_duration_threshold=min_duration)

        # ── Step 3: Split ─────────────────────────────────────────────────
        df_anomaly = df_ct[is_anomaly].copy().reset_index(drop=True)
        df_online  = df_ct[~is_anomaly].copy().reset_index(drop=True)

        # ── Step 4: Compute Failure BEFORE stripping (Exceeded still present) ──
        # Must read Exceeded BEFORE drop_computed removes it
        exc_col_pre = find_col(df_online.columns, 'exceeded')
        exc_series  = df_online[exc_col_pre] if exc_col_pre else pd.Series(dtype=float)
        failure_list = _compute_failure_series(exc_series)
        total_fail   = failure_list.count('fail')

        # ── Step 5: Strip pre-existing computed cols (now safe to remove Exceeded) ──
        df_online = drop_computed(df_online, CT_COMPUTED)

        # ── Step 6: Store results ─────────────────────────────────────────
        _phase2_store['online']       = df_online
        _phase2_store['anomaly']      = df_anomaly
        _phase2_store['total_fail']   = total_fail
        _phase2_store['failure_list'] = failure_list

        # ── Step 7: Build xlsx (pass pre-computed failure values explicitly) ──
        # Retrieve the filter dates stored by Phase 1 for the chart title
        _p1_start = _phase1_store.get('start_dt')
        _p1_end   = _phase1_store.get('end_dt')
        xlsx_bytes = _write_phase2_xlsx(
            df_online, df_anomaly, total_fail, failure_list,
            report_date=datetime.now(),
            start_dt=_p1_start,
            end_dt=_p1_end,
        )
        _phase2_store['xlsx'] = xlsx_bytes

        # Build preview for UI
        def preview_df(df, max_rows=10):
            p = df.head(max_rows).copy()
            for c in p.columns:
                if pd.api.types.is_datetime64_any_dtype(p[c]):
                    p[c] = p[c].dt.strftime('%d-%m-%Y %H:%M')
            return {
                'columns':    list(p.columns),
                'rows':       p.where(pd.notnull(p), None).values.tolist(),
                'total_rows': len(df),
            }

        meta = {
            'online_rows':   len(df_online),
            'anomaly_rows':  len(df_anomaly),
            'total_rows':    len(df_ct),
            'total_fail':    total_fail,
            'anomaly_threshold': anomaly_threshold,
            'sorted_by':     arr_col or '—',
            'online_preview':  preview_df(df_online),
            'anomaly_preview': preview_df(df_anomaly),
        }

        resp = make_response(json.dumps({'ok': True}))
        resp.headers['Content-Type'] = 'application/json'
        resp.headers['X-Meta']       = json.dumps(meta)
        return resp

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@app.route('/phase2/download')
def phase2_download():
    """Download the Phase 2 output xlsx (Online Data + Anomaly sheets)."""
    global _phase2_store
    xlsx = _phase2_store.get('xlsx')
    if xlsx is None:
        return jsonify({'error': 'No Phase 2 output yet. Run Phase 2 first.'}), 404
    resp = make_response(xlsx)
    resp.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = 'attachment; filename="Phase2_CT_Report.xlsx"'
    return resp


# ═════════════════════════════════════════════════════════════════════════════
#  WB PHASE — Weighbridge Data Processing
#  Sections: validation, rejected loads, WB Total sheet with all formatting
# ═════════════════════════════════════════════════════════════════════════════

# In-memory store for WB phase
_wb_store = {
    'total':    None,   # df — accepted rows ready for output
    'rejected': None,   # df — rows with ACCEPTED=No + blank WBOut reviewed as No
    'pending':  None,   # df — rows with blank WB Out Time awaiting user review
    'xlsx':     None,   # bytes — last generated workbook
    'flagged_tokens': [],  # list of {token, in_weight, row_idx} for UI popup
}

# ── WB Style constants ────────────────────────────────────────────────────────
WB_FONT_NAME   = 'Calibri'
WB_FONT_SIZE   = 11
WB_HDR_BG      = '0070C0'   # Standard Blue (Excel Standard Colors)
WB_HDR_FG      = 'FFFFFF'   # White text
WB_HDR_HEIGHT  = 25.20
WB_ROW_HEIGHT  = 14.40
WB_BLACK_BDR   = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)

# WB Total fixed column widths
WB_TOTAL_WIDTHS = {
    'TOKEN':              29.22,
    'SITE CODE':          20.22,
    'DATETIME ARRIVAL':   19.89,
    'DATETIME EXIT':      24.22,
    'VEHICLE NO':         24.22,
    'MATERIAL':           24.22,
    'ACCEPTED':           24.22,
    'REJECT REASON':      24.22,
    'TRAN STATUS':        24.22,
    'WB IN TIME':         24.22,
    'Format Hour':        24.22,
    'Peak Calculation':   24.22,
    'IN WEIGHT':          24.22,
    'WB OUT TIME':        24.22,
    'OUT WEIGHT':         24.22,
    'NET WEIGHT':         24.22,
    'HAULAGE CONTRACTOR': 24.22,
    'IN LANE':            24.22,
    'OUT LANE':           24.22,
    'STAGING GROUND':     24.22,
    'RE-CLASSIFIED':      24.22,
}

# Rejected sheet — same but Reject Reason wider
WB_REJ_WIDTHS = {**WB_TOTAL_WIDTHS, 'REJECT REASON': 34.56}


def _wb_cell(cell, value, bold=False, font_color='000000', bg='FFFFFF',
             h_align='center', v_align='center', wrap=False, red_text=False):
    """Apply standard WB cell formatting."""
    fc = 'FF0000' if red_text else font_color
    cell.value     = value
    cell.font      = Font(name=WB_FONT_NAME, size=WB_FONT_SIZE,
                          bold=bold, color=fc)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align,
                               wrap_text=wrap)
    cell.border    = WB_BLACK_BDR


def _wb_header_row(ws, row, col_names):
    """Write a header row with light-blue bg, white bold text."""
    for ci, name in enumerate(col_names, 1):
        cell = ws.cell(row=row, column=ci)
        _wb_cell(cell, name, bold=True,
                 font_color=WB_HDR_FG, bg=WB_HDR_BG,
                 h_align='center')
    ws.row_dimensions[row].height = WB_HDR_HEIGHT


def _fmt_time(v):
    """Return value unchanged — no reformatting, no AM/PM conversion."""
    if isinstance(v, pd.Timestamp):
        return v.strftime('%d-%m-%Y %H:%M:%S') if v.second else v.strftime('%d-%m-%Y %H:%M')
    if isinstance(v, str) and v.strip():
        return v.strip()
    return ''


def _fmt_weight(v):
    """Format weight — no unnecessary decimals, 0 shows as 0."""
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return v if v not in (None, '') else ''


def _wb_set_col_widths(ws, col_names, width_map):
    for ci, name in enumerate(col_names, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width_map.get(name, 24.22)


# ── Core WB processing ────────────────────────────────────────────────────────

def process_wb_phase(df1, df2, start_dt, end_dt):
    """
    1. Merge file1 + file2
    2. Filter by WB IN TIME within [start_dt, end_dt]
    3. Split by ACCEPTED column:
         ACCEPTED = No  → df_rejected (full rows)
         blank WB OUT TIME → df_pending (needs user review popup)
         rest → df_valid
    Returns (df_valid, df_rejected, df_pending, flagged_tokens)
    """
    # Merge
    df1 = drop_unnamed(df1.copy())
    dfs = [df1]
    if df2 is not None:
        df2 = drop_unnamed(df2.copy())
        for c in df1.columns:
            if c not in df2.columns:
                df2[c] = pd.NA
        df2 = df2[df1.columns]
        dfs.append(df2)
    df = pd.concat(dfs, ignore_index=True)

    # Drop pre-existing Format Hour / Peak Calculation
    df = drop_computed(df, WB_COMPUTED)

    # Locate WB IN TIME column
    wb_in_col = find_col(df.columns, 'wb', 'in', 'time')
    if not wb_in_col:
        raise ValueError(f"'WB IN TIME' not found. Columns: {list(df.columns)}")

    # Parse & filter
    df[wb_in_col] = pd.to_datetime(df[wb_in_col], dayfirst=True, errors='coerce')
    df = df[(df[wb_in_col] >= start_dt) & (df[wb_in_col] <= end_dt)].reset_index(drop=True)

    # Locate key columns
    acc_col    = find_col(df.columns, 'accepted')
    wb_out_col = find_col(df.columns, 'wb', 'out', 'time')
    token_col  = find_col(df.columns, 'token')
    in_wt_col  = find_col(df.columns, 'in', 'weight')

    # Step 1: Split ACCEPTED = No → rejected
    if acc_col:
        is_rej = df[acc_col].astype(str).str.strip().str.upper() == 'NO'
        df_rejected = df[is_rej].copy().reset_index(drop=True)
        df_rest     = df[~is_rej].copy().reset_index(drop=True)
    else:
        df_rejected = pd.DataFrame(columns=df.columns)
        df_rest     = df.copy()

    # Step 2: From remaining, find rows with blank WB OUT TIME → pending
    if wb_out_col:
        df_rest[wb_out_col] = pd.to_datetime(
            df_rest[wb_out_col], dayfirst=True, errors='coerce')
        is_blank_out = df_rest[wb_out_col].isna()
        df_pending = df_rest[is_blank_out].copy().reset_index(drop=True)
        df_valid   = df_rest[~is_blank_out].copy().reset_index(drop=True)
    else:
        df_pending = pd.DataFrame(columns=df_rest.columns)
        df_valid   = df_rest.copy()

    # Build flagged tokens list for UI popup
    flagged = []
    if token_col and in_wt_col:
        for _, row in df_pending.iterrows():
            flagged.append({
                'token':     str(row.get(token_col, '')),
                'in_weight': _fmt_weight(row.get(in_wt_col, '')),
            })
    elif token_col:
        for _, row in df_pending.iterrows():
            flagged.append({'token': str(row.get(token_col, '')), 'in_weight': ''})

    return df_valid, df_rejected, df_pending, flagged


def apply_wb_pending_decisions(df_valid, df_rejected, df_pending, decisions):
    """
    Apply user decisions for each pending (blank WB Out Time) row.
    decisions: list of dicts {token, choice: 'yes'|'no', out_weight: float|None}

    'no'  → mark rejected (OutWeight = InWeight, NetWeight = 0, red font flag)
    'yes' → add to valid with user-entered OutWeight, yellow highlight
    """
    dec_map = {d['token']: d for d in decisions}

    in_wt_col  = find_col(df_pending.columns, 'in', 'weight')
    out_wt_col = find_col(df_pending.columns, 'out', 'weight')
    net_wt_col = find_col(df_pending.columns, 'net', 'weight')
    token_col  = find_col(df_pending.columns, 'token')

    yes_rows, no_rows = [], []

    for _, row in df_pending.iterrows():
        token = str(row.get(token_col, '')) if token_col else ''
        dec   = dec_map.get(token, {})
        choice = dec.get('choice', 'no').lower()
        row   = row.copy()

        in_wt = float(row[in_wt_col]) if (in_wt_col and pd.notna(row.get(in_wt_col))) else 0.0

        if choice == 'yes':
            entered_out = dec.get('out_weight')
            try:
                out_wt = float(entered_out) if entered_out not in (None, '') else in_wt
            except (TypeError, ValueError):
                out_wt = in_wt
            net_wt = in_wt - out_wt

            if out_wt_col:
                row[out_wt_col] = _fmt_weight(out_wt)
            if net_wt_col:
                row[net_wt_col] = _fmt_weight(net_wt)
            row['_yellow'] = True
            row['_red_cols'] = [out_wt_col, net_wt_col]
            yes_rows.append(row)
        else:
            # No → rejected: OutWeight = InWeight, NetWeight = 0
            if out_wt_col:
                row[out_wt_col] = _fmt_weight(in_wt)
            if net_wt_col:
                row[net_wt_col] = 0
            row['_red_rejected'] = True
            no_rows.append(row)

    # Merge yes rows into df_valid
    if yes_rows:
        df_yes = pd.DataFrame(yes_rows)
        df_valid = pd.concat([df_valid, df_yes], ignore_index=True)

    # Merge no rows into df_rejected
    if no_rows:
        df_no = pd.DataFrame(no_rows)
        df_rejected = pd.concat([df_rejected, df_no], ignore_index=True)

    return df_valid, df_rejected


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_wb_phase_xlsx(df_valid, df_rejected):
    """
    Build WB workbook:
      Sheet 1: WB Total     — accepted + pending-yes rows, Format Hour + Peak Calc
      Sheet 2: WB Rejected Loads — rejected rows with heading
    """
    wb_xls = Workbook()
    wb_xls.remove(wb_xls.active)

    # ── Sheet 1: WB Total ─────────────────────────────────────────────────
    ws = wb_xls.create_sheet('WB Total')
    ws.sheet_format.defaultRowHeight = WB_ROW_HEIGHT
    ws.sheet_format.customHeight     = True

    # Locate WB IN TIME to insert Format Hour + Peak Calculation after it
    wb_in_col = find_col(df_valid.columns, 'wb', 'in', 'time')
    orig_cols = list(df_valid.columns)

    # Strip any internal marker cols
    orig_cols = [c for c in orig_cols if not c.startswith('_')]

    if wb_in_col and wb_in_col in orig_cols:
        wb_in_pos = orig_cols.index(wb_in_col)
        out_cols  = (orig_cols[:wb_in_pos + 1]
                     + ['Format Hour', 'Peak Calculation']
                     + orig_cols[wb_in_pos + 1:])
    else:
        out_cols = orig_cols + ['Format Hour', 'Peak Calculation']

    N_T = len(out_cols)

    # Column widths
    _wb_set_col_widths(ws, out_cols,
                       {**WB_TOTAL_WIDTHS, 'Format Hour': 24.22,
                        'Peak Calculation': 24.22})

    # Header row 1
    _wb_header_row(ws, 1, out_cols)

    # WB IN TIME letter for formulas
    wb_in_excel_ci = (out_cols.index(wb_in_col) + 1) if wb_in_col in out_cols else None
    wb_in_L = get_column_letter(wb_in_excel_ci) if wb_in_excel_ci else 'A'

    # Data rows
    for ri, (_, row) in enumerate(df_valid.iterrows(), 2):
        ws.row_dimensions[ri].height = WB_ROW_HEIGHT
        is_yellow  = bool(row.get('_yellow', False))
        red_cols   = row.get('_red_cols', []) or []
        row_bg     = 'FFFF00' if is_yellow else 'FFFFFF'

        for ci, col_name in enumerate(out_cols, 1):
            cell = ws.cell(row=ri, column=ci)

            if col_name == 'Format Hour':
                # =IF(TEXT(J2,"hh")=TEXT(J1,"hh"),"",TEXT(J2,"hh"))
                if ri == 2:
                    val = f'=TEXT({wb_in_L}{ri},"hh")'
                else:
                    val = (f'=IF(TEXT({wb_in_L}{ri},"hh")='
                           f'TEXT({wb_in_L}{ri-1},"hh"),"",TEXT({wb_in_L}{ri},"hh"))')
                _wb_cell(cell, val, bg=row_bg)

            elif col_name == 'Peak Calculation':
                val = f'=VALUE(TEXT({wb_in_L}{ri},"hhmm"))'
                _wb_cell(cell, val, bg=row_bg)

            else:
                if col_name in row.index:
                    raw = row[col_name]
                    # Time values passed through exactly — no reformatting
                    if not isinstance(raw, str) and pd.isna(raw):
                        display = ''
                    else:
                        display = raw

                    is_red = (col_name in red_cols)
                    _wb_cell(cell, display, bg=row_bg, red_text=is_red)
                else:
                    _wb_cell(cell, '', bg=row_bg)

    # ── Sheet 2: WB Rejected Loads ────────────────────────────────────────
    ws_r = wb_xls.create_sheet('WB Rejected Loads')
    ws_r.sheet_format.defaultRowHeight = WB_ROW_HEIGHT
    ws_r.sheet_format.customHeight     = True

    rej_cols = [c for c in df_rejected.columns if not c.startswith('_')]

    _wb_set_col_widths(ws_r, rej_cols, WB_REJ_WIDTHS)

    # Row 1: Big heading
    ws_r.merge_cells(start_row=1, start_column=1,
                     end_row=1, end_column=max(len(rej_cols), 1))
    hcell = ws_r.cell(row=1, column=1)
    hcell.value     = 'WEIGHBRIDGE REJECTED AND RECLASSIFIED LOADS'
    hcell.font      = Font(name='Arial', size=26, bold=True, color='FF0000')
    hcell.fill      = PatternFill('solid', fgColor='FFFFFF')
    hcell.alignment = Alignment(horizontal='left', vertical='center')
    hcell.border    = WB_BLACK_BDR
    ws_r.row_dimensions[1].height = WB_HDR_HEIGHT

    # Row 2: Column headers
    _wb_header_row(ws_r, 2, rej_cols)

    # Data rows
    for ri, (_, row) in enumerate(df_rejected.iterrows(), 3):
        ws_r.row_dimensions[ri].height = WB_ROW_HEIGHT
        is_red_rej  = bool(row.get('_red_rejected', False))
        red_cols_rj = row.get('_red_cols', []) or []

        for ci, col_name in enumerate(rej_cols, 1):
            cell = ws_r.cell(row=ri, column=ci)
            if col_name in row.index:
                raw = row[col_name]
                # Time values passed through exactly — no reformatting
                if not isinstance(raw, str) and pd.isna(raw):
                    display = ''
                else:
                    display = raw

                # Red font: if this col is in red_cols OR whole row is rejected-No
                is_red = (col_name in red_cols_rj) or (is_red_rej and col_name in
                          [find_col(rej_cols, 'out', 'weight'),
                           find_col(rej_cols, 'net', 'weight')])
                _wb_cell(cell, display, red_text=is_red)
            else:
                _wb_cell(cell, '')

    buf = io.BytesIO()
    wb_xls.save(buf)
    buf.seek(0)
    return buf.read()


# ── WB Phase Routes ────────────────────────────────────────────────────────────

@app.route('/wb')
def wb_ui():
    return app.send_static_file('wb.html')


@app.route('/wb/upload', methods=['POST', 'OPTIONS'])
def wb_upload():
    """
    Phase 1 of WB: upload, merge, filter, detect pending rows.
    Returns metadata + flagged tokens (blank WB Out Time) for UI popup.
    """
    global _wb_store
    try:
        file1 = request.files.get('wb_file1')
        file2 = request.files.get('wb_file2')
        sh1   = request.form.get('wb_sheet1', '0')
        sh2   = request.form.get('wb_sheet2', '0')
        start_str = request.form.get('start_dt', '')
        end_str   = request.form.get('end_dt',   '')

        if not file1:
            return jsonify({'error': 'File 1 is required.'}), 400
        if not start_str or not end_str:
            return jsonify({'error': 'Start and End date/time are required.'}), 400

        start_dt = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        end_dt   = datetime.strptime(end_str,   '%Y-%m-%dT%H:%M')
        if start_dt >= end_dt:
            return jsonify({'error': 'Start must be before End.'}), 400

        def sid(s):
            try: return int(s)
            except: return s or 0

        df1 = pd.read_excel(file1, sheet_name=sid(sh1))
        df2 = pd.read_excel(file2, sheet_name=sid(sh2)) if file2 else None

        df_valid, df_rejected, df_pending, flagged = process_wb_phase(
            df1, df2, start_dt, end_dt)

        _wb_store['total']           = df_valid
        _wb_store['rejected']        = df_rejected
        _wb_store['pending']         = df_pending
        _wb_store['flagged_tokens']  = flagged
        _wb_store['start_dt']        = start_dt
        _wb_store['end_dt']          = end_dt

        meta = {
            'valid_rows':    len(df_valid),
            'rejected_rows': len(df_rejected),
            'pending_rows':  len(df_pending),
            'flagged':       flagged,
        }
        return jsonify(meta)

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@app.route('/wb/resolve', methods=['POST', 'OPTIONS'])
def wb_resolve():
    """
    Apply user decisions for blank-WB-Out-Time rows, generate final xlsx.
    Body JSON: { decisions: [{token, choice, out_weight?}, ...] }
    """
    global _wb_store
    try:
        body      = request.get_json(force=True) or {}
        decisions = body.get('decisions', [])

        df_valid   = _wb_store.get('total')
        df_rejected = _wb_store.get('rejected')
        df_pending  = _wb_store.get('pending')

        if df_valid is None:
            return jsonify({'error': 'No WB data. Run /wb/upload first.'}), 400

        # Apply decisions
        if df_pending is not None and len(df_pending) > 0 and decisions:
            df_valid, df_rejected = apply_wb_pending_decisions(
                df_valid, df_rejected, df_pending, decisions)

        _wb_store['total']    = df_valid
        _wb_store['rejected'] = df_rejected

        # Generate xlsx
        xlsx_bytes = write_wb_phase_xlsx(df_valid, df_rejected)
        _wb_store['xlsx'] = xlsx_bytes

        return jsonify({
            'ok': True,
            'total_rows':    len(df_valid),
            'rejected_rows': len(df_rejected),
        })

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@app.route('/wb/download')
def wb_download():
    global _wb_store
    xlsx = _wb_store.get('xlsx')
    if not xlsx:
        return jsonify({'error': 'No WB output yet.'}), 404
    resp = make_response(xlsx)
    resp.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = 'attachment; filename="WB_Report.xlsx"'
    return resp


# ═════════════════════════════════════════════════════════════════════════════
#  SHEETS 5–7: Pivot Table  ·  Sorting  ·  Charts
#  All three are derived from df_online (Online Total) and df_wb (WB Total)
# ═════════════════════════════════════════════════════════════════════════════

# ── Constants ─────────────────────────────────────────────────────────────────

# Hour blocks: report runs 07:00 → 05:00 next day  (22 hours, wrap through midnight)
HOUR_SLOTS = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,0,1,2,3,4]

HOUR_LABELS = {           # Used in Charts / Sorting headers
    7:'7AM - 8AM',   8:'8AM - 9AM',   9:'9AM - 10AM',  10:'10AM - 11AM',
   11:'11AM - 12PM',12:'12PM - 1PM',  13:'1PM - 2PM',  14:'2PM - 3PM',
   15:'3PM - 4PM',  16:'4PM - 5PM',   17:'5PM - 6PM',  18:'6PM - 7PM',
   19:'7PM - 8PM',  20:'8PM - 9PM',   21:'9PM - 10PM', 22:'10PM - 11PM',
   23:'11PM - 12AM', 0:'12AM - 1AM',   1:'1AM - 2AM',   2:'2AM - 3AM',
    3:'3AM - 4AM',   4:'4AM - 5AM',
}

SORT_HOUR_LABELS = {      # Sorting sheet column group headers (HHMM-HHMM format)
    7:'0700-0759',  8:'0800-0859',  9:'0900-0959', 10:'1000-1059',
   11:'1100-1159', 12:'1200-1259', 13:'1300-1359', 14:'1400-1459',
   15:'1500-1559', 16:'1600-1659', 17:'1700-1759', 18:'1800-1859',
   19:'1900-1959', 20:'2000-2059', 21:'2100-2159', 22:'2200-2259',
   23:'2300-2359',  0:'0000-0059',  1:'0100-0159',  2:'0200-0259',
    3:'0300-0359',  4:'0400-0459',
}

# Provision defaults (hardcoded, manual adjustment after download if needed)
PROV_GE_MONTHLY = 200_000
PROV_SC_MONTHLY = 800_000
PROV_DAYS       = 29        # divisor matching demo (200000/29 = 6896.55..)
PROV_GE_DAILY   = PROV_GE_MONTHLY / PROV_DAYS
PROV_SC_DAILY   = PROV_SC_MONTHLY / PROV_DAYS
PROV_HOURS      = len(HOUR_SLOTS)   # 22 operating hours per day
PROV_GE_HOURLY  = PROV_GE_DAILY / PROV_HOURS
PROV_SC_HOURLY  = PROV_SC_DAILY / PROV_HOURS


# ── Shared cell style helpers (full-border, white bg) ─────────────────────────

def _xb():
    """Black thin border all sides."""
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def _xc(cell, value, bold=False, size=10, color='000000', bg='FFFFFF',
        h='center', v='center', wrap=False, italic=False):
    """Write value + full style to cell."""
    cell.value     = value
    cell.font      = Font(name='Calibri', size=size, bold=bold,
                          color=color, italic=italic)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border    = _xb()

def _xmerge(ws, r1, c1, r2, c2, value, bold=False, size=10,
            color='000000', bg='FFFFFF', h='center', wrap=False):
    """Merge cells and write styled value."""
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    _xc(cell, value, bold=bold, size=size, color=color, bg=bg, h=h, wrap=wrap)
    # Apply border to all cells in the merged range for clean grid
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = _xb()

def _col_bg(ws, ci, bg, row_start, row_end):
    """Fill background colour on a column range without changing other style."""
    for r in range(row_start, row_end+1):
        ws.cell(r, ci).fill = PatternFill('solid', fgColor=bg)


# ── Data helpers ──────────────────────────────────────────────────────────────

def _prep_online(df_online):
    """Normalise Online Total df for aggregation."""
    df = df_online.copy()
    arr_col = find_col(df.columns, 'datetime', 'arrival') or find_col(df.columns, 'date', 'arrival')
    mat_col = find_col(df.columns, 'material')
    dur_col = find_col(df.columns, 'duration')
    fail_col = find_col(df.columns, 'failure')

    if arr_col:
        df[arr_col] = pd.to_datetime(df[arr_col], dayfirst=True, errors='coerce')
        df['_hour'] = df[arr_col].dt.hour
    else:
        df['_hour'] = np.nan

    df['_mat'] = (df[mat_col].astype(str).str.strip().str.upper()
                  if mat_col else '')
    df['_dur'] = (pd.to_numeric(df[dur_col], errors='coerce')
                  if dur_col else np.nan)
    df['_fail'] = (df[fail_col].fillna('').astype(str).str.strip().str.lower() == 'fail'
                   if fail_col else False)
    return df


def _prep_wb(df_wb):
    """Normalise WB Total df for aggregation."""
    df = df_wb.copy()
    wb_in = find_col(df.columns, 'wb', 'in', 'time')
    mat_col = find_col(df.columns, 'material')

    if wb_in:
        df[wb_in] = pd.to_datetime(df[wb_in], dayfirst=True, errors='coerce')
        df['_hour'] = df[wb_in].dt.hour
    else:
        df['_hour'] = np.nan

    df['_mat'] = (df[mat_col].astype(str).str.strip().str.upper()
                  if mat_col else '')
    return df


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Pivot Table
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot_sheet(wb_out, df_online=None, df_wb=None):
    """
    Builds the Pivot Table sheet with formulas copied exactly from the demo file,
    cell-by-cell, row-by-row. No computed values — all live Excel formulas
    referencing the Sorting sheet (which supplies the per-hour vehicle data).

    Layout (A1:S25, rows 1-2 headers, rows 3-24 data, row 25 empty):
      Col A  = Row Labels (hour numbers: 7,8,9...23,0,1,2,3,4)
      B-G    = Soft Clay: Online SC | WB SC | Min Dur | Max Dur | Avg Dur | Failures
      H-M    = Good Earth: same structure
      N-S    = Totals: Online Total | WB Total | Min | Max | Avg | Total Failures
    All data cells reference Sorting!<column> via COUNTIF/MIN/MAX/AVERAGE/SUM.
    """
    ws = wb_out.create_sheet('Pivot Table')

    from openpyxl.styles import Font, Alignment

    FONT_NAME = 'Calibri'
    FONT_SIZE = 11

    def _w(cell, value):
        """Write value and apply demo formatting: Calibri 11, centre-aligned, no fill."""
        cell.value = value
        cell.font      = Font(name=FONT_NAME, size=FONT_SIZE)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Column widths (exact from demo) ──────────────────────────────────────
    col_widths = {
        'A': 10.86, 'B': 15.00, 'C': 11.86, 'D': 14.86, 'E': 15.14,
        'F': 18.86, 'G': 21.71, 'H': 15.14, 'I': 12.00, 'J': 14.86,
        'K': 15.14, 'L': 18.86, 'M': 21.71, 'N': 28.29, 'O': 25.14,
        'P': 19.86, 'Q': 20.14, 'R': 23.86, 'S': 34.29,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: section span headers ──────────────────────────────────────────
    row1 = {
        'B': 'Soft Clay',
        'H': 'Good Earth',
        'N': 'Online Total Count of material',
        'O': 'WB Total Count of material',
        'P': 'Total Min of duration',
        'Q': 'Total Max of duration',
        'R': 'Total Average of duration',
        'S': 'Total Count of Format No. of Failures',
    }
    for col, val in row1.items():
        _w(ws[f'{col}1'], val)

    # ── Row 2: column headers ─────────────────────────────────────────────────
    row2 = {
        'A': 'Row Labels',
        'B': 'Online No of SC',
        'C': 'WB No of SC',
        'D': 'Min of duration',
        'E': 'Max of duration',
        'F': 'Average of duration',
        'G': 'Count of No of Failures',
        'H': 'Online No of GE',
        'I': 'WB No of GE',
        'J': 'Min of duration',
        'K': 'Max of duration',
        'L': 'Average of duration',
        'M': 'Count of No of Failures',
    }
    for col, val in row2.items():
        _w(ws[f'{col}2'], val)

    # ── Data rows 3-24: exact formulas from demo (one row per hour) ───────────
    # Each row maps to one hour slot and references specific Sorting sheet columns.
    # The Sorting sheet has 8 columns per hour block:
    #   block_start + 0 = Online GE
    #   block_start + 1 = WB GE
    #   block_start + 2 = GE Duration
    #   block_start + 3 = GE Failure
    #   block_start + 4 = Online SC
    #   block_start + 5 = WB SC
    #   block_start + 6 = SC Duration
    #   block_start + 7 = SC Failure
    # Pivot formulas reference full columns of the Sorting sheet.
    # Extracted verbatim from demo file (data_only=False):

    # (excel_row, hour_label, formulas_B_thru_S)
    # formulas are the exact strings from the demo, spaces preserved
    DATA_ROWS = [
        # Row 3 — Hour 7
        (3, 7,
         '=COUNTIF( Sorting!E:E,"SOFT CLAY")',
         '=COUNTIF( Sorting!F:F,"SOFT CLAY")',
         '=IF(B3=0,"",MIN( Sorting!G:G))',
         '=IF(B3=0,"",MAX( Sorting!G:G))',
         '=IF(B3=0,"",AVERAGE( Sorting!G:G))',
         '=COUNTIF( Sorting!H:H,"fail")',
         '=COUNTIF( Sorting!A:A,"GOOD EARTH")',
         '=COUNTIF( Sorting!B:B,"GOOD EARTH")',
         '=IF(H3=0,"",MIN( Sorting!C:C))',
         '=IF(H3=0,"",MAX( Sorting!C:C))',
         '=IF(H3=0,"",AVERAGE( Sorting!C:C))',
         '=COUNTIF( Sorting!D:D,"fail")',
         '=SUM(B3,H3)', '=SUM(C3,I3)',
         '=IF(N3=0,"",MIN(D3,J3))',
         '=IF(N3=0,"",MAX(E3,K3))',
         '=IF(N3=0,"",AVERAGE( Sorting!C:C, Sorting!G:G))',
         '=SUM(G3,M3)',
        ),
        # Row 4 — Hour 8
        (4, 8,
         '=COUNTIF( Sorting!M:M,"SOFT CLAY")',
         '=COUNTIF( Sorting!N:N,"SOFT CLAY")',
         '=IF(B4=0,"",MIN( Sorting!O:O))',
         '=IF(B4=0,"",MAX( Sorting!O:O))',
         '=IF(B4=0,"",AVERAGE( Sorting!O:O))',
         '=COUNTIF( Sorting!P:P,"fail")',
         '=COUNTIF( Sorting!I:I,"GOOD EARTH")',
         '=COUNTIF( Sorting!J:J,"GOOD EARTH")',
         '=IF(H4=0,"",MIN( Sorting!K:K))',
         '=IF(H4=0,"",MAX( Sorting!K:K))',
         '=IF(H4=0,"",AVERAGE( Sorting!K:K))',
         '=COUNTIF( Sorting!L:L,"fail")',
         '=SUM(B4,H4)', '=SUM(C4,I4)',
         '=IF(N4=0,"",MIN(D4,J4))',
         '=IF(N4=0,"",MAX(E4,K4))',
         '=IF(N4=0,"",AVERAGE( Sorting!K:K, Sorting!O:O))',
         '=SUM(G4,M4)',
        ),
        # Row 5 — Hour 9
        (5, 9,
         '=COUNTIF( Sorting!U:U,"SOFT CLAY")',
         '=COUNTIF( Sorting!V:V,"SOFT CLAY")',
         '=IF(B5=0,"",MIN( Sorting!W:W))',
         '=IF(B5=0,"",MAX( Sorting!W:W))',
         '=IF(B5=0,"",AVERAGE( Sorting!W:W))',
         '=COUNTIF( Sorting!X:X,"fail")',
         '=COUNTIF( Sorting!Q:Q,"GOOD EARTH")',
         '=COUNTIF( Sorting!R:R,"GOOD EARTH")',
         '=IF(H5=0,"",MIN( Sorting!S:S))',
         '=IF(H5=0,"",MAX( Sorting!S:S))',
         '=IF(H5=0,"",AVERAGE( Sorting!S:S))',
         '=COUNTIF( Sorting!T:T,"fail")',
         '=SUM(B5,H5)', '=SUM(C5,I5)',
         '=IF(N5=0,"",MIN(D5,J5))',
         '=IF(N5=0,"",MAX(E5,K5))',
         '=IF(N5=0,"",AVERAGE( Sorting!S:S, Sorting!W:W))',
         '=SUM(G5,M5)',
        ),
        # Row 6 — Hour 10
        (6, 10,
         '=COUNTIF( Sorting!AC:AC,"SOFT CLAY")',
         '=COUNTIF( Sorting!AD:AD,"SOFT CLAY")',
         '=IF(B6=0,"",MIN( Sorting!AE:AE))',
         '=IF(B6=0,"",MAX( Sorting!AE:AE))',
         '=IF(B6=0,"",AVERAGE( Sorting!AE:AE))',
         '=COUNTIF( Sorting!AF:AF,"fail")',
         '=COUNTIF( Sorting!Y:Y,"GOOD EARTH")',
         '=COUNTIF( Sorting!Z:Z,"GOOD EARTH")',
         '=IF(H6=0,"",MIN( Sorting!AA:AA))',
         '=IF(H6=0,"",MAX( Sorting!AA:AA))',
         '=IF(H6=0,"",AVERAGE( Sorting!AA:AA))',
         '=COUNTIF( Sorting!AB:AB,"fail")',
         '=SUM(B6,H6)', '=SUM(C6,I6)',
         '=IF(N6=0,"",MIN(D6,J6))',
         '=IF(N6=0,"",MAX(E6,K6))',
         '=IF(N6=0,"",AVERAGE( Sorting!AA:AA, Sorting!AE:AE))',
         '=SUM(G6,M6)',
        ),
        # Row 7 — Hour 11
        (7, 11,
         '=COUNTIF( Sorting!AK:AK,"SOFT CLAY")',
         '=COUNTIF( Sorting!AL:AL,"SOFT CLAY")',
         '=IF(B7=0,"",MIN( Sorting!AM:AM))',
         '=IF(B7=0,"",MAX( Sorting!AM:AM))',
         '=IF(B7=0,"",AVERAGE( Sorting!AM:AM))',
         '=COUNTIF( Sorting!AN:AN,"fail")',
         '=COUNTIF( Sorting!AG:AG,"GOOD EARTH")',
         '=COUNTIF( Sorting!AH:AH,"GOOD EARTH")',
         '=IF(H7=0,"",MIN( Sorting!AI:AI))',
         '=IF(H7=0,"",MAX( Sorting!AI:AI))',
         '=IF(H7=0,"",AVERAGE( Sorting!AI:AI))',
         '=COUNTIF( Sorting!AJ:AJ,"fail")',
         '=SUM(B7,H7)', '=SUM(C7,I7)',
         '=IF(N7=0,"",MIN(D7,J7))',
         '=IF(N7=0,"",MAX(E7,K7))',
         '=IF(N7=0,"",AVERAGE( Sorting!AI:AI, Sorting!AM:AM))',
         '=SUM(G7,M7)',
        ),
        # Row 8 — Hour 12
        (8, 12,
         '=COUNTIF( Sorting!AS:AS,"SOFT CLAY")',
         '=COUNTIF( Sorting!AT:AT,"SOFT CLAY")',
         '=IF(B8=0,"",MIN( Sorting!AU:AU))',
         '=IF(B8=0,"",MAX( Sorting!AU:AU))',
         '=IF(B8=0,"",AVERAGE( Sorting!AU:AU))',
         '=COUNTIF( Sorting!AV:AV,"fail")',
         '=COUNTIF( Sorting!AO:AO,"GOOD EARTH")',
         '=COUNTIF( Sorting!AP:AP,"GOOD EARTH")',
         '=IF(H8=0,"",MIN( Sorting!AQ:AQ))',
         '=IF(H8=0,"",MAX( Sorting!AQ:AQ))',
         '=IF(H8=0,"",AVERAGE( Sorting!AQ:AQ))',
         '=COUNTIF( Sorting!AR:AR,"fail")',
         '=SUM(B8,H8)', '=SUM(C8,I8)',
         '=IF(N8=0,"",MIN(D8,J8))',
         '=IF(N8=0,"",MAX(E8,K8))',
         '=IF(N8=0,"",AVERAGE( Sorting!AQ:AQ, Sorting!AU:AU))',
         '=SUM(G8,M8)',
        ),
        # Row 9 — Hour 13
        (9, 13,
         '=COUNTIF( Sorting!BA:BA,"SOFT CLAY")',
         '=COUNTIF( Sorting!BB:BB,"SOFT CLAY")',
         '=IF(B9=0,"",MIN( Sorting!BC:BC))',
         '=IF(B9=0,"",MAX( Sorting!BC:BC))',
         '=IF(B9=0,"",AVERAGE( Sorting!BC:BC))',
         '=COUNTIF( Sorting!BD:BD,"fail")',
         '=COUNTIF( Sorting!AW:AW,"GOOD EARTH")',
         '=COUNTIF( Sorting!AX:AX,"GOOD EARTH")',
         '=IF(H9=0,"",MIN( Sorting!AY:AY))',
         '=IF(H9=0,"",MAX( Sorting!AY:AY))',
         '=IF(H9=0,"",AVERAGE( Sorting!AY:AY))',
         '=COUNTIF( Sorting!AZ:AZ,"fail")',
         '=SUM(B9,H9)', '=SUM(C9,I9)',
         '=IF(N9=0,"",MIN(D9,J9))',
         '=IF(N9=0,"",MAX(E9,K9))',
         "=IF(N9=0,\"\",AVERAGE( Sorting!AY:AY, Sorting!BC:BC))",
         '=SUM(G9,M9)',
        ),
        # Row 10 — Hour 14
        (10, 14,
         '=COUNTIF( Sorting!BI:BI,"SOFT CLAY")',
         '=COUNTIF( Sorting!BJ:BJ,"SOFT CLAY")',
         '=IF(B10=0,"",MIN( Sorting!BK:BK))',
         '=IF(B10=0,"",MAX( Sorting!BK:BK))',
         '=IF(B10=0,"",AVERAGE( Sorting!BK:BK))',
         '=COUNTIF( Sorting!BL:BL,"fail")',
         '=COUNTIF( Sorting!BE:BE,"GOOD EARTH")',
         '=COUNTIF( Sorting!BF:BF,"GOOD EARTH")',
         '=IF(H10=0,"",MIN( Sorting!BG:BG))',
         '=IF(H10=0,"",MAX( Sorting!BG:BG))',
         '=IF(H10=0,"",AVERAGE( Sorting!BG:BG))',
         '=COUNTIF( Sorting!BH:BH,"fail")',
         '=SUM(B10,H10)', '=SUM(C10,I10)',
         '=IF(N10=0,"",MIN(D10,J10))',
         '=IF(N10=0,"",MAX(E10,K10))',
         '=IF(N10=0,"",AVERAGE( Sorting!BG:BG, Sorting!BK:BK))',
         '=SUM(G10,M10)',
        ),
        # Row 11 — Hour 15
        (11, 15,
         '=COUNTIF( Sorting!BQ:BQ,"SOFT CLAY")',
         '=COUNTIF( Sorting!BR:BR,"SOFT CLAY")',
         '=IF(B11=0,"",MIN( Sorting!BS:BS))',
         '=IF(B11=0,"",MAX( Sorting!BS:BS))',
         '=IF(B11=0,"",AVERAGE( Sorting!BS:BS))',
         '=COUNTIF( Sorting!BT:BT,"fail")',
         '=COUNTIF( Sorting!BM:BM,"GOOD EARTH")',
         '=COUNTIF( Sorting!BN:BN,"GOOD EARTH")',
         '=IF(H11=0,"",MIN( Sorting!BO:BO))',
         '=IF(H11=0,"",MAX( Sorting!BO:BO))',
         '=IF(H11=0,"",AVERAGE( Sorting!BO:BO))',
         '=COUNTIF( Sorting!BP:BP,"fail")',
         '=SUM(B11,H11)', '=SUM(C11,I11)',
         '=IF(N11=0,"",MIN(D11,J11))',
         '=IF(N11=0,"",MAX(E11,K11))',
         '=IF(N11=0,"",AVERAGE( Sorting!BO:BO, Sorting!BS:BS))',
         '=SUM(G11,M11)',
        ),
        # Row 12 — Hour 16
        (12, 16,
         '=COUNTIF( Sorting!BY:BY,"SOFT CLAY")',
         '=COUNTIF( Sorting!BZ:BZ,"SOFT CLAY")',
         '=IF(B12=0,"",MIN( Sorting!CA:CA))',
         '=IF(B12=0,"",MAX( Sorting!CA:CA))',
         '=IF(B12=0,"",AVERAGE( Sorting!CA:CA))',
         '=COUNTIF( Sorting!CB:CB,"fail")',
         '=COUNTIF( Sorting!BU:BU,"GOOD EARTH")',
         '=COUNTIF( Sorting!BV:BV,"GOOD EARTH")',
         '=IF(H12=0,"",MIN( Sorting!BW:BW))',
         '=IF(H12=0,"",MAX( Sorting!BW:BW))',
         '=IF(H12=0,"",AVERAGE( Sorting!BW:BW))',
         '=COUNTIF( Sorting!BX:BX,"fail")',
         '=SUM(B12,H12)', '=SUM(C12,I12)',
         '=IF(N12=0,"",MIN(D12,J12))',
         '=IF(N12=0,"",MAX(E12,K12))',
         '=IF(N12=0,"",AVERAGE( Sorting!BW:BW, Sorting!CA:CA))',
         '=SUM(G12,M12)',
        ),
        # Row 13 — Hour 17
        (13, 17,
         '=COUNTIF( Sorting!CG:CG,"SOFT CLAY")',
         '=COUNTIF( Sorting!CH:CH,"SOFT CLAY")',
         '=IF(B13=0,"",MIN( Sorting!CI:CI))',
         '=IF(B13=0,"",MAX( Sorting!CI:CI))',
         '=IF(B13=0,"",AVERAGE( Sorting!CI:CI))',
         '=COUNTIF( Sorting!CJ:CJ,"fail")',
         '=COUNTIF( Sorting!CC:CC,"GOOD EARTH")',
         '=COUNTIF( Sorting!CD:CD,"GOOD EARTH")',
         '=IF(H13=0,"",MIN( Sorting!CE:CE))',
         '=IF(H13=0,"",MAX( Sorting!CE:CE))',
         '=IF(H13=0,"",AVERAGE( Sorting!CE:CE))',
         '=COUNTIF( Sorting!CF:CF,"fail")',
         '=SUM(B13,H13)', '=SUM(C13,I13)',
         '=IF(N13=0,"",MIN(D13,J13))',
         '=IF(N13=0,"",MAX(E13,K13))',
         '=IF(N13=0,"",AVERAGE( Sorting!CE:CE, Sorting!CI:CI))',
         '=SUM(G13,M13)',
        ),
        # Row 14 — Hour 18
        (14, 18,
         '=COUNTIF( Sorting!CO:CO,"SOFT CLAY")',
         '=COUNTIF( Sorting!CP:CP,"SOFT CLAY")',
         '=IF(B14=0,"",MIN( Sorting!CQ:CQ))',
         '=IF(B14=0,"",MAX( Sorting!CQ:CQ))',
         '=IF(B14=0,"",AVERAGE( Sorting!CQ:CQ))',
         '=COUNTIF( Sorting!CR:CR,"fail")',
         '=COUNTIF( Sorting!CK:CK,"GOOD EARTH")',
         '=COUNTIF( Sorting!CL:CL,"GOOD EARTH")',
         '=IF(H14=0,"",MIN( Sorting!CM:CM))',
         '=IF(H14=0,"",MAX( Sorting!CM:CM))',
         '=IF(H14=0,"",AVERAGE( Sorting!CM:CM))',
         '=COUNTIF( Sorting!CN:CN,"fail")',
         '=SUM(B14,H14)', '=SUM(C14,I14)',
         '=IF(N14=0,"",MIN(D14,J14))',
         '=IF(N14=0,"",MAX(E14,K14))',
         '=IF(N14=0,"",AVERAGE( Sorting!CM:CM, Sorting!CQ:CQ))',
         '=SUM(G14,M14)',
        ),
        # Row 15 — Hour 19
        (15, 19,
         '=COUNTIF( Sorting!CW:CW,"SOFT CLAY")',
         '=COUNTIF( Sorting!CX:CX,"SOFT CLAY")',
         '=IF(B15=0,"",MIN( Sorting!CY:CY))',
         '=IF(B15=0,"",MAX( Sorting!CY:CY))',
         '=IF(B15=0,"",AVERAGE( Sorting!CY:CY))',
         '=COUNTIF( Sorting!CZ:CZ,"fail")',
         '=COUNTIF( Sorting!CS:CS,"GOOD EARTH")',
         '=COUNTIF( Sorting!CT:CT,"GOOD EARTH")',
         '=IF(H15=0,"",MIN( Sorting!CU:CU))',
         '=IF(H15=0,"",MAX( Sorting!CU:CU))',
         '=IF(H15=0,"",AVERAGE( Sorting!CU:CU))',
         '=COUNTIF( Sorting!CV:CV,"fail")',
         '=SUM(B15,H15)', '=SUM(C15,I15)',
         '=IF(N15=0,"",MIN(D15,J15))',
         '=IF(N15=0,"",MAX(E15,K15))',
         '=IF(N15=0,"",AVERAGE( Sorting!CU:CU, Sorting!CY:CY))',
         '=SUM(G15,M15)',
        ),
        # Row 16 — Hour 20
        (16, 20,
         '=COUNTIF( Sorting!DE:DE,"SOFT CLAY")',
         '=COUNTIF( Sorting!DF:DF,"SOFT CLAY")',
         '=IF(B16=0,"",MIN( Sorting!DG:DG))',
         '=IF(B16=0,"",MAX( Sorting!DG:DG))',
         '=IF(B16=0,"",AVERAGE( Sorting!DG:DG))',
         '=COUNTIF( Sorting!DH:DH,"fail")',
         '=COUNTIF( Sorting!DA:DA,"GOOD EARTH")',
         '=COUNTIF( Sorting!DB:DB,"GOOD EARTH")',
         '=IF(H16=0,"",MIN( Sorting!DC:DC))',
         '=IF(H16=0,"",MAX( Sorting!DC:DC))',
         '=IF(H16=0,"",AVERAGE( Sorting!DC:DC))',
         '=COUNTIF( Sorting!DD:DD,"fail")',
         '=SUM(B16,H16)', '=SUM(C16,I16)',
         '=IF(N16=0,"",MIN(D16,J16))',
         '=IF(N16=0,"",MAX(E16,K16))',
         '=IF(N16=0,"",AVERAGE( Sorting!DC:DC, Sorting!DG:DG))',
         '=SUM(G16,M16)',
        ),
        # Row 17 — Hour 21
        (17, 21,
         '=COUNTIF( Sorting!DM:DM,"SOFT CLAY")',
         '=COUNTIF( Sorting!DN:DN,"SOFT CLAY")',
         '=IF(B17=0,"",MIN( Sorting!DO:DO))',
         '=IF(B17=0,"",MAX( Sorting!DO:DO))',
         '=IF(B17=0,"",AVERAGE( Sorting!DO:DO))',
         '=COUNTIF( Sorting!DP:DP,"fail")',
         '=COUNTIF( Sorting!DI:DI,"GOOD EARTH")',
         '=COUNTIF( Sorting!DJ:DJ,"GOOD EARTH")',
         '=IF(H17=0,"",MIN( Sorting!DK:DK))',
         '=IF(H17=0,"",MAX( Sorting!DK:DK))',
         '=IF(H17=0,"",AVERAGE( Sorting!DK:DK))',
         '=COUNTIF( Sorting!DL:DL,"fail")',
         '=SUM(B17,H17)', '=SUM(C17,I17)',
         '=IF(N17=0,"",MIN(D17,J17))',
         '=IF(N17=0,"",MAX(E17,K17))',
         '=IF(N17=0,"",AVERAGE( Sorting!DK:DK, Sorting!DO:DO))',
         '=SUM(G17,M17)',
        ),
        # Row 18 — Hour 22
        (18, 22,
         '=COUNTIF( Sorting!DU:DU,"SOFT CLAY")',
         '=COUNTIF( Sorting!DV:DV,"SOFT CLAY")',
         '=IF(B18=0,"",MIN( Sorting!DW:DW))',
         '=IF(B18=0,"",MAX( Sorting!DW:DW))',
         '=IF(B18=0,"",AVERAGE( Sorting!DW:DW))',
         '=COUNTIF( Sorting!DX:DX,"fail")',
         '=COUNTIF( Sorting!DQ:DQ,"GOOD EARTH")',
         '=COUNTIF( Sorting!DR:DR,"GOOD EARTH")',
         '=IF(H18=0,"",MIN( Sorting!DS:DS))',
         '=IF(H18=0,"",MAX( Sorting!DS:DS))',
         '=IF(H18=0,"",AVERAGE( Sorting!DS:DS))',
         '=COUNTIF( Sorting!DT:DT,"fail")',
         '=SUM(B18,H18)', '=SUM(C18,I18)',
         '=IF(N18=0,"",MIN(D18,J18))',
         '=IF(N18=0,"",MAX(E18,K18))',
         '=IF(N18=0,"",AVERAGE( Sorting!DS:DS, Sorting!DW:DW))',
         '=SUM(G18,M18)',
        ),
        # Row 19 — Hour 23
        (19, 23,
         '=COUNTIF( Sorting!EC:EC,"SOFT CLAY")',
         '=COUNTIF( Sorting!ED:ED,"SOFT CLAY")',
         '=IF(B19=0,"",MIN( Sorting!EE:EE))',
         '=IF(B19=0,"",MAX( Sorting!EE:EE))',
         '=IF(B19=0,"",AVERAGE( Sorting!EE:EE))',
         '=COUNTIF( Sorting!EF:EF,"fail")',
         '=COUNTIF( Sorting!DY:DY,"GOOD EARTH")',
         '=COUNTIF( Sorting!DZ:DZ,"GOOD EARTH")',
         '=IF(H19=0,"",MIN( Sorting!EA:EA))',
         '=IF(H19=0,"",MAX( Sorting!EA:EA))',
         '=IF(H19=0,"",AVERAGE( Sorting!EA:EA))',
         '=COUNTIF( Sorting!EB:EB,"fail")',
         '=SUM(B19,H19)', '=SUM(C19,I19)',
         '=IF(N19=0,"",MIN(D19,J19))',
         '=IF(N19=0,"",MAX(E19,K19))',
         '=IF(N19=0,"",AVERAGE( Sorting!EA:EA, Sorting!EE:EE))',
         '=SUM(G19,M19)',
        ),
        # Row 20 — Hour 0
        (20, 0,
         '=COUNTIF( Sorting!EK:EK,"SOFT CLAY")',
         '=COUNTIF( Sorting!EL:EL,"SOFT CLAY")',
         None, None, None,
         '=COUNTIF( Sorting!EN:EN,"fail")',
         '=COUNTIF( Sorting!EG:EG,"GOOD EARTH")',
         '=COUNTIF( Sorting!EH:EH,"GOOD EARTH")',
         None, None, None,
         '=COUNTIF( Sorting!EJ:EJ,"fail")',
         '=SUM(B20,H20)', '=SUM(C20,I20)',
         None, None, None,
         '=SUM(G20,M20)',
        ),
        # Row 21 — Hour 1
        (21, 1,
         '=COUNTIF( Sorting!ES:ES,"SOFT CLAY")',
         '=COUNTIF( Sorting!ET:ET,"SOFT CLAY")',
         None, None, None,
         '=COUNTIF( Sorting!EV:EV,"fail")',
         '=COUNTIF( Sorting!EO:EO,"GOOD EARTH")',
         '=COUNTIF( Sorting!EP:EP,"GOOD EARTH")',
         None, None, None,
         '=COUNTIF( Sorting!ER:ER,"fail")',
         '=SUM(B21,H21)', '=SUM(C21,I21)',
         None, None, None,
         '=SUM(G21,M21)',
        ),
        # Row 22 — Hour 2
        (22, 2,
         '=COUNTIF( Sorting!FA:FA,"SOFT CLAY")',
         '=COUNTIF( Sorting!FB:FB,"SOFT CLAY")',
         None, None, None,
         '=COUNTIF( Sorting!FD:FD,"fail")',
         '=COUNTIF( Sorting!EW:EW,"GOOD EARTH")',
         '=COUNTIF( Sorting!EX:EX,"GOOD EARTH")',
         None, None, None,
         '=COUNTIF( Sorting!EZ:EZ,"fail")',
         '=SUM(B22,H22)', '=SUM(C22,I22)',
         None, None, None,
         '=SUM(G22,M22)',
        ),
        # Row 23 — Hour 3
        (23, 3,
         '=COUNTIF( Sorting!FI:FI,"SOFT CLAY")',
         '=COUNTIF( Sorting!FJ:FJ,"SOFT CLAY")',
         None, None, None,
         '=COUNTIF( Sorting!FL:FL,"fail")',
         '=COUNTIF( Sorting!FE:FE,"GOOD EARTH")',
         '=COUNTIF( Sorting!FF:FF,"GOOD EARTH")',
         None, None, None,
         '=COUNTIF( Sorting!FH:FH,"fail")',
         '=SUM(B23,H23)', '=SUM(C23,I23)',
         None, None, None,
         '=SUM(G23,M23)',
        ),
        # Row 24 — Hour 4
        (24, 4,
         '=COUNTIF( Sorting!FI:FI,"SOFT CLAY")',
         '=COUNTIF( Sorting!FR:FR,"SOFT CLAY")',
         None, None, None,
         '=COUNTIF( Sorting!FT:FT,"fail")',
         '=COUNTIF( Sorting!FM:FM,"GOOD EARTH")',
         '=COUNTIF( Sorting!FN:FN,"GOOD EARTH")',
         None, None, None,
         '=COUNTIF( Sorting!FP:FP,"fail")',
         '=SUM(B24,H24)', '=SUM(C24,I24)',
         None, None, None,
         '=SUM(G24,M24)',
        ),
    ]

    # Column mapping: index 0=B, 1=C ... 17=S
    COL_LETTERS = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S']

    for row_data in DATA_ROWS:
        r       = row_data[0]
        hour    = row_data[1]
        formulas = row_data[2:]   # 18 values for cols B–S

        # Column A: hour label
        _w(ws.cell(r, 1), hour)

        # Columns B–S: formulas (or None = leave blank)
        for ci, (col_letter, formula) in enumerate(zip(COL_LETTERS, formulas)):
            if formula is not None:
                _w(ws[f'{col_letter}{r}'], formula)




def build_sorting_sheet(wb_out, df_wb=None):
    """
    Sorting sheet: exact formulas from demo.
    - 23 hour-blocks, each 8 columns wide (A-H, I-P, Q-X, ...)
    - Row 1: hour label in first col of each block only
    - Row 2: 8 sub-headers per block (Online GE | WB GE | GE Duration | GE Failure |
                                       Online SC | WB SC | SC Duration | SC Failure)
    - Rows 3+: formulas referencing 'Online Total' and 'WB Total'
    - Bold medium border on column H of each block (every 8th column right-side)
    - NO color fills — plain text only
    - Source row = sorting_row - 1 (row 3 → Online Total row 2, row 4 → row 3, etc.)
    """
    ws = wb_out.create_sheet('Sorting')

    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    FONT_NAME = 'Calibri'
    FONT_SIZE = 10
    PLAIN_FILL = PatternFill(fill_type=None)

    THIN  = Side(style='thin',   color='000000')
    MED   = Side(style='medium', color='000000')
    NO_BD = Side(style=None)

    def _border(right_med=False):
        return Border(
            left=NO_BD, top=NO_BD, bottom=NO_BD,
            right=MED if right_med else NO_BD
        )

    def _w(cell, value, bold=False, right_med=False):
        cell.value = value
        cell.font  = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border(right_med)

    # 23 hour blocks in demo order, with their start times and column ranges
    HOUR_BLOCKS = [
        ('0700-0759', 700,  800),
        ('0800-0859', 800,  900),
        ('0900-0959', 900,  1000),
        ('1000-1059', 1000, 1100),
        ('1100-1159', 1100, 1200),
        ('1200-1259', 1200, 1300),
        ('1300-1359', 1300, 1400),
        ('1400-1459', 1400, 1500),
        ('1500-1559', 1500, 1600),
        ('1600-1659', 1600, 1700),
        ('1700-1759', 1700, 1800),
        ('1800-1859', 1800, 1900),
        ('1900-1959', 1900, 2000),
        ('2000-2059', 2000, 2100),
        ('2100-2159', 2100, 2200),
        ('2200-2259', 2200, 2300),
        ('2300-2359', 2300, 2400),
        ('0000-0059', 0,    100),
        ('0100-0159', 100,  200),
        ('0200-0259', 200,  300),
        ('0300-0359', 300,  400),
        ('0400-0459', 400,  500),
        ('0500-0559', 500,  600),
    ]

    SUB_HDRS = ['Online GE','WB GE','GE Duration','GE Failure',
                'Online SC','WB SC','SC Duration','SC Failure']

    # ── Row 1: hour labels ────────────────────────────────────────────────────
    for bi, (label, lo, hi) in enumerate(HOUR_BLOCKS):
        base_col = bi * 8 + 1         # 1, 9, 17, 25 ...
        last_col = base_col + 7        # 8, 16, 24, 32 ...
        _w(ws.cell(1, base_col), label, bold=True)
        # Bold right border on the 8th column of every block
        _w(ws.cell(1, last_col), None, right_med=True)

    # ── Row 2: sub-headers ───────────────────────────────────────────────────
    for bi in range(len(HOUR_BLOCKS)):
        base_col = bi * 8 + 1
        for si, sub in enumerate(SUB_HDRS):
            is_last = (si == 7)
            _w(ws.cell(2, base_col + si), sub, right_med=is_last)

    # ── Data rows 3 onwards: formulas ─────────────────────────────────────────
    # Source row in Online Total / WB Total = sorting_row - 1
    # (sorting row 3 → source row 2, sorting row 4 → source row 3, ...)
    # Demo has formulas up to row 7453 (3039 Online + WB rows × headroom)
    # We write formulas for all source rows covering Online Total (3040 data + header)
    # plus WB Total (3040 data + header): use 7452 data rows to match demo
    # Use actual data size instead of fixed 7451 — dramatically faster
    _n_online = max(df_wb.shape[0] if df_wb is not None else 0, 100)
    N_DATA_ROWS = min(_n_online + 50, 3500)   # cap at 3500 for performance

    for data_idx in range(N_DATA_ROWS):
        sr = data_idx + 3      # sorting row (3-based)
        src = sr - 1           # source row in Online Total / WB Total

        for bi, (label, lo, hi) in enumerate(HOUR_BLOCKS):
            base_col = bi * 8 + 1

            # Build the time range condition string
            # For blocks crossing midnight (0000-0059 etc.) lo < hi always
            # except 2300-2359 which is 2300 <= x < 2400
            # The demo uses numeric HHMM: 700 <= F2 < 800
            lo_str = str(lo)
            hi_str = str(hi)
            pc_col = 'F'    # Peak Calculation column in Online Total (col F)
            wb_pc  = 'L'    # Peak Calculation column in WB Total (col L)
            mat_col_ot = 'K'  # Material col in Online Total
            mat_col_wb = 'F'  # Material col in WB Total
            dur_col_ot  = 'J' # Duration col in Online Total
            fail_col_ot = 'Q' # Failure col in Online Total (col Q = Failure in our sheet)

            # ── Col 0 (Online GE): from Online Total ──────────────────────────
            c0 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"GOOD EARTH\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${mat_col_ot}{src},\"\")"
            cell = ws.cell(sr, base_col)
            _w(cell, c0)

            # ── Col 1 (WB GE): from WB Total ─────────────────────────────────
            c1 = f"=IF(AND('WB Total'!${mat_col_wb}{src}=\"GOOD EARTH\",{lo_str}<='WB Total'!${wb_pc}{src},'WB Total'!${wb_pc}{src}<{hi_str}),'WB Total'!${mat_col_wb}{src},\"\")"
            _w(ws.cell(sr, base_col + 1), c1)

            # ── Col 2 (GE Duration): from Online Total ────────────────────────
            c2 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"GOOD EARTH\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${dur_col_ot}{src},\"\")"
            _w(ws.cell(sr, base_col + 2), c2)

            # ── Col 3 (GE Failure): from Online Total ─────────────────────────
            c3 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"GOOD EARTH\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${fail_col_ot}{src},\"\")"
            _w(ws.cell(sr, base_col + 3), c3)

            # ── Col 4 (Online SC): from Online Total ──────────────────────────
            c4 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"SOFT CLAY\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${mat_col_ot}{src},\"\")"
            _w(ws.cell(sr, base_col + 4), c4)

            # ── Col 5 (WB SC): from WB Total ──────────────────────────────────
            c5 = f"=IF(AND('WB Total'!${mat_col_wb}{src}=\"SOFT CLAY\",{lo_str}<='WB Total'!${wb_pc}{src},'WB Total'!${wb_pc}{src}<{hi_str}),'WB Total'!${mat_col_wb}{src},\"\")"
            _w(ws.cell(sr, base_col + 5), c5)

            # ── Col 6 (SC Duration): from Online Total ────────────────────────
            c6 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"SOFT CLAY\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${dur_col_ot}{src},\"\")"
            _w(ws.cell(sr, base_col + 6), c6)

            # ── Col 7 (SC Failure): from Online Total + bold right border ──────
            c7 = f"=IF(AND('Online Total'!${mat_col_ot}{src}=\"SOFT CLAY\",{lo_str}<='Online Total'!${pc_col}{src},'Online Total'!${pc_col}{src}<{hi_str}),'Online Total'!${fail_col_ot}{src},\"\")"
            _w(ws.cell(sr, base_col + 7), c7, right_med=True)


def build_charts_sheet(wb_out, df_online=None, df_wb=None, report_dt=None):
    """
    Charts sheet — exact formulas from demo + full formatting spec:
    - Row 1 : merged entire row (empty)
    - Row 2 : h=14.40, L2:N2 merged, font size 11
    - Row 3 : h=105.40, column headers + red/yellow fills
    - Rows 4-26: h=14.40, all borders, formulas intact
    - Row 27 : no borders, formulas only (MIN/MAX/AVG)
    - Below row 27: 3 embedded bar charts (Total Trucks, GE, SC)
      with yellow/grey/orange bars and exact title colors from images
    - Column widths per spec
    - All font size 11
    """
    ws = wb_out.create_sheet('Charts')

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.utils import get_column_letter

    FN = 'Calibri'
    FS = 11                                      # global font size 11
    RED_FILL = PatternFill('solid', fgColor='FF0000')
    YEL_FILL = PatternFill('solid', fgColor='FFFF00')
    NO_FILL  = PatternFill(fill_type=None)

    _thin = Side(style='thin', color='000000')
    ALL_BDR = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    NO_BDR  = Border()

    def _w(cell, value, bold=False, fill=None, border=ALL_BDR,
           h_align='center', wrap=True, font_color='000000'):
        cell.value = value
        cell.font  = Font(name=FN, size=FS, bold=bold, color=font_color)
        cell.alignment = Alignment(horizontal=h_align, vertical='center',
                                   wrap_text=wrap)
        if fill:
            cell.fill = fill
        if border is not None:
            cell.border = border

    # ── Column widths (exact per spec) ────────────────────────────────────────
    COL_WIDTHS = {
        'A': 15.11, 'B': 10.67, 'C': 10.78,
        'D': 12.67, 'E': 13.78, 'F':  9.56,
        'G': 13.56, 'H': 11.11, 'I': 14.33,
        'J':  8.89, 'K':  8.89,
        'L': 15.78, 'M':  9.33, 'N':  7.67,
        'O':  2.22,
        'P': 35.56, 'Q': 10.44, 'R':  6.89,
    }
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: merged empty row ───────────────────────────────────────────────
    ws.merge_cells('A1:R1')
    ws.row_dimensions[1].height = 14.40
    _w(ws['A1'], None, border=NO_BDR)

    # ── Row 2: h=14.40, L2:N2 merged, font 11, all borders ───────────────────
    ws.row_dimensions[2].height = 14.40
    _w(ws['A2'], "='Online Total'!D2")
    _w(ws['J2'], 72)
    _w(ws['K2'], 288)
    ws.merge_cells('L2:N2')
    _w(ws['L2'], 'Cycle Time Duration (minutes)')
    _w(ws['P2'], 'Provision Number of Good Earth Trucks')
    _w(ws['Q2'], 200000)
    _w(ws['R2'], '/month')

    # ── Row 3: h=105.40, column headers + fills + all borders ─────────────────
    ws.row_dimensions[3].height = 105.40
    HDR_SPECS = {
        'A': ('Time',                  NO_FILL),
        'B': ('Exceedance',            NO_FILL),
        'C': ('Applicable',            NO_FILL),
        'D': ('Online Truck No. (GE)', RED_FILL),
        'E': ('WB Truck No. (GE)',     YEL_FILL),
        'F': ('Online Truck no. (SC)', RED_FILL),
        'G': ('WB Truck no. (SC)',     YEL_FILL),
        'H': ('Online Total Trucks',   RED_FILL),
        'I': ('WB Total Trucks',       YEL_FILL),
        'J': ('WB Hour that exceeded the expected no. of trucks GE', NO_FILL),
        'K': ('WB Hour that exceeded the expected no. of trucks SC', NO_FILL),
        'L': ('Minimum',  NO_FILL),
        'M': ('Maximum',  NO_FILL),
        'N': ('Average',  NO_FILL),
    }
    for col, (text, fill) in HDR_SPECS.items():
        _w(ws[f'{col}3'], text, bold=True, fill=fill)
    # Right panel row 3
    _w(ws['Q3'], '=Q2/29', border=NO_BDR)
    _w(ws['R3'], '/day',   border=NO_BDR)

    # ── Rows 4-26: h=14.40, all borders, exact formulas ───────────────────────
    HOUR_LABELS = [
        '7AM - 8AM','8AM - 9AM','9AM - 10AM','10AM - 11AM',
        '11AM - 12PM','12PM - 1PM','1PM - 2PM','2PM - 3PM',
        '3PM - 4PM','4PM - 5PM','5PM - 6PM','6PM - 7PM',
        '7PM - 8PM','8PM - 9PM','9PM - 10PM','10PM - 11PM',
        '11PM - 12AM','12AM - 1AM','1AM - 2AM','2AM - 3AM',
        '3AM - 4AM','4AM - 5AM',
    ]

    for i, label in enumerate(HOUR_LABELS):
        r   = i + 4
        pvr = i + 3
        ws.row_dimensions[r].height = 14.40
        _w(ws[f'A{r}'], label)
        _w(ws[f'B{r}'], f"='Pivot Table'!S{pvr}")
        _w(ws[f'C{r}'], f'=IF(J{r}="Yes", "0", B{r})')
        _w(ws[f'D{r}'], f"='Pivot Table'!H{pvr}", fill=RED_FILL)
        _w(ws[f'E{r}'], f"='Pivot Table'!I{pvr}", fill=YEL_FILL)
        _w(ws[f'F{r}'], f"='Pivot Table'!B{pvr}", fill=RED_FILL)
        _w(ws[f'G{r}'], f"='Pivot Table'!C{pvr}", fill=YEL_FILL)
        _w(ws[f'H{r}'], f'=D{r}+F{r}', fill=RED_FILL)
        _w(ws[f'I{r}'], f'=E{r}+G{r}', fill=YEL_FILL)
        _w(ws[f'J{r}'], f'=IF(E{r}>72, "Yes", "No")')
        _w(ws[f'K{r}'], f'=IF(G{r}>288, "Yes", "No")')
        _w(ws[f'L{r}'], f"='Pivot Table'!P{pvr}")
        _w(ws[f'M{r}'], f"='Pivot Table'!Q{pvr}")
        _w(ws[f'N{r}'], f"='Pivot Table'!R{pvr}")

    # Right panel provision (outside table — no borders)
    _w(ws['P4'], 'Phase 1A (10m3 per truck)', bold=True, border=NO_BDR)
    _w(ws['Q4'], '=Q3/10', bold=True, border=NO_BDR)
    _w(ws['R4'], '/day',   bold=True, border=NO_BDR)
    _w(ws['P6'], 'Provision Number of Good Earth Trucks', border=NO_BDR)
    _w(ws['Q6'], 800000,   border=NO_BDR)
    _w(ws['R6'], '/month',  border=NO_BDR)
    _w(ws['Q7'], '=Q6/29', border=NO_BDR)
    _w(ws['R7'], '/day',   border=NO_BDR)
    _w(ws['P8'], 'Phase 1B (10m3 per truck)', bold=True, border=NO_BDR)
    _w(ws['Q8'], '=Q7/10', bold=True, border=NO_BDR)
    _w(ws['R8'], '/day',   bold=True, border=NO_BDR)

    # Row 26: Sum row — inside table, all borders
    ws.row_dimensions[26].height = 14.40
    _w(ws['A26'], 'Sum', bold=True)
    _w(ws['B26'], "='Pivot Table'!S25")
    _w(ws['C26'], '=SUM(C4:C20)')
    _w(ws['D26'], '=SUM(D4:D20)', fill=RED_FILL)
    _w(ws['E26'], '=SUM(E4:E25)', fill=YEL_FILL)
    _w(ws['F26'], '=SUM(F4:F20)', fill=RED_FILL)
    _w(ws['G26'], '=SUM(G4:G25)', fill=YEL_FILL)
    _w(ws['H26'], '=SUM(H4:H20)', fill=RED_FILL)
    _w(ws['I26'], '=SUM(I4:I25)', fill=YEL_FILL)
    _w(ws['J26'], '')
    _w(ws['K26'], '')
    _w(ws['L26'], 'Total Exceedence')
    _w(ws['M26'], '=IF(E26>$Q$4, 0, IF(G26>Q8, 0, (SUM(C3:C20))))')
    _w(ws['N26'], '')

    # Row 27: Totals — NO borders, formulas only
    ws.row_dimensions[27].height = 14.40
    _w(ws['J27'], 'Total', border=NO_BDR)
    _w(ws['L27'], "=MIN('Online Total'!M:M)",     border=NO_BDR)
    _w(ws['M27'], "=MAX('Online Total'!M:M)",     border=NO_BDR)
    _w(ws['N27'], "=AVERAGE('Online Total'!M:M)", border=NO_BDR)
    ws['L27'].number_format = '0.0'
    ws['M27'].number_format = '0.0'
    ws['N27'].number_format = '0.0'

    # ── 3 Bar Charts (Total Trucks=yellow, GE=grey, SC=orange) ───────────────
    # Data range: rows 4-25 (22 hour rows), col A = categories
    # Col I = WB Total Trucks (yellow), col E = WB GE, col G = WB SC
    DATA_START = 4
    DATA_END   = 25

    def _bar_chart(title, title_color, series_refs, bar_color, anchor):
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                            ParagraphProperties, CharacterProperties,
                                            RegularTextRun)
        bc = BarChart()
        bc.type = 'col'
        bc.grouping = 'clustered'
        bc.style = 10
        bc.width  = 30.52   # correct dimensions per spec
        bc.height = 17
        bc.x_axis.title = None
        bc.y_axis.title = None
        bc.x_axis.delete = False
        bc.y_axis.delete = False
        bc.legend = None
        try:
            rpr  = CharacterProperties(sz=2800, solidFill=title_color, b=False)
            run  = RegularTextRun(t=title, rPr=rpr)
            para = Paragraph(r=[run], pPr=ParagraphProperties(defRPr=rpr))
            rich = RichText(bodyPr=RichTextProperties(), p=[para])
            from openpyxl.chart.title import Title
            from openpyxl.chart.text import Text
            bc.title = Title(tx=Text(rich=rich), overlay=False)
        except Exception:
            bc.title = title
        for col_idx in series_refs:
            ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                            min_row=DATA_START, max_row=DATA_END)
            s = Series(ref, title_from_data=False)
            s.graphicalProperties.solidFill = bar_color
            s.graphicalProperties.line.solidFill = bar_color
            bc.series.append(s)
        cats = Reference(ws, min_col=1, max_col=1,
                         min_row=DATA_START, max_row=DATA_END)
        bc.set_categories(cats)
        ws.add_chart(bc, anchor)
        return bc

    # Charts: 17cm each. Rows 4-27 = data (26 rows).
    # Row 28 = small separator. Chart 1 at A29.
    # chart1 ends ~row 61 → gap row 62 → chart2 at A63.
    # chart2 ends ~row 95 → gap row 96 → chart3 at A97.
    _bar_chart('Total Trucks',    'FF0000', [9],  'FFFF00', 'A29')
    _bar_chart('Truck No. (GE)', '00B050', [5],  '808080', 'A63')
    _bar_chart('Truck no. (SC)', '0070C0', [7],  'FF6600', 'A97')



def build_full_report(df_online, df_anomaly, df_wb_total, df_wb_rejected,
                      failure_list=None, report_dt=None, start_dt=None, end_dt=None):
    """
    Build the complete 7-sheet report in the specified order:
      1. Online Total  (Online Data)
      2. WB Total
      3. WB Rejected Loads
      4. Anomaly and Not Considered
      5. Pivot Table
      6. Sorting
      7. Charts
    """
    # Generate the base 4-sheet workbook bytes from existing writer
    base_bytes = _write_phase2_xlsx(
        df_online, df_anomaly,
        failure_list.count('fail') if failure_list else 0,
        failure_list,
        report_date=report_dt or datetime.now(),
        start_dt=start_dt,
        end_dt=end_dt,
    )

    # Reload and add WB + extra sheets
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(base_bytes))

    # Rename "Online Data" → "Online Total" to match spec
    if 'Online Data' in wb.sheetnames:
        wb['Online Data'].title = 'Online Total'

    # Inject WB Total and WB Rejected sheets
    write_wb_phase_xlsx_into(wb, df_wb_total, df_wb_rejected)

    # Add the three analytical sheets
    build_pivot_sheet(wb, df_online, df_wb_total)
    build_sorting_sheet(wb, df_wb_total)
    build_charts_sheet(wb, df_online, df_wb_total, report_dt=report_dt)

    # Show all 7 sheet tabs without horizontal scrolling.
    # tabRatio=800 gives 80% to tabs (enough for 7 tab names) and
    # keeps the remaining 20% as a medium-sized horizontal scroll/resize bar.
    if wb.views:
        wb.views[0].tabRatio = 800

    # Enforce sheet order: Online Total, WB Total, WB Rejected Loads,
    #                      Anomaly and Not Considered, Pivot Table, Sorting, Charts
    desired_order = [
        'Online Total',
        'WB Total',
        'WB Rejected Loads',
        'Anomaly and Not Considered',
        'Pivot Table',
        'Sorting',
        'Charts',
    ]
    existing = wb.sheetnames
    ordered = [s for s in desired_order if s in existing]
    # Any extra sheets go at the end
    ordered += [s for s in existing if s not in ordered]
    wb._sheets.sort(key=lambda ws: ordered.index(ws.title) if ws.title in ordered else 99)

    # Save to bytes first
    buf = io.BytesIO()
    wb.save(buf)
    raw_bytes = buf.getvalue()

    # ── Patch chart XML: fix stale 'Online Data' sheet references ────────────
    # When the sheet is renamed from 'Online Data' to 'Online Total', openpyxl
    # does NOT update formula strings inside chart XML files.
    # Excel then reports a broken "Workbook Link" error and the chart goes blank.
    # Fix: rewrite the ZIP in-memory replacing all occurrences in chart XML files.
    import zipfile as _zf
    in_zip  = _zf.ZipFile(io.BytesIO(raw_bytes), 'r')
    out_buf = io.BytesIO()
    with _zf.ZipFile(out_buf, 'w', compression=_zf.ZIP_DEFLATED) as out_zip:
        for item in in_zip.infolist():
            data = in_zip.read(item.filename)
            # Patch chart XML files only
            if item.filename.startswith('xl/charts/') and item.filename.endswith('.xml'):
                text = data.decode('utf-8', errors='replace')
                # Replace old sheet name in formula references (both quoted and single-quoted)
                text = text.replace("'Online Data'!", "'Online Total'!")
                text = text.replace('"Online Data"!', '"Online Total"!')
                data = text.encode('utf-8')
            out_zip.writestr(item, data)
    in_zip.close()
    out_buf.seek(0)
    return out_buf.read()


def write_wb_phase_xlsx_into(wb_xls, df_valid, df_rejected):
    """Add WB Total + WB Rejected Loads sheets into an already-open workbook."""
    # Remove if they already exist (idempotent)
    for name in ['WB Total', 'WB Rejected Loads']:
        if name in wb_xls.sheetnames:
            del wb_xls[name]

    # FIX 1+3: Sort both WB dataframes strictly by WB In Time (A→Z)
    wb_in_col_sort = find_col(df_valid.columns, 'wb', 'in', 'time')
    if wb_in_col_sort:
        df_valid = df_valid.copy()
        import pandas as _pd
        df_valid[wb_in_col_sort] = _pd.to_datetime(df_valid[wb_in_col_sort], dayfirst=True, errors='coerce')
        df_valid = df_valid.sort_values(wb_in_col_sort, ascending=True, kind='stable').reset_index(drop=True)
    wb_in_col_rej = find_col(df_rejected.columns, 'wb', 'in', 'time')
    if wb_in_col_rej:
        df_rejected = df_rejected.copy()
        df_rejected[wb_in_col_rej] = _pd.to_datetime(df_rejected[wb_in_col_rej], dayfirst=True, errors='coerce')
        df_rejected = df_rejected.sort_values(wb_in_col_rej, ascending=True, kind='stable').reset_index(drop=True)

    # Create fresh WB sheets using the existing WB writer logic
    # We re-use write_wb_phase_xlsx but extract the sheet-writing part
    wb_in_col = find_col(df_valid.columns, 'wb', 'in', 'time')
    orig_cols = [c for c in df_valid.columns if not c.startswith('_')]

    if wb_in_col and wb_in_col in orig_cols:
        wb_in_pos = orig_cols.index(wb_in_col)
        out_cols  = (orig_cols[:wb_in_pos + 1]
                     + ['Format Hour', 'Peak Calculation']
                     + orig_cols[wb_in_pos + 1:])
    else:
        out_cols = orig_cols + ['Format Hour', 'Peak Calculation']

    wb_in_excel_ci = (out_cols.index(wb_in_col) + 1) if wb_in_col in out_cols else None
    wb_in_L = get_column_letter(wb_in_excel_ci) if wb_in_excel_ci else 'A'
    N_T = len(out_cols)

    # ── WB Total ──────────────────────────────────────────────────────────
    ws = wb_xls.create_sheet('WB Total')
    ws.sheet_format.defaultRowHeight = WB_ROW_HEIGHT
    ws.sheet_format.customHeight     = True
    _wb_set_col_widths(ws, out_cols,
                       {**WB_TOTAL_WIDTHS, 'Format Hour': 24.22, 'Peak Calculation': 24.22})
    _wb_header_row(ws, 1, out_cols)

    for ri, (_, row) in enumerate(df_valid.iterrows(), 2):
        ws.row_dimensions[ri].height = WB_ROW_HEIGHT
        is_yellow = bool(row.get('_yellow', False))
        red_cols  = row.get('_red_cols', []) or []
        row_bg    = 'FFFF00' if is_yellow else 'FFFFFF'

        for ci, col_name in enumerate(out_cols, 1):
            cell = ws.cell(row=ri, column=ci)
            if col_name == 'Format Hour':
                val = (f'=TEXT({wb_in_L}{ri},"hh")' if ri == 2 else
                       f'=IF(TEXT({wb_in_L}{ri},"hh")=TEXT({wb_in_L}{ri-1},"hh"),"",TEXT({wb_in_L}{ri},"hh"))')
                _wb_cell(cell, val, bg=row_bg)
            elif col_name == 'Peak Calculation':
                _wb_cell(cell, f'=VALUE(TEXT({wb_in_L}{ri},"hhmm"))', bg=row_bg)
            else:
                raw = row[col_name] if col_name in row.index else ''
                if not isinstance(raw, str) and pd.isna(raw): raw = ''
                _wb_cell(cell, raw, bg=row_bg, red_text=(col_name in red_cols))

    # ── WB Rejected Loads ─────────────────────────────────────────────────
    ws_r = wb_xls.create_sheet('WB Rejected Loads')
    ws_r.sheet_format.defaultRowHeight = WB_ROW_HEIGHT
    ws_r.sheet_format.customHeight     = True
    rej_cols = [c for c in df_rejected.columns if not c.startswith('_')]
    _wb_set_col_widths(ws_r, rej_cols, WB_REJ_WIDTHS)

    ws_r.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(rej_cols),1))
    hcell = ws_r.cell(row=1, column=1)
    hcell.value     = 'WEIGHBRIDGE REJECTED AND RECLASSIFIED LOADS'
    hcell.font      = Font(name='Arial', size=26, bold=True, color='FF0000')
    hcell.fill      = PatternFill('solid', fgColor='FFFFFF')
    hcell.alignment = Alignment(horizontal='left', vertical='center')
    hcell.border    = WB_BLACK_BDR
    ws_r.row_dimensions[1].height = WB_HDR_HEIGHT
    _wb_header_row(ws_r, 2, rej_cols)

    for ri, (_, row) in enumerate(df_rejected.iterrows(), 3):
        ws_r.row_dimensions[ri].height = WB_ROW_HEIGHT
        for ci, col_name in enumerate(rej_cols, 1):
            raw = row[col_name] if col_name in row.index else ''
            if not isinstance(raw, str) and pd.isna(raw): raw = ''
            _wb_cell(ws_r.cell(row=ri, column=ci), raw)


# ═════════════════════════════════════════════════════════════════════════════
#  /fullreport — Combined 7-sheet report route
# ═════════════════════════════════════════════════════════════════════════════

_fullreport_store = {'xlsx': None}


@app.route('/fullreport')
def fullreport_ui():
    return app.send_static_file('fullreport.html')


@app.route('/fullreport/generate', methods=['POST', 'OPTIONS'])
def fullreport_generate():
    """
    Generate the complete 7-sheet report using data already in the stores
    from Phase 2 and WB processing.

    Requires:
      - _phase2_store['online']       — clean Online df
      - _phase2_store['anomaly']      — anomaly df
      - _phase2_store['failure_list'] — precomputed failure list
      - _phase1_store['start_dt']     — filter start datetime
      - _phase1_store['end_dt']       — filter end datetime
      - _wb_store['total']            — WB valid df
      - _wb_store['rejected']         — WB rejected df
    """
    global _fullreport_store, _phase2_store, _phase1_store, _wb_store

    try:
        df_online   = _phase2_store.get('online')
        df_anomaly  = _phase2_store.get('anomaly')
        failure_list = _phase2_store.get('failure_list', [])
        df_wb_total  = _wb_store.get('total')
        df_wb_rej    = _wb_store.get('rejected')
        start_dt     = _phase1_store.get('start_dt')
        end_dt       = _phase1_store.get('end_dt')

        missing = []
        if df_online  is None: missing.append('Phase 2 CT data (run Phase 2 first)')
        if df_anomaly is None: missing.append('Anomaly data (run Phase 2 first)')
        if df_wb_total is None: missing.append('WB Total data (run WB Upload first)')
        if df_wb_rej  is None: missing.append('WB Rejected data (run WB Upload first)')
        if missing:
            return jsonify({'error': 'Missing data: ' + ' | '.join(missing)}), 400

        # If WB has not been resolved yet, run resolve with empty decisions
        if _wb_store.get('xlsx') is None and _wb_store.get('pending') is not None:
            df_wb_total, df_wb_rej = apply_wb_pending_decisions(
                df_wb_total, df_wb_rej, _wb_store['pending'], [])

        xlsx_bytes = build_full_report(
            df_online, df_anomaly,
            df_wb_total, df_wb_rej,
            failure_list=failure_list,
            report_dt=datetime.now(),
            start_dt=start_dt,
            end_dt=end_dt,
        )
        _fullreport_store['xlsx'] = xlsx_bytes

        return jsonify({
            'ok': True,
            'size_kb': round(len(xlsx_bytes) / 1024),
        })

    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@app.route('/fullreport/download')
def fullreport_download():
    global _fullreport_store
    xlsx = _fullreport_store.get('xlsx')
    if not xlsx:
        return jsonify({'error': 'No report generated yet.'}), 404
    resp = make_response(xlsx)
    resp.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = 'attachment; filename="Full_Report_7Sheets.xlsx"'
    return resp
