"""
Hourly Trucks Quantity Report — v3 with all corrections.

Corrections applied:
  Sheet 1 (Hourly Report):
    - Header BG = D9A49A (Tan, Darker 10%) for SC & GE Time/Trucks rows
    - Col A=9.89, B=4.89, C=2.22, D=5.11, E/F/G=6.67, H-Z=6.56, AA=9.89
    - B+C merged for rows 6+7 (and equivalent rows in GE/Total blocks)
    - Wrap Text ON, H=Center, V=Center everywhere in table area

  Sheet 2 (Charts — renamed from Histogram Chart):
    - Sheet name = "Charts"
    - Remove: working hours text, extra yellow box
    - Keep: yellow summary box (Soft Clay stats + time-split counts)
    - Single outer border around entire content area
    - Layout matches reference image
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils import get_column_letter

# ── Colour constants ──────────────────────────────────────────────────────────
TAN_BG  = 'D9A49A'   # Tan Background 2, Darker 10%
YEL_BG  = 'FFFF00'   # Yellow — TOTAL section & summary box
WHT     = 'FFFFFF'
BLK     = '000000'
GRN     = '00B050'   # green bars  — IN
DKRED   = '7B0000'   # dark maroon — OUT

def _fill(h):
    return PatternFill('solid', fgColor=h)

_thin_s = Side(style='thin',   color=BLK)
_med_s  = Side(style='medium', color=BLK)
_no_s   = Side(style=None)

def _bdr(l='thin', r='thin', t='thin', b='thin'):
    def _s(x):
        if   x == 'thin':   return _thin_s
        elif x == 'medium': return _med_s
        else:               return _no_s
    return Border(left=_s(l), right=_s(r), top=_s(t), bottom=_s(b))

ALL_BDR = _bdr()
NO_BDR  = Border()
BOT_BDR = Border(bottom=_thin_s)
MED_ALL = _bdr('medium','medium','medium','medium')

def _w(cell, val, bold=False, sz=11, h='center', v='center',
       fill=None, bdr=ALL_BDR, ul=None, color=BLK, wrap=True):
    cell.value = val
    cell.font  = Font(name='Calibri', bold=bold, size=sz,
                      underline=ul, color=color)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if fill is not None:
        cell.fill = fill
    if bdr is not None:
        cell.border = bdr


# ── Time slot definitions ─────────────────────────────────────────────────────
HOUR_SLOTS = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,0,1,2,3,4]

# Labels use newline for wrap within the merged time-header cell
TBL_LABELS = [
    '0700\n-\n0800', '0800\n-\n0900', '0900\n-\n1000', '1000\n-\n1100',
    '1100\n-\n1200', '1200\n-\n1300', '1300\n-\n1400', '1400\n-\n1500',
    '1500\n-\n1600', '1600\n-\n1700', '1700\n-\n1800', '1800\n-\n1900',
    '1900\n-\n2000', '2000\n-\n2100', '2100\n-\n2200', '2200\n-\n2300',
    '2300\n-\n0000', '0000\n-\n0100', '0100\n-\n0200', '0200\n-\n0300',
    '0300\n-\n0400', '0400\n-\n0500',
]
LATE_SC = '0500\n-\n1100'
LATE_GE = '0500\n-\n0600'

DATA_COLS = [get_column_letter(c) for c in range(4, 27)]   # D..Z (23 cols)
TOT_COL   = 'AA'

CHART_LABELS = [
    '0700','0800','0900','1000','1100','1200','1300','1400',
    '1500','1600','1700','1800','1900','2000','2100','2200',
    '2300','0000','0100','0200','0300','0400',
]


# ── Data computation ──────────────────────────────────────────────────────────
def _count_hourly(df, mat, hour_col):
    """Return 23 counts: 22 regular slots + 1 late slot (hour 5)."""
    sub    = df[df['_mat'] == mat.upper()]
    counts = [int((sub[hour_col] == h).sum()) for h in HOUR_SLOTS]
    late   = int(((sub[hour_col] >= 5) & (sub[hour_col] < 7)).sum())
    counts.append(late)
    return counts   # length 23


# ── Column widths per spec ────────────────────────────────────────────────────
def _set_col_widths(ws):
    ws.column_dimensions['A'].width  = 9.89
    ws.column_dimensions['B'].width  = 4.89
    ws.column_dimensions['C'].width  = 2.22
    ws.column_dimensions['D'].width  = 5.11
    for col in ['E', 'F', 'G']:
        ws.column_dimensions[col].width = 6.67
    for c in range(8, 27):          # H..Z
        ws.column_dimensions[get_column_letter(c)].width = 6.56
    ws.column_dimensions[TOT_COL].width = 9.89


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — HOURLY REPORT
# ═════════════════════════════════════════════════════════════════════════════

def _write_mat_block(ws, mat_label, in_vals, out_vals, late_label,
                     r_lbl, r_hdr1, r_hdr2, r_in, r_out, r_gap, hdr_bg):
    """
    Write one material block (Soft Clay or Good Earth).
    r_lbl  : material name row
    r_hdr1 : first header row — B:C = 'Time', D:Z+AA = time labels (merged hdr1:hdr2)
    r_hdr2 : second header row — B:C = 'Trucks'
    r_in   : IN data row
    r_out  : OUT data row
    r_gap  : spacer row (or None)
    """
    fh = _fill(hdr_bg)
    fw = _fill(WHT)

    ws.row_dimensions[r_lbl].height  = 18.75
    ws.row_dimensions[r_hdr1].height = 42.00   # tall for wrapped time text
    ws.row_dimensions[r_hdr2].height = 20.00
    ws.row_dimensions[r_in].height   = 49.90
    ws.row_dimensions[r_out].height  = 49.90
    if r_gap:
        ws.row_dimensions[r_gap].height = 30.00

    # Col A = material label; B:E = empty spacer merged
    _w(ws[f'A{r_lbl}'], mat_label, bold=True, sz=14, h='left', fill=fw, bdr=BOT_BDR)
    ws.merge_cells(f'B{r_lbl}:E{r_lbl}')
    _w(ws[f'B{r_lbl}'], '', fill=fw, bdr=NO_BDR)

    # B+C merged across hdr1 and hdr2 (spec: B & C rows 6 & 7 merged)
    ws.merge_cells(f'B{r_hdr1}:C{r_hdr1}')
    ws.merge_cells(f'B{r_hdr2}:C{r_hdr2}')
    ws.merge_cells(f'B{r_in}:C{r_in}')
    ws.merge_cells(f'B{r_out}:C{r_out}')

    _w(ws[f'B{r_hdr1}'], 'Time',   bold=True, sz=11, h='center', v='center', fill=fh, wrap=True)
    _w(ws[f'B{r_hdr2}'], 'Trucks', bold=True, sz=11, h='center', v='center', fill=fh, wrap=True)
    _w(ws[f'B{r_in}'],   'IN',     bold=True, sz=11, h='center', v='center', fill=fw, wrap=True)
    _w(ws[f'B{r_out}'],  'OUT',    bold=True, sz=11, h='center', v='center', fill=fw, wrap=True)

    all_labels = TBL_LABELS + [late_label]   # 23 entries

    for ci, (col, lbl) in enumerate(zip(DATA_COLS, all_labels)):
        # Each data column header = hdr1 merged with hdr2
        ws.merge_cells(f'{col}{r_hdr1}:{col}{r_hdr2}')
        _w(ws[f'{col}{r_hdr1}'], lbl, bold=True, sz=9,
           h='center', v='center', fill=fh, wrap=True)
        _w(ws[f'{col}{r_in}'],  in_vals[ci],  bold=True, sz=14,
           h='center', v='center', fill=fw, wrap=True)
        _w(ws[f'{col}{r_out}'], out_vals[ci], bold=True, sz=14,
           h='center', v='center', fill=fw, wrap=True)

    # TOTAL column — same pattern
    ws.merge_cells(f'{TOT_COL}{r_hdr1}:{TOT_COL}{r_hdr2}')
    _w(ws[f'{TOT_COL}{r_hdr1}'], 'TOTAL', bold=True, sz=11,
       h='center', v='center', fill=fh, wrap=True)
    _w(ws[f'{TOT_COL}{r_in}'],  f'=SUM(D{r_in}:Z{r_in})',
       bold=True, sz=14, h='center', v='center', fill=fw)
    _w(ws[f'{TOT_COL}{r_out}'], f'=SUM(D{r_out}:Z{r_out})',
       bold=True, sz=14, h='center', v='center', fill=fw)


def _write_total_block(ws, r_lbl, r_hdr1, r_hdr2, r_in, r_out,
                       sc_in_row, sc_out_row, ge_in_row, ge_out_row):
    """Combined TOTAL block — yellow background + formula references."""
    fy = _fill(YEL_BG)
    fw = _fill(WHT)

    ws.row_dimensions[r_lbl].height  = 18.75
    ws.row_dimensions[r_hdr1].height = 42.00
    ws.row_dimensions[r_hdr2].height = 20.00
    ws.row_dimensions[r_in].height   = 49.90
    ws.row_dimensions[r_out].height  = 49.90

    _w(ws[f'A{r_lbl}'], 'TOTAL OF SOFT CLAY & GOOD EARTH',
       bold=True, sz=14, h='left', fill=fw, bdr=BOT_BDR)
    ws.merge_cells(f'B{r_lbl}:E{r_lbl}')
    _w(ws[f'B{r_lbl}'], '', fill=fw, bdr=NO_BDR)

    ws.merge_cells(f'B{r_hdr1}:C{r_hdr1}')
    ws.merge_cells(f'B{r_hdr2}:C{r_hdr2}')
    ws.merge_cells(f'B{r_in}:C{r_in}')
    ws.merge_cells(f'B{r_out}:C{r_out}')

    _w(ws[f'B{r_hdr1}'], 'Time',   bold=True, sz=11, h='center', v='center', fill=fy, wrap=True)
    _w(ws[f'B{r_hdr2}'], 'Trucks', bold=True, sz=11, h='center', v='center', fill=fy, wrap=True)
    _w(ws[f'B{r_in}'],   'IN',     bold=True, sz=11, h='center', v='center', fill=fw, wrap=True)
    _w(ws[f'B{r_out}'],  'OUT',    bold=True, sz=11, h='center', v='center', fill=fw, wrap=True)

    all_labels = TBL_LABELS + [LATE_SC]
    for ci, (col, lbl) in enumerate(zip(DATA_COLS, all_labels)):
        ws.merge_cells(f'{col}{r_hdr1}:{col}{r_hdr2}')
        _w(ws[f'{col}{r_hdr1}'], lbl, bold=True, sz=9,
           h='center', v='center', fill=fy, wrap=True)
        _w(ws[f'{col}{r_in}'],
           f'={col}{sc_in_row}+{col}{ge_in_row}',
           bold=True, sz=14, h='center', v='center', fill=fy, wrap=True)
        _w(ws[f'{col}{r_out}'],
           f'={col}{sc_out_row}+{col}{ge_out_row}',
           bold=True, sz=14, h='center', v='center', fill=fy, wrap=True)

    ws.merge_cells(f'{TOT_COL}{r_hdr1}:{TOT_COL}{r_hdr2}')
    _w(ws[f'{TOT_COL}{r_hdr1}'], 'TOTAL', bold=True, sz=11,
       h='center', v='center', fill=fy, wrap=True)
    _w(ws[f'{TOT_COL}{r_in}'],  f'=SUM(D{r_in}:Z{r_in})',
       bold=True, sz=14, h='center', v='center', fill=fy)
    _w(ws[f'{TOT_COL}{r_out}'], f'=SUM(D{r_out}:Z{r_out})',
       bold=True, sz=14, h='center', v='center', fill=fy)


def build_summary_sheet(wb, df, report_dates, proj_name, contractor):
    ws = wb.create_sheet('Hourly Report')
    _set_col_widths(ws)
    ws.sheet_view.showGridLines = False

    for r in [1, 2, 3]:
        ws.row_dimensions[r].height = 25.15
    ws.row_dimensions[4].height = 30.00

    # ── Header rows 1–3 ───────────────────────────────────────────────────────
    _w(ws['A1'], f'PROJECT NAME:  {proj_name}',
       bold=True, sz=14, h='left', bdr=NO_BDR, wrap=False)
    _w(ws['A2'], f'CONTRACTOR NAME: {contractor}',
       bold=True, sz=14, h='left', bdr=NO_BDR, wrap=False)
    _w(ws['A3'], 'HOURLY TRUCKS QUANTITY REPORT',
       bold=True, sz=18, h='left', ul='single', bdr=NO_BDR, wrap=False)
    _w(ws['O3'], 'DATE :', bold=False, sz=14, bdr=NO_BDR, wrap=False)
    ws.merge_cells('P3:S3')
    _w(ws['P3'], report_dates, bold=True, sz=14,
       bdr=Border(bottom=_thin_s), wrap=False)

    # ── Compute data ──────────────────────────────────────────────────────────
    sc_in  = _count_hourly(df, 'SOFT CLAY',  '_in_hour')
    sc_out = _count_hourly(df, 'SOFT CLAY',  '_out_hour')
    ge_in  = _count_hourly(df, 'GOOD EARTH', '_in_hour')
    ge_out = _count_hourly(df, 'GOOD EARTH', '_out_hour')

    # ── Block 1: Soft Clay  rows 5–10 ────────────────────────────────────────
    _write_mat_block(ws, 'SOFT CLAY', sc_in, sc_out, LATE_SC,
                     r_lbl=5, r_hdr1=6, r_hdr2=7,
                     r_in=8, r_out=9, r_gap=10, hdr_bg=TAN_BG)

    # ── Block 2: Good Earth  rows 11–16 ──────────────────────────────────────
    _write_mat_block(ws, 'GOOD EARTH', ge_in, ge_out, LATE_GE,
                     r_lbl=11, r_hdr1=12, r_hdr2=13,
                     r_in=14, r_out=15, r_gap=16, hdr_bg=TAN_BG)

    # ── Block 3: Total  rows 17–21 ────────────────────────────────────────────
    _write_total_block(ws, r_lbl=17, r_hdr1=18, r_hdr2=19,
                       r_in=20, r_out=21,
                       sc_in_row=8, sc_out_row=9,
                       ge_in_row=14, ge_out_row=15)
    return ws


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — CHARTS
# ═════════════════════════════════════════════════════════════════════════════

def build_chart_sheet(wb, df, report_dates, proj_name, contractor):
    """
    Sheet named 'Charts':
    - Project/Contractor/Title header
    - Two time-split boxes (7AM-7PM | 7PM-5AM)
    - Yellow Soft Clay summary box (top right, inside outer border)
    - Clustered bar chart: IN=green, OUT=dark red
    - Single outer border around rows 1–48, cols A–Y
    - No working hours text, no extra yellow boxes
    """
    ws = wb.create_sheet('Charts')
    ws.sheet_view.showGridLines = False

    # Column widths for chart sheet
    ws.column_dimensions['A'].width = 1.5   # left margin (inside border)
    for c in range(2, 26):
        ws.column_dimensions[get_column_letter(c)].width = 4.2
    for c in range(26, 33):                 # yellow box columns Z..AF
        ws.column_dimensions[get_column_letter(c)].width = 5.8

    # Row heights
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 36
    ws.row_dimensions[7].height = 6
    for r in range(8, 49):
        ws.row_dimensions[r].height = 14
    ws.row_dimensions[48].height = 8   # bottom padding

    fw = _fill(WHT)
    fy = _fill(YEL_BG)

    # ── Compute values ────────────────────────────────────────────────────────
    sc          = df[df['_mat'] == 'SOFT CLAY']
    sc_in_vals  = [int((sc['_in_hour']  == h).sum()) for h in HOUR_SLOTS]
    sc_out_vals = [int((sc['_out_hour'] == h).sum()) for h in HOUR_SLOTS]
    sc_trucks   = len(sc)
    sc_net_wt   = round(sc['NET WEIGHT'].sum(), 2) if 'NET WEIGHT' in sc.columns else 0

    day_hrs   = list(range(7, 19))
    night_hrs = list(range(19, 24)) + list(range(0, 5))
    day_in    = int(sc['_in_hour'].isin(day_hrs).sum())
    night_in  = int(sc['_in_hour'].isin(night_hrs).sum())

    # ── Header ────────────────────────────────────────────────────────────────
    ws.merge_cells('B1:Y1')
    _w(ws['B1'], f'PROJECT NAME: {proj_name}',
       bold=True, sz=11, h='left', bdr=NO_BDR, wrap=False)
    ws.merge_cells('B2:Y2')
    _w(ws['B2'], f'CONTRACTOR NAME: {contractor}',
       bold=True, sz=11, h='left', bdr=NO_BDR, wrap=False)
    ws.merge_cells('B4:Y4')
    _w(ws['B4'],
       f'Daily Trucks Record IN/OUT Histogram Chart Dated {report_dates}',
       bold=True, sz=12, h='left', ul='single', bdr=NO_BDR, wrap=False)

    # ── Time-split boxes row 6 ────────────────────────────────────────────────
    ws.merge_cells('B6:L6')
    _w(ws['B6'], f'7am to 7pm (12hrs) = {day_in:,} Trucks',
       bold=True, sz=11, h='center', v='center', fill=fw,
       bdr=MED_ALL, wrap=False)

    ws.merge_cells('M6:Y6')
    _w(ws['M6'], f'7pm to 5am (8hrs) = {night_in:,} Trucks',
       bold=True, sz=11, h='center', v='center', fill=fw,
       bdr=MED_ALL, wrap=False)

    # ── Yellow summary box (cols Z:AF, rows 1–4) ──────────────────────────────
    ws.merge_cells('Z1:AF1')
    _w(ws['Z1'], 'Soft Clay', bold=True, sz=12, h='left', ul='single',
       fill=fy, bdr=_bdr('medium','medium','medium','thin'))

    ws.merge_cells('Z2:AF2')
    _w(ws['Z2'], '', fill=fy, bdr=_bdr('medium','medium','thin','thin'))

    ws.merge_cells('Z3:AF3')
    _w(ws['Z3'], f'Total No. of Trucks - {sc_trucks:,}',
       bold=True, sz=11, h='left', fill=fy,
       bdr=_bdr('medium','medium','thin','thin'), wrap=False)

    ws.merge_cells('Z4:AF4')
    _w(ws['Z4'], f'Total Net Weight (T) - {sc_net_wt:,.2f}',
       bold=True, sz=11, h='left', fill=fy,
       bdr=_bdr('medium','medium','thin','medium'), wrap=False)

    for r in range(5, 9):
        ws.merge_cells(f'Z{r}:AF{r}')
        _w(ws[f'Z{r}'], '', fill=fy, bdr=NO_BDR)

    # ── Data rows for chart (rows 44–46) ─────────────────────────────────────
    D_LBL, D_IN, D_OUT = 44, 45, 46
    for ci, col_n in enumerate(range(2, 24)):
        ws.cell(D_LBL, col_n).value = CHART_LABELS[ci]
        ws.cell(D_IN,  col_n).value = sc_in_vals[ci]
        ws.cell(D_OUT, col_n).value = sc_out_vals[ci]

    # ── Bar chart ─────────────────────────────────────────────────────────────
    bc = BarChart()
    bc.type     = 'col'
    bc.grouping = 'clustered'
    bc.title    = None
    bc.style    = 10
    bc.width    = 22
    bc.height   = 13
    bc.x_axis.title = None
    bc.y_axis.title = None
    bc.legend   = None

    in_ref = Reference(ws, min_col=2, max_col=23, min_row=D_IN,  max_row=D_IN)
    s_in   = Series(in_ref, title_from_data=False)
    s_in.graphicalProperties.solidFill = GRN
    s_in.graphicalProperties.line.solidFill = GRN
    bc.series.append(s_in)

    out_ref = Reference(ws, min_col=2, max_col=23, min_row=D_OUT, max_row=D_OUT)
    s_out   = Series(out_ref, title_from_data=False)
    s_out.graphicalProperties.solidFill = DKRED
    s_out.graphicalProperties.line.solidFill = DKRED
    bc.series.append(s_out)

    cat_ref = Reference(ws, min_col=2, max_col=23, min_row=D_LBL, max_row=D_LBL)
    bc.set_categories(cat_ref)
    ws.add_chart(bc, 'B8')

    # ── Outer border: rows 1–48, cols A(1)–Y(25) ─────────────────────────────
    LAST_ROW, LAST_COL = 48, 25
    for r in range(1, LAST_ROW + 1):
        for c in range(1, LAST_COL + 1):
            cell  = ws.cell(r, c)
            left  = _med_s if c == 1         else None
            right = _med_s if c == LAST_COL  else None
            top   = _med_s if r == 1         else None
            bot   = _med_s if r == LAST_ROW  else None
            if any([left, right, top, bot]):
                ex = cell.border
                cell.border = Border(
                    left   = left   or ex.left,
                    right  = right  or ex.right,
                    top    = top    or ex.top,
                    bottom = bot    or ex.bottom,
                )
    return ws


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def build_hourly_report(df_wb_raw, report_dates='',
                        proj_name='MANAGEMENT OF STAGING GROUND & INFILLING WORKS (PHASE 3)',
                        contractor='TOA - SAMSUNG C&T JOINT VENTURE'):
    """
    Build the complete Hourly Trucks Quantity Report (2 sheets).

    Args:
        df_wb_raw    : WB Total DataFrame (columns must include MATERIAL,
                       WB IN TIME, WB OUT TIME, NET WEIGHT)
        report_dates : e.g. '14 April 2026 & 15 April 2026'
        proj_name    : Project name string
        contractor   : Contractor name string

    Returns:
        bytes: xlsx file content ready to write to disk or send as download
    """
    df = df_wb_raw.copy()
    df['WB IN TIME']  = pd.to_datetime(df['WB IN TIME'],  dayfirst=True, errors='coerce')
    df['WB OUT TIME'] = pd.to_datetime(df['WB OUT TIME'], dayfirst=True, errors='coerce')
    df['_in_hour']  = df['WB IN TIME'].dt.hour
    df['_out_hour'] = df['WB OUT TIME'].dt.hour
    df['_mat']      = df['MATERIAL'].astype(str).str.strip().str.upper()

    if not report_dates:
        dates = df['WB IN TIME'].dropna()
        if len(dates):
            d1 = dates.min().strftime('%d %B %Y').replace(' 0', ' ')
            d2 = dates.max().strftime('%d %B %Y').replace(' 0', ' ')
            report_dates = f'{d1} & {d2}' if d1 != d2 else d1

    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, df, report_dates, proj_name, contractor)
    build_chart_sheet(wb, df, report_dates, proj_name, contractor)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Standalone run ────────────────────────────────────────────────────────────
if __name__ == '__main__':
    CY = '/mnt/user-data/uploads/Cycle_Time-_20260414_15042026.xlsx'
    df_wb = pd.read_excel(CY, sheet_name='WB Total')

    xlsx = build_hourly_report(df_wb, report_dates='14 April 2026 & 15 April 2026')

    with open('/mnt/user-data/outputs/Hourly_Trucks_Report.xlsx', 'wb') as f:
        f.write(xlsx)

    from openpyxl import load_workbook
    wbv = load_workbook(io.BytesIO(xlsx))
    print("Sheets:", wbv.sheetnames)

    ws1 = wbv['Hourly Report']
    print("\n-- Hourly Report --")
    print(f"Col A width : {ws1.column_dimensions['A'].width}")    # 9.89
    print(f"Col B width : {ws1.column_dimensions['B'].width}")    # 4.89
    print(f"Col C width : {ws1.column_dimensions['C'].width}")    # 2.22
    print(f"Col D width : {ws1.column_dimensions['D'].width}")    # 5.11
    print(f"Col H width : {ws1.column_dimensions['H'].width}")    # 6.56
    print(f"Col AA width: {ws1.column_dimensions['AA'].width}")   # 9.89
    print(f"D6 header bg: {ws1['D6'].fill.fgColor.rgb}")          # D9A49A
    print(f"D6 wrap_text: {ws1['D6'].alignment.wrap_text}")       # True
    print(f"B6:C6 merged: {'B6:C6' in [str(m) for m in ws1.merged_cells.ranges]}")
    print(f"B7:C7 merged: {'B7:C7' in [str(m) for m in ws1.merged_cells.ranges]}")
    print(f"D8 SC IN 7AM: {ws1['D8'].value}")    # 51
    print(f"AA8 formula : {ws1['AA8'].value}")
    print(f"D20 formula : {ws1['D20'].value}")   # =D8+D14

    ws2 = wbv['Charts']
    print("\n-- Charts --")
    print(f"Charts: {len(ws2._charts)} chart(s)")
    print(f"B1: {ws2['B1'].value!r}")
    print(f"B6 day box : {ws2['B6'].value!r}")
    print(f"M6 night box: {ws2['M6'].value!r}")
    print(f"Z1 yellow  : {ws2['Z1'].value!r}")
    print(f"Z3 trucks  : {ws2['Z3'].value!r}")
    print(f"Z4 weight  : {ws2['Z4'].value!r}")
    print(f"Generated  : {len(xlsx):,} bytes")
