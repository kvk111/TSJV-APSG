"""
report_engine.py — Complete Report Processing Engine

Handles all 3 report types:
  1. 7-sheet Main Excel  (build_main_report)
  2. 2-sheet Summary Excel  (build_summary_report)
  3. 2-slide PowerPoint  (build_ppt_report)

Data flow:
  CT files  → prepare_ct_data()   → df_ct_clean, df_anomaly, failure_list, exceedances, applicable_hours
  Online    → prepare_online_data() → df_wb_total, df_wb_rejected
  WB Total  → build_summary_report() → filled Demo.xlsx
  CT + WB   → build_ppt_report()  → 2-slide PPTX
"""

import io, os, re, zipfile, sys, tempfile
import pandas as pd
import numpy as np
from datetime import datetime

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.ticker import MultipleLocator
from matplotlib.dates import HourLocator, DateFormatter

from openpyxl import load_workbook

_HERE = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _fc(columns, *kws):
    """Find column whose lower-normalised name contains all keywords."""
    for c in columns:
        n = c.lower().replace(' ','').replace('_','').replace('/','').replace(':','')
        if all(k.lower().replace(' ','') in n for k in kws):
            return c
    return None


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1: PREPARE CT DATA
# ═══════════════════════════════════════════════════════════════════════════

def prepare_ct_data(ct_file_objects, start_dt, end_dt,
                    queue_minutes=0, lag_minutes=0, duration_threshold=120,
                    min_duration_threshold=5):
    """
    Read one or more CT Excel files, merge, filter by date range,
    then apply the 3-condition anomaly split using the UI-configured thresholds.

    Anomaly conditions (any one TRUE → row goes to Anomaly sheet):
      Condition 1: WB In Time – Date Time Arrival > queue_minutes
      Condition 2: WB Out Time – Date Time Exit   > lag_minutes
      Condition 3: Duration column from uploaded Excel > duration_threshold (upper-limit)
      Condition 4 (lower-limit): Duration == 0  OR  Duration < min_duration_threshold (default 5 min)
                   This is always applied regardless of other conditions.
                   Any duration < 5 minutes is always treated as anomaly by default.

    Returns:
        df_clean         : filtered clean CT DataFrame (non-anomaly rows)
        df_anomaly       : anomaly rows
        failure_list     : list of 'fail'/'ok' per row in df_clean
        exceedances      : total count of FAIL records
        applicable_hours : count of hours with any exceedance
    """
    parts = []
    for fo in ct_file_objects:
        try:
            df = pd.read_excel(fo, header=0)
        except Exception:
            fo.seek(0)
            df = pd.read_excel(fo, header=None)
            df.columns = df.iloc[0]; df = df.iloc[1:].reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        parts.append(df)

    df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

    # Parse datetime columns
    for col in ['WB In Time','WB Out Time','Date Time Arrival','Date Time Exit']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')

    # ── SORT by Date Time Arrival ascending BEFORE any formula/calculation ──
    # This ensures row order matches chronological order so rolling averages,
    # Format Hour formulas, and Peak Hour calculations are all correct.
    arr_col_sort = _fc(df.columns, 'date', 'time', 'arrival') or _fc(df.columns, 'wb', 'in', 'time')
    if arr_col_sort and arr_col_sort in df.columns:
        df = df.sort_values(arr_col_sort, ascending=True, kind='stable').reset_index(drop=True)

    # Derived columns
    wbin = _fc(df.columns,'wb','in','time')
    dur  = _fc(df.columns,'duration')

    if wbin:
        df['format hour'] = df[wbin].dt.hour.astype('float')

    # Normalise exceeded to numeric
    exc_col = _fc(df.columns,'exceeded')
    if exc_col:
        df['_exceeded_num'] = df[exc_col].astype(str).str.upper() \
            .map({'YES':1.0,'NG':1.0,'Y':1.0,'1':1.0}).fillna(0.0)
    else:
        df['_exceeded_num'] = 0.0

    # Normalise failure flag
    fail_col = _fc(df.columns,'failure')
    if fail_col:
        df['_is_fail'] = df[fail_col].astype(str).str.upper().isin(['FAIL','YES','Y','F'])
    else:
        df['_is_fail'] = False

    # Date filter
    if wbin and start_dt and end_dt:
        mask = (df[wbin] >= start_dt) & (df[wbin] <= end_dt)
        df = df[mask].reset_index(drop=True)

    # ── Full 3-condition anomaly split (using UI-configured thresholds) ──────
    arr_col    = _fc(df.columns, 'date', 'time', 'arrival') or _fc(df.columns, 'date', 'arrival')
    exit_col   = _fc(df.columns, 'date', 'time', 'exit')   or _fc(df.columns, 'date', 'exit')
    wb_in_col  = _fc(df.columns, 'wb', 'in', 'time')
    wb_out_col = _fc(df.columns, 'wb', 'out', 'time')

    for col in [arr_col, exit_col, wb_in_col, wb_out_col]:
        if col and col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')

    # Condition 1: WB In Time – Date Time Arrival > queue_minutes
    # SKIPPED ENTIRELY when queue_minutes == 0 (zero means "ignore this check")
    if queue_minutes > 0 and wb_in_col and arr_col:
        cond1 = (df[wb_in_col] - df[arr_col]).dt.total_seconds() / 60 > queue_minutes
    else:
        cond1 = pd.Series(False, index=df.index)

    # Condition 2: WB Out Time – Date Time Exit > lag_minutes
    # SKIPPED ENTIRELY when lag_minutes == 0 (zero means "ignore this check")
    if lag_minutes > 0 and wb_out_col and exit_col:
        cond2 = (df[wb_out_col] - df[exit_col]).dt.total_seconds() / 60 > lag_minutes
    else:
        cond2 = pd.Series(False, index=df.index)

    # Condition 3: Duration column from uploaded Excel > duration_threshold
    # Uses the 'Duration' column directly from the CT file — no datetime subtraction.
    # Falls back to computing Date Time Exit – Date Time Arrival only if Duration column absent.
    dur_col_for_cond3 = _fc(df.columns, 'duration')
    if dur_col_for_cond3:
        dur_vals_cond3 = pd.to_numeric(df[dur_col_for_cond3], errors='coerce').fillna(0)
        cond3 = dur_vals_cond3 > duration_threshold
    elif arr_col and exit_col:
        total_dur = (df[exit_col] - df[arr_col]).dt.total_seconds() / 60
        cond3 = total_dur > duration_threshold
    else:
        cond3 = pd.Series(False, index=df.index)

    # ── Condition 4 (lower-limit): Duration == 0  OR  Duration < min_duration_threshold ──
    # Default min_duration_threshold = 5 minutes.
    # Any record with duration == 0 or duration < 5 min is ALWAYS flagged as anomaly,
    # regardless of the user-defined upper-limit threshold (duration_threshold).
    # Missing WB In Time is also always treated as an anomaly.
    if dur:
        df[dur] = pd.to_numeric(df[dur], errors='coerce').fillna(0)
        # lower-limit: zero or below the minimum allowed duration
        cond4_lower = (df[dur] == 0) | (df[dur] < min_duration_threshold)
        # missing WB In Time also treated as anomaly
        missing_wbin = df[wbin].isna() if wbin else pd.Series(False, index=df.index)
        simple_anom = cond4_lower | missing_wbin
    else:
        # No duration column at all — treat missing WB In Time as anomaly
        simple_anom = df[wbin].isna() if wbin else pd.Series(False, index=df.index)

    anom_mask = cond1 | cond2 | cond3 | simple_anom

    df_anomaly = df[anom_mask].copy()
    df_clean   = df[~anom_mask].copy()

    # Failure list (5-consecutive logic matching ct_module/app.py)
    # Reuses 'ct_app_module' cache — same key as _get_ct_app() in __init__.py
    try:
        import importlib.util as _ilu
        _mod_name    = 'ct_app_module'
        _ct_app_path = os.path.join(_HERE, 'app.py')
        if _mod_name not in sys.modules:
            _spec = _ilu.spec_from_file_location(_mod_name, _ct_app_path)
            _mod  = _ilu.module_from_spec(_spec)
            sys.modules[_mod_name] = _mod
            _spec.loader.exec_module(_mod)
        _compute_failure_series = sys.modules[_mod_name]._compute_failure_series
        failure_list = list(_compute_failure_series(df_clean['_exceeded_num']))
    except Exception:
        failure_list = ['fail' if v else 'ok' for v in df_clean['_is_fail']]

    # FIX: match Online Total COUNTIF(Failure,"fail") — use failure_list not raw _is_fail
    exceedances = failure_list.count('fail')

    # Applicable hours: count distinct hours where rolling-10 Avg Cycle Time
    # STRICTLY exceeds the Peak Hours threshold — matching exactly the chart rule
    # (blue line > yellow line). Same formula as _render_chart_image.
    applicable_hours = 0
    if wbin and dur and len(df_clean) >= 10:
        try:
            _ah_df = df_clean[[wbin, dur]].copy()
            _ah_df[dur] = pd.to_numeric(_ah_df[dur], errors='coerce')
            _ah_df = _ah_df.dropna().sort_values(wbin).reset_index(drop=True)
            # Rolling 10-row mean (matches Excel Avg Cycle Time formula)
            _ah_df['_avg'] = _ah_df[dur].rolling(10).mean()
            # Peak threshold per row (matches Excel IF formula)
            _ah_df['_pcalc'] = _ah_df[wbin].dt.hour * 100 + _ah_df[wbin].dt.minute
            _ah_df['_peak']  = _ah_df['_pcalc'].apply(
                lambda p: 45 if (900 <= p < 1130) or (1500 <= p < 1730) else 30
            )
            # Only rows where avg is computed (row >= 10) and strictly > threshold
            _ah_df = _ah_df.dropna(subset=['_avg'])
            _ah_df['_hr'] = _ah_df[wbin].dt.floor('h')
            # For each hour: check if ANY 15-min window has avg > peak
            _exc_per_hr = _ah_df.groupby('_hr').apply(
                lambda g: (g['_avg'] > g['_peak']).any()
            )
            applicable_hours = int(_exc_per_hr.sum())
        except Exception:
            applicable_hours = 0

    return df_clean, df_anomaly, failure_list, exceedances, applicable_hours


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2: PREPARE ONLINE/WB DATA
# ═══════════════════════════════════════════════════════════════════════════

def prepare_online_data(online_file_object, start_dt, end_dt):
    """
    Read Online Data file (xlsx with .csv extension or real xlsx).
    Filter by date range. Split into accepted / rejected.

    Returns:
        df_wb_total    : ACCEPTED=YES records
        df_wb_rejected : ACCEPTED=NO records
        net_weight     : sum of NET WEIGHT column
    """
    fo = online_file_object
    try:
        xl = pd.ExcelFile(fo)
        df = pd.read_excel(fo, sheet_name=xl.sheet_names[0], header=0)
    except Exception:
        fo.seek(0)
        try: df = pd.read_csv(fo, encoding='utf-8', on_bad_lines='skip')
        except Exception: df = pd.read_csv(fo, encoding='latin-1', on_bad_lines='skip')

    df.columns = [str(c).strip().upper() for c in df.columns]

    wbin = _fc(df.columns,'wb','in','time')
    acc  = _fc(df.columns,'accepted')

    if wbin:
        df[wbin] = pd.to_datetime(df[wbin], dayfirst=True, errors='coerce')
        if start_dt and end_dt:
            mask = (df[wbin] >= start_dt) & (df[wbin] <= end_dt)
            df = df[mask].reset_index(drop=True)

    if acc:
        df_yes = df[df[acc].astype(str).str.upper() == 'YES'].reset_index(drop=True)
        df_no  = df[df[acc].astype(str).str.upper() == 'NO'].reset_index(drop=True)
    else:
        df_yes = df.copy()
        df_no  = pd.DataFrame(columns=df.columns)

    # FIX 1+3: Sort accepted and rejected strictly by WB IN TIME (A→Z)
    if wbin:
        df_yes = df_yes.sort_values(wbin, ascending=True, kind='stable').reset_index(drop=True)
        if len(df_no) and wbin in df_no.columns:
            df_no = df_no.sort_values(wbin, ascending=True, kind='stable').reset_index(drop=True)

    nw_col = _fc(df_yes.columns,'net','weight')
    net_weight = round(
        pd.to_numeric(df_yes[nw_col], errors='coerce').dropna().sum(), 2
    ) if nw_col else 0.0

    return df_yes, df_no, net_weight


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3a: BUILD 7-SHEET MAIN EXCEL — Original logic restored
# Uses build_full_report() from ct_module/app.py (loaded via cached importlib).
# The original function produces the exact 7-sheet output with:
#   1. Online Total  — all CT rows, formulas (Format Hour, Peak Calc, Avg CT, etc.)
#   2. WB Total      — WB accepted rows
#   3. WB Rejected Loads
#   4. Anomaly and Not Considered
#   5. Pivot Table   — exact cross-sheet COUNTIF/MIN/MAX/AVERAGE formulas
#   6. Sorting       — 23 hour-blocks × 8 sub-columns of cross-sheet IF formulas
#   7. Charts        — bar/line charts referencing Sorting and Online Total
# Do NOT replace this with simplified aggregate logic — the business output
# depends on the cross-sheet Excel formulas being present and correct.
# ═══════════════════════════════════════════════════════════════════════════

def _fmt_dt(v):
    """Format a cell value that might be a Timestamp, NaT, date, or string.
    NaT must be checked BEFORE isinstance(pd.Timestamp) because NaT is a
    subclass of Timestamp and calling .strftime() on NaT raises NaTType error.
    """
    try:
        if v is None or pd.isnull(v):
            return ''
    except Exception:
        pass
    if isinstance(v, pd.Timestamp):
        return v.strftime('%d-%m-%Y %H:%M')
    if hasattr(v, 'strftime'):
        return v.strftime('%d-%m-%Y %H:%M')
    return str(v) if v is not None else ''


def build_main_report(df_ct_clean, df_anomaly, df_wb_total, df_wb_rejected,
                      failure_list, start_dt, end_dt):
    """
    Build the 7-sheet CT Excel report using the original build_full_report()
    from ct_module/app.py.

    Loads ct_module/app.py by absolute file path under the stable module name
    'ct_app_module' (same key used by ct_module/__init__._get_ct_app()) so the
    3,500-line module is executed exactly ONCE and reused on every subsequent call.
    This avoids both the sys.modules name collision with root app.py AND the
    repeated expensive module execution that caused timeout on free tier.
    """
    import gc as _gc
    import importlib.util as _ilu

    print("[CT build] Loading original build_full_report from ct_module/app.py")

    _mod_name    = 'ct_app_module'          # matches _get_ct_app() in __init__.py
    _ct_app_path = os.path.join(_HERE, 'app.py')

    if _mod_name not in sys.modules:
        print("[CT build] First load — executing ct_module/app.py (once only)")
        _spec = _ilu.spec_from_file_location(_mod_name, _ct_app_path)
        _mod  = _ilu.module_from_spec(_spec)
        sys.modules[_mod_name] = _mod     # register BEFORE exec → avoids circular refs
        _spec.loader.exec_module(_mod)
        print("[CT build] ct_module/app.py loaded and cached")
    else:
        print("[CT build] ct_module/app.py already cached — reusing")

    _ct_app = sys.modules[_mod_name]

    # Extract the three functions we need (all defined in ct_module/app.py)
    build_full_report = _ct_app.build_full_report
    drop_unnamed      = _ct_app.drop_unnamed
    drop_computed     = _ct_app.drop_computed
    CT_COMPUTED       = _ct_app.CT_COMPUTED

    # ── Prepare dataframes — strip junk columns, sort chronologically ─────────
    df_c = drop_unnamed(df_ct_clean)
    try:
        df_c = drop_computed(df_c, CT_COMPUTED)
    except Exception:
        pass

    # Sort CT clean rows by WB In Time (chronological, ascending)
    wbin = _fc(df_c.columns, 'wb', 'in', 'time')
    if wbin and wbin in df_c.columns:
        df_c[wbin] = pd.to_datetime(df_c[wbin], dayfirst=True, errors='coerce')
        df_c = df_c.sort_values(wbin, ascending=True, kind='stable').reset_index(drop=True)
        failure_list = list(failure_list)   # ensure list copy after sort

    # Sort WB Total by WB In Time
    wbin_wb = _fc(df_wb_total.columns, 'wb', 'in', 'time')
    if wbin_wb and wbin_wb in df_wb_total.columns:
        df_wb_total[wbin_wb] = pd.to_datetime(df_wb_total[wbin_wb], dayfirst=True, errors='coerce')
        df_wb_total = df_wb_total.sort_values(wbin_wb, ascending=True, kind='stable').reset_index(drop=True)

    # Sort WB Rejected by WB In Time
    wbin_rj = _fc(df_wb_rejected.columns, 'wb', 'in', 'time')
    if wbin_rj and wbin_rj in df_wb_rejected.columns:
        df_wb_rejected[wbin_rj] = pd.to_datetime(df_wb_rejected[wbin_rj], dayfirst=True, errors='coerce')
        df_wb_rejected = df_wb_rejected.sort_values(wbin_rj, ascending=True, kind='stable').reset_index(drop=True)

    print(f"[CT build] Inputs: CT={len(df_c)} AN={len(df_anomaly)} WB={len(df_wb_total)} RJ={len(df_wb_rejected)}")
    print("[CT build] Calling build_full_report (original 7-sheet builder)...")

    result = build_full_report(
        df_c, df_anomaly, df_wb_total, df_wb_rejected,
        failure_list=failure_list,
        report_dt=datetime.now(),
        start_dt=start_dt,
        end_dt=end_dt,
    )

    # Free dataframes after build — memory cleanup for free tier
    df_c = df_anomaly = df_wb_total = df_wb_rejected = None
    build_full_report = drop_unnamed = drop_computed = None
    _gc.collect()

    print(f"[CT build] Complete. Size: {len(result)//1024} KB")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3b: BUILD SUMMARY EXCEL (fill Demo.xlsx template)
# ═══════════════════════════════════════════════════════════════════════════

def build_summary_report(df_wb_total, demo_template_path=None,
                         start_dt=None, end_dt=None):
    """Fill the Demo.xlsx template with WB data."""
    if demo_template_path is None:
        demo_template_path = os.path.join(_HERE, 'Demo.xlsx')

    sys.path.insert(0, _HERE)
    sys.path.insert(0, os.path.dirname(_HERE))
    from fill_demo import fill_demo_report

    return fill_demo_report(df_wb_total, template_path=demo_template_path,
                            start_dt=start_dt, end_dt=end_dt)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3c: BUILD POWERPOINT
# ═══════════════════════════════════════════════════════════════════════════

_DENSITY = 54043.31 / 37154.77   # T per m³ (from demo reference data)

def _esc(s):
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def _rt(xml, old, new):
    return xml.replace(f'<a:t>{old}</a:t>', f'<a:t>{new}</a:t>', 1)


def _prepare_ct_chart_data(df_ct):
    """
    Prepare CT chart data by calling the SAME function used for the online Excel chart.

    Delegates entirely to _get_ct_chart_data() in ct_module/app.py — the function
    that is ALSO called by _write_phase2_xlsx to build the online "Cycle Time
    Existence Chart". This guarantees the PPT chart uses IDENTICAL data, sort
    order, rolling mean, and peak hours logic as the online chart.
    """
    import math as _m
    # Load ct_app_module (cached after first load — same as build_main_report uses)
    _ct_app = sys.modules.get('ct_app_module')
    if _ct_app is None:
        import importlib.util as _ilu
        _spec = _ilu.spec_from_file_location('ct_app_module', os.path.join(_HERE, 'app.py'))
        _ct_app = _ilu.module_from_spec(_spec)
        sys.modules['ct_app_module'] = _ct_app
        _spec.loader.exec_module(_ct_app)

    # Call the SAME shared function used by the online Excel chart
    raw = _ct_app._get_ct_chart_data(df_ct)
    if raw is None:
        return None

    # Compute Y-axis scaling (same rules as Excel chart)
    data_max   = raw['dur_max']
    y_top      = max(70, int(_m.ceil(data_max / 10.0)) * 10)
    if y_top // 10 <= 10:
        y_interval = 10
    else:
        y_interval = 20
        if y_top % 20 != 0:
            y_top += 10
        while y_top // 20 < 7:
            y_top += 20

    return {
        'arr_col':    raw['arr_col'],
        'df':         raw['df'],
        'y_top':      y_top,
        'y_interval': y_interval,
        # Expose arrays directly for the plot
        'x_data':     raw['x_data'],
        'avg_ct':     raw['avg_ct'],
        'peak_hours': raw['peak_hours'],
    }


def _render_chart_image(df_ct, date_title, out_path):
    """
    Render the Cycle Time Exceedance chart for PPT Slide 1.

    Uses _prepare_ct_chart_data() — the SAME shared preparation function
    that mirrors the online Excel chart logic exactly.
    This guarantees: data, sorting, rolling mean, peak hours, and x-axis
    are all IDENTICAL between the PPT chart and the online Excel chart.
    """
    # Use the shared data preparation — single source of truth
    prepared = _prepare_ct_chart_data(df_ct)
    if prepared is None:
        _blank_img(out_path, 'No chart data')
        return

    df        = prepared['df']
    arr_col   = prepared['arr_col']
    y_top     = prepared['y_top']
    y_interval= prepared['y_interval']

    # Use arrays directly from the shared preparation function.
    # These are identical to what _write_phase2_xlsx uses for the online Excel chart.
    import numpy as _np
    import pandas as _pd
    t    = _pd.Series(_pd.to_datetime(prepared['x_data']))   # Date Time Arrival timestamps
    avg  = _pd.Series(prepared['avg_ct'])                     # rolling 10-mean (NaN rows 1-9)
    pk_v = _pd.Series(prepared['peak_hours'])                 # 45 or 30 per row

    # ── Plot: Fixed outer container 30.79 cm × 17.65 cm (spec) ──────────────
    # 30.79/2.54 = 12.12", 17.65/2.54 = 6.95"
    fig, ax = plt.subplots(figsize=(12.12, 6.95), facecolor='white')
    ax.set_facecolor('white')
    ax.set_ylim(0, y_top)
    ax.yaxis.set_major_locator(MultipleLocator(y_interval))
    ax.xaxis.set_major_locator(HourLocator(interval=1))
    ax.xaxis.set_major_formatter(DateFormatter('%H'))
    ax.tick_params(axis='x', labelsize=9, rotation=0, pad=4)
    ax.tick_params(axis='y', labelsize=9)
    ax.yaxis.grid(True, linestyle='-', color='#BBBBBB', linewidth=0.8)
    ax.xaxis.grid(False)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.tick_params(axis='y', which='both', length=0)
    ax.tick_params(axis='x', which='both', length=3)
    ax.set_title(date_title, fontsize=11, fontweight='bold', pad=8)

    # FIX 4: Yellow Peak Hours as LINE ONLY (no fill) — exact match to Excel chart
    ax.step(t, pk_v, where='post', color='#FFFF00', linewidth=2.5, zorder=2)
    # NO fill_between — removed to match Excel which shows only the yellow line

    # Blue Avg Cycle Time line — on top
    ax.plot(t, avg, color='#4472C4', linewidth=1.5, zorder=5)

    # Legend well below x-axis
    ax.legend(
        handles=[
            plt.Line2D([0],[0], color='#4472C4', linewidth=2, label='Avg Cycle Time'),
            plt.Line2D([0],[0], color='#FFFF00', linewidth=2.5, label='Peak Hours'),
        ],
        loc='lower center',
        bbox_to_anchor=(0.5, -0.18),   # pushed further down
        ncol=2,
        frameon=False,
        fontsize=10,
        handlelength=2.5,
    )

    plt.tight_layout(pad=0.5)
    # pad_inches=0.35 ensures legend doesn't clip at bottom edge
    plt.savefig(out_path, dpi=150, bbox_inches='tight',
                facecolor='white', pad_inches=0.35)
    plt.close(fig)

    # Flatten RGBA -> RGB for PPT compatibility
    try:
        from PIL import Image as _PILfix
        _im = _PILfix.open(out_path)
        if _im.mode != 'RGB':
            _bg = _PILfix.new('RGB', _im.size, (255,255,255))
            _bg.paste(_im, mask=_im.split()[3] if _im.mode=='RGBA' else None)
            _bg.save(out_path, 'PNG')
    except Exception:
        pass


def _render_slide2_image(summary_xlsx_bytes, out_path, _unused=None):
    """
    Render the Summary sheet (B1:AA21) as a full-width image for PPT Slide 2.

    Cross-platform approach (works on Windows, Mac, Linux):
    1. Try LibreOffice/pdftoppm (Linux/Mac) - best quality
    2. Fallback: matplotlib table render (Windows-safe, no external tools needed)
    """
    import subprocess, shutil, glob, tempfile, zipfile
    import numpy as _np

    try:
        from PIL import Image as _PILImg
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'Pillow', '-q'],
                       capture_output=True)
        from PIL import Image as _PILImg

    # ── Try LibreOffice approach first (Linux/Mac) ───────────────────────────
    _lo_available = False
    try:
        _r = subprocess.run(
            ['libreoffice', '--version'],
            capture_output=True, timeout=10
        )
        _lo_available = (_r.returncode == 0)
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
        _lo_available = False

    if _lo_available:
        try:
            _render_slide2_libreoffice(summary_xlsx_bytes, out_path)
            return
        except Exception as _le:
            pass  # Fall through to matplotlib approach

    # ── Fallback: matplotlib render (Windows-safe) ───────────────────────────
    _render_slide2_matplotlib(summary_xlsx_bytes, out_path)


def _render_slide2_libreoffice(summary_xlsx_bytes, out_path):
    """Render using LibreOffice + pdftoppm (Linux/Mac only)."""
    import subprocess, shutil, glob, tempfile, zipfile, re
    import numpy as _np
    from PIL import Image as _PILImg

    zf    = zipfile.ZipFile(io.BytesIO(summary_xlsx_bytes))
    _fmap = {_i.filename: zf.read(_i.filename) for _i in zf.infolist()}
    zf.close()

    for _k in list(_fmap.keys()):
        if 'chartsheets' in _k or (_k.startswith('xl/charts/') and 'chart1' in _k):
            del _fmap[_k]

    _wb_xml = _fmap.get('xl/workbook.xml', b'').decode('utf-8')
    _wb_xml = re.sub(r'<sheet[^>]+name="Chart"[^>]*/>', '', _wb_xml)
    _fmap['xl/workbook.xml'] = _wb_xml.encode('utf-8')

    _wr = _fmap.get('xl/_rels/workbook.xml.rels', b'').decode('utf-8')
    _wr = re.sub(r'<Relationship[^>]+chartsheet[^>]*/>', '', _wr)
    _fmap['xl/_rels/workbook.xml.rels'] = _wr.encode('utf-8')

    _ct = _fmap.get('[Content_Types].xml', b'').decode('utf-8')
    _ct = re.sub(r'<Override PartName="/xl/chartsheets[^"]+"[^/]*/>', '', _ct)
    _fmap['[Content_Types].xml'] = _ct.encode('utf-8')

    _sx = _fmap['xl/worksheets/sheet1.xml'].decode('utf-8')
    _sx = re.sub(r'<pageSetUp[^/]*/>', '', _sx)
    _sx = re.sub(r'<pageSetup[^/]*/>', '', _sx)
    _sx = re.sub(r'<pageMargins[^/]*/>', '', _sx)
    _ps = ('<pageMargins left="0.15" right="0.15" top="0.2" bottom="0.2"'
           ' header="0" footer="0"/>'
           '<pageSetup paperSize="8" orientation="landscape"'
           ' fitToPage="1" fitToWidth="1" fitToHeight="0"/>')
    _sx = _sx.replace('</worksheet>', _ps + '</worksheet>')
    if '<sheetPr' not in _sx:
        _sx = _sx.replace('<sheetData>',
                          '<sheetPr><pageSetUpPr fitToPage="1"/></sheetPr><sheetData>', 1)
    _fmap['xl/worksheets/sheet1.xml'] = _sx.encode('utf-8')

    _tmp = tempfile.mkdtemp(prefix='apsg_s2_')
    _xl  = os.path.join(_tmp, 'sum.xlsx')
    _pdf = os.path.join(_tmp, 'sum.pdf')

    _ob = io.BytesIO()
    with zipfile.ZipFile(_ob, 'w', compression=zipfile.ZIP_DEFLATED) as _oz:
        for _fn, _fb in _fmap.items():
            _oz.writestr(_fn, _fb)
    with open(_xl, 'wb') as _f:
        _f.write(_ob.getvalue())

    _r = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', _tmp, _xl],
        capture_output=True, text=True, timeout=60
    )
    if not os.path.exists(_pdf):
        raise RuntimeError('LibreOffice failed')

    _base = os.path.join(_tmp, 'pg')
    subprocess.run(['pdftoppm', '-r', '220', '-png', _pdf, _base],
                   capture_output=True, timeout=30)
    _pages = sorted(glob.glob(os.path.join(_tmp, 'pg-*.png')))
    if not _pages:
        raise RuntimeError('pdftoppm produced no pages')

    _strips = []
    _ref_w  = None
    for _fp in _pages:
        _arr = _np.array(_PILImg.open(_fp).convert('RGB'))
        _rc  = _np.where(_arr.min(axis=(1, 2)) < 245)[0]
        if len(_rc) == 0:
            continue
        _s = _arr[_rc[0]:_rc[-1] + 30, :]
        if _ref_w is None:
            _ref_w = _s.shape[1]
        elif _s.shape[1] != _ref_w:
            _hn = int(_s.shape[0] * _ref_w / _s.shape[1])
            _s  = _np.array(_PILImg.fromarray(_s).resize((_ref_w, _hn), _PILImg.LANCZOS))
        _strips.append(_s)

    if not _strips:
        raise RuntimeError('No content found')

    _GAP = 15
    _gap = _np.ones((_GAP, _ref_w, 3), dtype=_np.uint8) * 255
    _parts = []
    for _i, _s in enumerate(_strips):
        _parts.append(_s)
        if _i < len(_strips) - 1:
            _parts.append(_gap)
    _combined = _np.vstack(_parts)

    _B     = 25
    _res   = _PILImg.fromarray(_combined)
    _final = _PILImg.new('RGB', (_res.width + 2*_B, _res.height + 2*_B), 'white')
    _final.paste(_res, (_B, _B))
    _final.save(out_path, 'PNG', dpi=(220, 220))
    shutil.rmtree(_tmp, ignore_errors=True)


def _render_slide2_matplotlib(summary_xlsx_bytes, out_path):
    """
    Render the Summary sheet table as a tight PNG matching the PPT slide preview.
    Sized exactly to content — no blank whitespace — so the image fills Slide 2.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as _plt
    import matplotlib.patches as _mpa

    # ── Read data from xlsx ───────────────────────────────────────────────────
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(io.BytesIO(summary_xlsx_bytes))
        _ws = (_wb['Summary'] if 'Summary' in _wb.sheetnames else _wb.active)
        _date_val = str(_ws['P3'].value or '')
        _LCOLS = ['D','E','F','G','H','I','J','K','L','M','N','O',
                  'P','Q','R','S','T','U','V','W','X','Y','Z']
        def _rd(row):
            vals = []
            for L in _LCOLS:
                v = _ws[f'{L}{row}'].value
                try:    vals.append(int(float(v)) if v is not None else 0)
                except: vals.append(0)
            return vals
        _sc_in  = _rd(8);  _sc_out = _rd(9)
        _ge_in  = _rd(14); _ge_out = _rd(15)
        _tot_in  = [_sc_in[i]  + _ge_in[i]  for i in range(23)]
        _tot_out = [_sc_out[i] + _ge_out[i] for i in range(23)]
    except Exception:
        _date_val = ''
        _sc_in = _sc_out = _ge_in = _ge_out = _tot_in = _tot_out = [0]*23

    # ── Layout constants (all in inches) ─────────────────────────────────────
    _TAN = '#FDE9D9'
    _YEL = '#FFFF00'
    _WHT = '#FFFFFF'
    _BLK = '#000000'

    _TM   = 0.10   # top margin
    _BM   = 0.10   # bottom margin
    _LM   = 0.10   # left margin
    _H1   = 0.21   # header text row height
    _HG   = 0.08   # gap after last header line
    _HS   = 0.19   # section label height
    _HHR  = 0.52   # time-header row height
    _HDR  = 0.56   # IN / OUT data row height
    _GAP  = 0.15   # gap between blocks

    # Total figure height = all content + margins (no extra blank space)
    _FH = (_TM + 3*_H1 + _HG
           + 3*(_HS + _HHR + 2*_HDR + _GAP)
           - _GAP             # no trailing gap after last block
           + _BM)

    # Figure width — wide enough for 24 cols + label col
    _FW  = 13.60
    _RM  = _FW - 0.10

    # Column widths
    _LBL_W = 0.68
    _N     = 24     # 23 time slots + TOTAL
    _DATA_W = (_RM - _LM - _LBL_W) / _N

    def _xc(i):
        return _LM + _LBL_W + i * _DATA_W

    # ── 24 column headers ────────────────────────────────────────────────────
    _HDRS = [
        '0700-0800','0800-0900','0900-1000','1000-1100',
        '1100-1200','1200-1300','1300-1400','1400-1500',
        '1500-1600','1600-1700','1700-1800','1800-1900',
        '1900-2000','2000-2100','2100-2200','2200-2300',
        '2300-0000','0000-0100','0100-0200','0200-0300',
        '0300-0400','0400-0500','0500-0600','TOTAL',
    ]

    def _mk(v):
        row = list(v[:22]) + [v[22] if len(v) > 22 else 0, sum(v)]
        return [str(x) for x in row]

    # ── Figure setup ─────────────────────────────────────────────────────────
    _fig = _plt.figure(figsize=(_FW, _FH), dpi=150, facecolor='white')
    _fig.patch.set_facecolor('white')
    _ax  = _fig.add_axes([0, 0, 1, 1])
    _ax.set_xlim(0, _FW)
    _ax.set_ylim(0, _FH)
    _ax.invert_yaxis()
    _ax.set_facecolor('white')
    _ax.axis('off')

    # ── Drawing helpers ───────────────────────────────────────────────────────
    def _rect(x, y, w, h, fc=_WHT, ec=_BLK, lw=0.5, zorder=2):
        _ax.add_patch(_mpa.Rectangle(
            (x, y), w, h, facecolor=fc, edgecolor=ec,
            linewidth=lw, zorder=zorder, clip_on=False))

    def _txt(x, y, s, fs=7, fw='normal', ha='center', va='center', col=_BLK):
        _ax.text(x, y, str(s), fontsize=fs, fontweight=fw,
                 ha=ha, va=va, color=col, zorder=5,
                 multialignment='center', clip_on=False)

    # ── HEADER (3 lines + DATE box) ───────────────────────────────────────────
    _y = _TM
    _txt(_LM, _y + _H1/2,
         'Project Name:  MANAGEMENT OF STAGING GROUND & INFILLING WORKS (PHASE 3)',
         fs=8.5, fw='bold', ha='left')
    _y += _H1

    _txt(_LM, _y + _H1/2,
         'Contractor Name: TOA - SAMSUNG C&T JOINT VENTURE',
         fs=8.5, fw='bold', ha='left')
    _y += _H1

    _txt(_LM, _y + _H1/2, 'HOURLY TRUCKS QUANTITY REPORT',
         fs=9.5, fw='bold', ha='left')

    # DATE label + underlined box (right side of same row)
    _date_lbl_x = _FW * 0.46
    _date_box_x = _FW * 0.515
    _date_box_w = _FW * 0.33
    _date_box_h = _H1 * 0.80
    _txt(_date_lbl_x, _y + _H1/2, 'DATE :', fs=8, ha='left')
    _rect(_date_box_x, _y + _H1 * 0.10, _date_box_w, _date_box_h,
          fc=_WHT, lw=0.7)
    _txt(_date_box_x + _date_box_w/2, _y + _H1/2,
         _date_val, fs=8, fw='bold')
    _y += _H1 + _HG

    # ── Draw one material block ───────────────────────────────────────────────
    def _block(y0, label, in_v, out_v, hdr_clr, row_clr=None,
               hdr_font_size=5.0, hdr_font_size_total=5.8,
               hdr_diag_size=6):
        _y = y0
        # Section label
        _txt(_LM, _y + _HS/2, label, fs=8.5, fw='bold', ha='left')
        _y += _HS

        # Diagonal Time/Trucks header cell
        _rect(_LM, _y, _LBL_W, _HHR, fc=hdr_clr, lw=0.6)
        _ax.plot([_LM, _LM+_LBL_W], [_y, _y+_HHR],
                 color=_BLK, lw=0.5, zorder=6, clip_on=False)
        _txt(_LM + _LBL_W*0.72, _y + _HHR*0.25, 'Time',   fs=hdr_diag_size, fw='bold')
        _txt(_LM + _LBL_W*0.28, _y + _HHR*0.75, 'Trucks', fs=hdr_diag_size, fw='bold')

        # Time column headers
        for _i, _h in enumerate(_HDRS):
            _x = _xc(_i)
            _rect(_x, _y, _DATA_W, _HHR, fc=hdr_clr, lw=0.4)
            _fss = hdr_font_size if _i < 23 else hdr_font_size_total
            _txt(_x + _DATA_W/2, _y + _HHR/2, _h, fs=_fss, fw='bold')
        _y += _HHR

        # IN row
        _rect(_LM, _y, _LBL_W, _HDR, lw=0.5)
        _txt(_LM + _LBL_W/2, _y + _HDR/2, 'IN', fs=8.5, fw='bold')
        for _i, _v in enumerate(_mk(in_v)):
            _x = _xc(_i)
            _rect(_x, _y, _DATA_W, _HDR,
                  fc=row_clr if row_clr else _WHT, lw=0.4)
            # Load count values: bold + slightly larger (8.0 normal cols, 8.5 TOTAL)
            _txt(_x + _DATA_W/2, _y + _HDR/2, _v,
                 fs=8.0 if _i < 23 else 8.5,
                 fw='bold')
        _y += _HDR

        # OUT row
        _rect(_LM, _y, _LBL_W, _HDR, lw=0.5)
        _txt(_LM + _LBL_W/2, _y + _HDR/2, 'OUT', fs=8.5, fw='bold')
        for _i, _v in enumerate(_mk(out_v)):
            _x = _xc(_i)
            _rect(_x, _y, _DATA_W, _HDR,
                  fc=row_clr if row_clr else _WHT, lw=0.4)
            # Load count values: bold + slightly larger (8.0 normal cols, 8.5 TOTAL)
            _txt(_x + _DATA_W/2, _y + _HDR/2, _v,
                 fs=8.0 if _i < 23 else 8.5,
                 fw='bold')
        _y += _HDR
        return _y + _GAP

    # ── Color constants ───────────────────────────────────────────────────────
    # SC / GE header row: RGB(235,241,222) = #EBF1DE (light sage green)
    # TOTAL header row:   #FFFF00 (yellow) — unchanged per spec
    _SC_GE_HDR = '#EBF1DE'   # new: Soft Clay + Good Earth heading row

    # Draw 3 blocks
    # SC and GE: new #EBF1DE header color, slightly larger header fonts (5.5 / 6.2 / 6.5)
    _y = _block(_y, 'SOFT CLAY',  _sc_in,  _sc_out,
                 hdr_clr=_SC_GE_HDR,
                 hdr_font_size=5.5, hdr_font_size_total=6.2, hdr_diag_size=6.5)
    _y = _block(_y, 'GOOD EARTH', _ge_in,  _ge_out,
                 hdr_clr=_SC_GE_HDR,
                 hdr_font_size=5.5, hdr_font_size_total=6.2, hdr_diag_size=6.5)
    # TOTAL: yellow unchanged, original font sizes unchanged
    _block(_y, 'TOTAL OF SOFT CLAY & GOOD EARTH',
           _tot_in, _tot_out, hdr_clr=_YEL)   # row_clr omitted → data cells white

    # Save with no extra padding
    _plt.savefig(out_path, dpi=150, bbox_inches='tight',
                 facecolor='white', edgecolor='none', pad_inches=0.02)
    _plt.close(_fig)

    # Flatten RGBA -> RGB with white background for PowerPoint compatibility
    try:
        from PIL import Image as _PILfix
        _im = _PILfix.open(out_path)
        if _im.mode != 'RGB':
            _bg = _PILfix.new('RGB', _im.size, (255, 255, 255))
            if _im.mode == 'RGBA':
                _bg.paste(_im, mask=_im.split()[3])
            else:
                _bg.paste(_im)
            _bg.save(out_path, 'PNG')
    except Exception:
        pass
def _render_slide2_full(sc_in, sc_out, ge_in, ge_out,
                         tot_in, tot_out, date_str, out_path):
    """
    Render the complete Hourly Trucks Quantity Report as a full-slide image.
    Matches the demo slide 2 exactly: 3 blocks (SC / GE / Total), proper colors,
    column headers, row labels, totals.
    """
    # 24 columns: 22 hourly + 1 late + TOTAL
    COL_HDRS = [
        '0700-\n0800','0800-\n0900','0900-\n1000','1000-\n1100',
        '1100-\n1200','1200-\n1300','1300-\n1400','1400-\n1500',
        '1500-\n1600','1600-\n1700','1700-\n1800','1800-\n1900',
        '1900-\n2000','2000-\n2100','2100-\n2200','2200-\n2300',
        '2300-\n0000','0000-\n0100','0100-\n0200','0200-\n0300',
        '0300-\n0400','0400-\n0500','0500-\n1100','TOTAL',
    ]

    def _mk(v):
        row = list(v[:22]) + [v[22] if len(v) > 22 else 0, sum(v)]
        return row

    # ── Figure setup: A3 landscape ratio, white background ───────────────
    fig = plt.figure(figsize=(19.69, 13.9), dpi=120, facecolor='white')
    fig.patch.set_facecolor('white')

    # ── Header area ───────────────────────────────────────────────────────
    PROJ = 'MANAGEMENT OF STAGING GROUND & INFILLING WORKS (PHASE 3)'
    CONT = 'TOA - SAMSUNG C&T JOINT VENTURE'

    ax_hdr = fig.add_axes([0.01, 0.88, 0.98, 0.11])
    ax_hdr.set_facecolor('white'); ax_hdr.axis('off')
    ax_hdr.text(0.0, 0.90, f'Project Name:  {PROJ}',
                fontsize=9, fontweight='bold', va='top', ha='left',
                transform=ax_hdr.transAxes)
    ax_hdr.text(0.0, 0.60, f'Contractor Name: {CONT}',
                fontsize=9, fontweight='bold', va='top', ha='left',
                transform=ax_hdr.transAxes)
    ax_hdr.text(0.0, 0.28, 'HOURLY TRUCKS QUANTITY REPORT',
                fontsize=10.5, fontweight='bold', va='top', ha='left',
                transform=ax_hdr.transAxes,
)

    ax_hdr.text(0.56, 0.28, 'DATE :',
                fontsize=9, va='top', ha='left', transform=ax_hdr.transAxes)
    ax_hdr.text(0.63, 0.28, date_str,
                fontsize=9, fontweight='bold', va='top', ha='left',
                transform=ax_hdr.transAxes,
                bbox=dict(boxstyle='square,pad=0.3', facecolor='white',
                          edgecolor='black', linewidth=0.8))

    # ── Draw one material block ───────────────────────────────────────────
    def _draw_block(ax_title_y, ax_y, ax_h,
                    title, in_v, out_v, hdr_bg, row_bg=None):
        # Title label above the table
        ax_t = fig.add_axes([0.01, ax_title_y, 0.98, 0.03])
        ax_t.set_facecolor('white'); ax_t.axis('off')
        ax_t.text(0.005, 0.5, title,
                  fontsize=9, fontweight='bold', va='center', ha='left',
                  transform=ax_t.transAxes)

        sub = fig.add_axes([0.01, ax_y, 0.98, ax_h])
        sub.set_facecolor('white'); sub.axis('off')

        cell_text = [_mk(in_v), _mk(out_v)]
        n = len(COL_HDRS)

        tbl = sub.table(
            cellText=cell_text,
            rowLabels=['IN', 'OUT'],
            colLabels=COL_HDRS,
            loc='center',
            cellLoc='center',
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(6.8)
        tbl.scale(1, 2.8)

        # Style header row (row 0)
        for j in range(n):
            cell = tbl[0, j]
            cell.set_facecolor(hdr_bg)
            cell.set_text_props(fontsize=6.0, fontweight='bold')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.5)

        # Style data rows (rows 1, 2)
        for i in [1, 2]:
            for j in range(n):
                cell = tbl[i, j]
                if row_bg:
                    cell.set_facecolor(row_bg)
                else:
                    cell.set_facecolor('white')
                cell.set_edgecolor('black')
                cell.set_linewidth(0.4)
                cell.set_text_props(fontsize=7.0)

            # TOTAL column: bold, slightly different bg
            tbl[i, n-1].set_text_props(fontweight='bold', fontsize=7.5)
            tbl[i, n-1].set_facecolor('#D0D0D0' if not row_bg else '#E8E800')

            # Row label cell
            tbl[i, -1].set_text_props(fontweight='bold', fontsize=7.5)
            tbl[i, -1].set_facecolor(hdr_bg)
            tbl[i, -1].set_edgecolor('black')

    # Block positions (y from bottom in figure fraction)
    _draw_block(ax_title_y=0.84, ax_y=0.60, ax_h=0.23,
                title='SOFT CLAY',
                in_v=sc_in, out_v=sc_out, hdr_bg='#FDE9D9')

    _draw_block(ax_title_y=0.55, ax_y=0.31, ax_h=0.23,
                title='GOOD EARTH',
                in_v=ge_in, out_v=ge_out, hdr_bg='#FDE9D9')

    _draw_block(ax_title_y=0.26, ax_y=0.02, ax_h=0.23,
                title='TOTAL OF SOFT CLAY & GOOD EARTH',
                in_v=tot_in, out_v=tot_out,
                hdr_bg='#FFFF00', row_bg='#FFFF00')

    plt.savefig(out_path, dpi=130, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)


def _render_table_img(sc_in, sc_out, ge_in, ge_out, tot_in, tot_out, date_str, out_path):
    COLS = ['0700-0800','0800-0900','0900-1000','1000-1100','1100-1200',
            '1200-1300','1300-1400','1400-1500','1500-1600','1600-1700',
            '1700-1800','1800-1900','1900-2000','2000-2100','2100-2200',
            '2200-2300','2300-0000','0000-0100','0100-0200','0200-0300',
            '0300-0400','0400-0500','0500-1100','TOTAL']

    def _mk(v):
        return list(v[:22]) + [v[22] if len(v)>22 else 0, sum(v)]

    fig = plt.figure(figsize=(20,13), facecolor='white', dpi=110)
    ax_bg = fig.add_axes([0,0,1,1]); ax_bg.set_facecolor('white'); ax_bg.axis('off')

    ax_bg.text(.01,.975,'Project Name:  MANAGEMENT OF STAGING GROUND & INFILLING WORKS (PHASE 3)',
               fontsize=9,fontweight='bold',va='top',transform=ax_bg.transAxes)
    ax_bg.text(.01,.950,'Contractor Name: TOA - SAMSUNG C&T JOINT VENTURE',
               fontsize=9,fontweight='bold',va='top',transform=ax_bg.transAxes)
    ax_bg.text(.01,.925,'HOURLY TRUCKS QUANTITY REPORT',
               fontsize=10,fontweight='bold',va='top',transform=ax_bg.transAxes)
    ax_bg.text(.55,.925,f'DATE :  {date_str}',fontsize=9,fontweight='bold',va='top',
               transform=ax_bg.transAxes,
               bbox=dict(boxstyle='square,pad=.3',facecolor='white',edgecolor='black',lw=.8))

    for (y_t,y_a,h,title,iv,ov,hbg,rbg) in [
        (.88,.60,.26,'SOFT CLAY',              sc_in, sc_out,'#FDE9D9',None),
        (.57,.29,.26,'GOOD EARTH',             ge_in, ge_out,'#FDE9D9',None),
        (.265,.01,.25,'TOTAL OF SOFT CLAY & GOOD EARTH',tot_in,tot_out,'#FFFF00','#FFFF00'),
    ]:
        ax_bg.text(.01,y_t,title,fontsize=9,fontweight='bold',va='bottom',transform=ax_bg.transAxes)
        sub=fig.add_axes([.015,y_a,.97,h]); sub.axis('off')
        tbl=sub.table(cellText=[_mk(iv),_mk(ov)],rowLabels=['IN','OUT'],
                      colLabels=COLS,loc='center',cellLoc='center')
        tbl.auto_set_font_size(False); tbl.set_fontsize(6.5); tbl.scale(1,2.6)
        n=len(COLS)
        for j in range(n):
            tbl[0,j].set_facecolor(hbg)
            tbl[0,j].set_text_props(fontsize=5.5,fontweight='bold')
        for i in [1,2]:
            if rbg:
                for j in range(n): tbl[i,j].set_facecolor(rbg)
            tbl[i,n-1].set_text_props(fontweight='bold',fontsize=8)
            tbl[i,n-1].set_facecolor('#E0E0E0' if not rbg else '#F5F500')

    plt.savefig(out_path, dpi=110, bbox_inches='tight', facecolor='white')
    plt.close(fig)


def _blank_img(path, msg='No data'):
    fig,ax=plt.subplots(figsize=(16,9),facecolor='white')
    ax.text(.5,.5,msg,ha='center',va='center',fontsize=14,color='#888'); ax.axis('off')
    plt.savefig(path,dpi=100,facecolor='white',bbox_inches='tight'); plt.close(fig)


def build_ppt_report(df_ct_clean, df_wb_total,
                     reason, exceedances, applicable_hours,
                     summary_xlsx_bytes,
                     ppt_template_path=None):
    """
    Build the 2-slide PowerPoint.
    Slide 1: CT exceedance chart + metrics box + summary table
    Slide 2: Summary Excel table snapshot image
    """
    if ppt_template_path is None:
        ppt_template_path = os.path.join(_HERE, 'demo.pptx')

    # Compute dates and WB summary
    wbin_ct = _fc(df_ct_clean.columns,'wb','in')
    if wbin_ct and df_ct_clean[wbin_ct].dtype == 'object':
        df_ct_clean[wbin_ct] = pd.to_datetime(df_ct_clean[wbin_ct], dayfirst=True, errors='coerce')

    dates = df_ct_clean[wbin_ct].dropna() if wbin_ct else pd.Series(dtype='datetime64[ns]')
    d1 = dates.min() if len(dates) else None
    d2 = dates.max() if len(dates) else None
    date_chart = (f"Cycle Time Exceedance Chart on {d1.day} & {d2.day} {d2.strftime('%b %Y')}"
                  if d1 and d2 else "Cycle Time Exceedance Chart")
    date_table = (f"{d1.day} & {d2.day} {d2.strftime('%b %Y')}" if d1 and d2 else "")

    nw_col = _fc(df_wb_total.columns,'net','weight')
    tot_loads  = len(df_wb_total)
    tot_weight = round(pd.to_numeric(df_wb_total[nw_col], errors='coerce').dropna().sum(), 2) if nw_col else 0.0
    tot_vol    = round(tot_weight / _DENSITY, 2)

    # Render images
    chart_img  = os.path.join(tempfile.gettempdir(), '_apsg_chart.png')
    slide2_img = os.path.join(tempfile.gettempdir(), '_apsg_s2.png')
    _render_chart_image(df_ct_clean, date_chart, chart_img)

    # Save summary_xlsx_bytes to disk first so LibreOffice can open it fully
    # (the in-memory bytes may lack printerSettings that LibreOffice needs)
    _tmp_sum = os.path.join(tempfile.gettempdir(), '_apsg_summary_input.xlsx')
    with open(_tmp_sum, 'wb') as _tmpf:
        _tmpf.write(summary_xlsx_bytes)
    with open(_tmp_sum, 'rb') as _tmpf:
        _bytes_for_render = _tmpf.read()
    _render_slide2_image(_bytes_for_render, slide2_img)

    # Read template ZIP
    tmpl = zipfile.ZipFile(ppt_template_path, 'r')
    files = {item.filename: tmpl.read(item.filename) for item in tmpl.infolist()}
    tmpl.close()

    # Patch slide 1 XML
    s1 = files['ppt/slides/slide1.xml'].decode('utf-8')

    # Remove the pre-existing template reason text box (it will be replaced
    # by the newly injected one after chart insertion, below).
    import re as _re_inner
    s1 = _re_inner.sub(
        r'<p:sp>(?:(?!</p:sp>).)*?queue condition(?:(?!</p:sp>).)*?</p:sp>',
        '', s1, count=1, flags=re.DOTALL
    )

    # Exceedances
    s1 = re.sub(r'<a:t>Number of Exceedance: \d+</a:t>',
                f'<a:t>Number of Exceedance: {exceedances}</a:t>', s1, count=1)
    # Applicable hours
    s1 = re.sub(r'<a:t>Applicable Hours: \d+\t</a:t>',
                f'<a:t>Applicable Hours: {applicable_hours}\t</a:t>', s1, count=1)
    # Date in table
    s1 = re.sub(r'<a:t>\d+\s*(?:&amp;|&)\s*\d+\s+\w+\s+\d{4}</a:t>',
                f'<a:t>{_esc(date_table)}</a:t>', s1, count=1)

    # FIX 6.2: SC column = ALL combined totals; Total column = same combined values
    # Replace ALL occurrences of the demo values (SC col + Total col both get combined)
    for old_val, new_val in [
        ('3039',     str(tot_loads)),
        ('54043.31', f'{tot_weight:.2f}'),
        ('37154.77', f'{tot_vol:.2f}'),
    ]:
        s1 = s1.replace(f'<a:t>{old_val}</a:t>', f'<a:t>{new_val}</a:t>')

    # ── Replace native chart (rId3) with rendered PNG in slide 1 ───────────
    # The template has a native pptx chart (chart1.xml via rId3).
    # Replace it with our matplotlib PNG at the identical position,
    # so PPT chart is an exact visual replica of the Excel Online Total chart.
    with open(chart_img, 'rb') as _cf:
        _chart_png_bytes = _cf.read()

    files['ppt/media/chart1_img.png'] = _chart_png_bytes

    # Update rId3 in slide1 rels: chart type -> image type
    _s1r = files['ppt/slides/_rels/slide1.xml.rels'].decode('utf-8')
    _s1r = _s1r.replace(
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="../charts/chart1.xml"',
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/chart1_img.png"'
    )
    files['ppt/slides/_rels/slide1.xml.rels'] = _s1r.encode('utf-8')

    # Replace graphicFrame (native chart) with p:pic (image) at same EMU position
    _CX_OFF = 907586; _CY_OFF = 1179095; _CX = 10986428; _CY = 6376737
    _pic_xml = (
        '<p:pic>'
        '<p:nvPicPr>'
        '<p:cNvPr id="2" name="Chart 1"/>'
        '<p:cNvPicPr><a:picLocks noChangeAspect="1"/></p:cNvPicPr>'
        '<p:nvPr/>'
        '</p:nvPicPr>'
        '<p:blipFill>'
        '<a:blip r:embed="rId3" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/>'
        '<a:stretch><a:fillRect/></a:stretch>'
        '</p:blipFill>'
        '<p:spPr>'
        '<a:xfrm>'
        '<a:off x="' + str(_CX_OFF) + '" y="' + str(_CY_OFF) + '"/>'
        '<a:ext cx="' + str(_CX) + '" cy="' + str(_CY) + '"/>'
        '</a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        '</p:spPr>'
        '</p:pic>'
    )
    s1 = re.sub(
        r'<p:graphicFrame>(?:(?!</p:graphicFrame>).)*<c:chart(?:(?!</p:graphicFrame>).)*</p:graphicFrame>',
        _pic_xml, s1, count=1, flags=re.DOTALL
    )
    files.pop('ppt/charts/chart1.xml', None)
    files.pop('ppt/charts/_rels/chart1.xml.rels', None)

    # ── Inject reason text INSIDE the chart plot region — left side, vertically centered ──
    # Chart image on A3 slide (EMU): x=907586, y=1179095, w=10986428, h=6376737
    #
    # The matplotlib figure has internal margins (tight_layout):
    #   left  ~4.5% of chart width  = Y-axis labels area  → skip past this
    #   top   ~5.8% of chart height = title area           → skip past this
    #   bottom ~17.3% from chart bottom = X-axis + legend  → stay above this
    #
    # Plot region boundaries (EMU):
    #   plot_left   = chart_x + 4.5% * chart_w  = 1,401,975  (1.533in from slide left)
    #   plot_top    = chart_y + 5.8% * chart_h  = 1,549,011  (1.694in from slide top)
    #   plot_bottom = chart_y + 82.7% * chart_h = 6,450,494  (7.057in from slide top)
    #
    # Text box: LEFT side of plot, VERTICALLY CENTERED in the plot region.
    #   - x: 0.05in past plot_left  (just inside plot area, past Y-axis number labels)
    #   - y: vertical center of plot region minus half box height
    #   - w: 5.5in wide — auto-wraps to 2 lines for long sentences
    #   - h: 0.75in tall — fits 2 lines at 12pt
    # All four edges verified inside plot boundary.
    _EMU_unit      = 914400
    _chart_x       = 907586
    _chart_y       = 1179095
    _chart_w       = 10986428
    _chart_h       = 6376737
    _plot_left_emu = _chart_x + int(0.045 * _chart_w)   # 1,401,975 EMU
    _plot_top_emu  = _chart_y + int(0.058 * _chart_h)   # 1,548,911 EMU
    _plot_bot_emu  = _chart_y + int(0.827 * _chart_h)   # 6,450,494 EMU
    _plot_ctr_y    = (_plot_top_emu + _plot_bot_emu) // 2
    _txt_box_w     = int(5.5  * _EMU_unit)              # 5.5in wide — 2-line wrap
    _txt_box_h     = int(0.75 * _EMU_unit)              # 0.75in tall — 2 lines at 12pt
    _txt_box_x     = _plot_left_emu + int(0.05 * _EMU_unit)  # 0.05in past plot left
    _txt_box_y     = _plot_ctr_y - _txt_box_h // 2           # vertically centered in plot
    _reason_esc = _esc(str(reason))
    _reason_sp = (
        '<p:sp>'
        '<p:nvSpPr>'
        '<p:cNvPr id="99" name="ReasonText"/>'
        '<p:cNvSpPr><a:spLocks noGrp="1"/></p:cNvSpPr>'
        '<p:nvPr/>'
        '</p:nvSpPr>'
        '<p:spPr>'
        '<a:xfrm>'
        f'<a:off x="{_txt_box_x}" y="{_txt_box_y}"/>'
        f'<a:ext cx="{_txt_box_w}" cy="{_txt_box_h}"/>'
        '</a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        '<a:noFill/>'
        '</p:spPr>'
        '<p:txBody>'
        '<a:bodyPr wrap="square" lIns="45720" rIns="45720" tIns="36576" bIns="36576" anchor="t"/>'
        '<a:lstStyle/>'
        '<a:p>'
        '<a:pPr algn="l"/>'
        '<a:r>'
        '<a:rPr lang="en-US" sz="1200" b="0" dirty="0">'
        '<a:solidFill><a:srgbClr val="1F2D3D"/></a:solidFill>'
        '</a:rPr>'
        f'<a:t>{_reason_esc}</a:t>'
        '</a:r>'
        '</a:p>'
        '</p:txBody>'
        '</p:sp>'
    )
    # Insert this shape just before </p:spTree> — renders on top of the chart image
    s1 = s1.replace('</p:spTree>', _reason_sp + '</p:spTree>', 1)

    files['ppt/slides/slide1.xml'] = s1.encode('utf-8')

    # ── Replace slide 2 image and resize to fill slide ──────────────────────
    with open(slide2_img,'rb') as f: s2_img_bytes = f.read()

    # FIX: template uses image5.emf; replace it with our PNG and update rels
    # Remove old EMF, add PNG under same logical slot
    files.pop('ppt/media/image5.emf', None)
    files['ppt/media/image5.png'] = s2_img_bytes

    # FIX: Update slide2 rels so rId2 points to the PNG (not the old EMF)
    s2_rels = files['ppt/slides/_rels/slide2.xml.rels'].decode('utf-8')
    s2_rels = s2_rels.replace(
        'Target="../media/image5.emf"',
        'Target="../media/image5.png"'
    )
    files['ppt/slides/_rels/slide2.xml.rels'] = s2_rels.encode('utf-8')

    # Resize image transform in slide2.xml to fill available slide area
    try:
        from PIL import Image as _PIL
        import io as _io2, re as _re2
        _img = _PIL.open(_io2.BytesIO(s2_img_bytes))
        _iw, _ih = _img.size
        _asp = _iw / _ih   # image aspect ratio

        # Slide content area (EMU): below header, above red line
        _EMU = 914400
        _SLW = 12801600; _SLH = 9601200
        _LM = int(0.30 * _EMU); _RM = int(0.30 * _EMU)
        _TM = int(1.25 * _EMU); _BM = int(0.25 * _EMU)
        _aw = _SLW - _LM - _RM
        _ah = _SLH - _TM - _BM

        # Scale: fill width, maintain aspect ratio
        _cx = _aw
        _cy = int(_cx / _asp)
        if _cy > _ah:
            _cy = _ah
            _cx = int(_cy * _asp)

        # Center in available area
        _x = _LM + (_aw - _cx) // 2
        _y = _TM + (_ah - _cy) // 2

        s2_xml = files['ppt/slides/slide2.xml'].decode('utf-8')
        # FIX: target the pic element's spPr xfrm specifically
        # The pic's transform lives inside <p:spPr> after <p:blipFill>
        s2_xml = _re2.sub(
            r'(<p:blipFill>.*?</p:blipFill>\s*<p:spPr[^>]*>)\s*'
            r'<a:xfrm><a:off x="\d+" y="\d+"/><a:ext cx="\d+" cy="\d+"/></a:xfrm>',
            lambda m: m.group(1) +
                f'<a:xfrm><a:off x="{_x}" y="{_y}"/>'
                f'<a:ext cx="{_cx}" cy="{_cy}"/></a:xfrm>',
            s2_xml, count=1, flags=_re2.DOTALL
        )
        files['ppt/slides/slide2.xml'] = s2_xml.encode('utf-8')
    except Exception:
        pass   # keep original position if anything fails

    # Write output
    out = io.BytesIO()
    with zipfile.ZipFile(out,'w',compression=zipfile.ZIP_DEFLATED) as oz:
        for fn, fb in files.items(): oz.writestr(fn, fb)
    out.seek(0)
    return out.read()
