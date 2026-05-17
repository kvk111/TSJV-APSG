"""
wb_engine.py — WB Server data processing engine.

Purpose  : Compare WB server data against Online server data for consistency.
           Generate a view-only pivot table after validation.

Core rules:
  - Accepted = 1  → Accepted;  Accepted = 0 → Rejected (direct, no extra logic)
  - Blank Out Weight / Out Time in WB → match by E-Token to Online → use Online values
  - Online corrections (Out Weight entered by user) → synced to WB via E-Token
  - Net Weight pre-validation: In Weight − Out Weight must match file value
  - Pivot: Material as group-header row, Source Sites as data rows, NO subtotal rows
  - Zero Streamlit imports
"""

import io as _io_mod
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def safe_float(v, default=0.0):
    """Return float(v), substituting default for NaN/Inf/None."""
    try:
        result = float(v)
        return default if (math.isnan(result) or math.isinf(result)) else result
    except (TypeError, ValueError):
        return default

# ─────────────────────────────────────────────────────────────
# Robust blank-check for Reject Reason (handles pandas StringDtype NA)
# Pandas StringDtype NA converts to float nan via astype(str), which does NOT
# match the string "nan" in .isin(). Must use .isna() check first.
# ─────────────────────────────────────────────────────────────
def _rr_is_blank(series):
    """True where Reject Reason is effectively blank/empty."""
    return series.isna() | series.astype(str).str.strip().str.lower().isin(
        ["", "nan", "none", "<na>", "pd.na"]
    )


def _normalize_accepted(val):
    """
    Normalize any accepted/rejected value to canonical "Yes" or "No".
    Handles:  1 / "1" / "yes" / "true" / "accepted"  → "Yes"
              0 / "0" / "no"  / "false"/ "rejected"   → "No"
    Returns the canonical string, or the original stripped string for unknown values.
    """
    s = str(val).strip().lower()
    if s in ("1", "yes", "true", "accepted"):
        return "Yes"
    if s in ("0", "no", "false", "rejected", "reject"):
        return "No"
    return str(val).strip()


def _is_rejected_val(val):
    """Return True if val represents a rejected state (No / 0)."""
    return _normalize_accepted(val) == "No"


def _is_accepted_val(val):
    """Return True if val represents an accepted state (Yes / 1)."""
    return _normalize_accepted(val) == "Yes"


def _rr_has_value(series):
    """True where Reject Reason has a real non-blank value."""
    return ~_rr_is_blank(series)


# ─────────────────────────────────────────────────────────────
# Column name map (raw header → canonical)
# ─────────────────────────────────────────────────────────────
WB_COL_MAP = {
    "ticket no.":"Ticket No","ticket no":"Ticket No","ticket":"Ticket No",
    "veh no.":"Vehicle No","veh no":"Vehicle No","vehicle no":"Vehicle No","vehicle number":"Vehicle No",
    "material":"Material","source site":"Source Site",
    "date in":"Date In","time in":"Time In",
    "weight in(t)":"In Weight","weight in (t)":"In Weight","in weight":"In Weight","inweight":"In Weight",
    "date out":"Date Out","time out":"Time Out",
    "weight out(t)":"Out Weight","weight out (t)":"Out Weight","out weight":"Out Weight","outweight":"Out Weight",
    "net weight":"Net Weight","netweight":"Net Weight",
    "e-token":"E-Token","etoken":"E-Token","e token":"E-Token",
    "source site entry":"Source Site Entry","source site exit":"Source Site Exit",
    "accepted":"Accepted","reject reason":"Reject Reason","rejectreason":"Reject Reason",
    "in lane no.":"In Lane","in lane no":"In Lane","in lane":"In Lane",
    "out lane no.":"Out Lane","out lane no":"Out Lane","out lane":"Out Lane",
}

WB_REQUIRED = ["Ticket No","Vehicle No","Material","Source Site",
               "In Weight","Out Weight","Net Weight","E-Token","Accepted"]


# ─────────────────────────────────────────────────────────────
# File loading
# ─────────────────────────────────────────────────────────────
def load_wb_file(uploaded_file) -> dict:
    """Load WB file, normalise columns, parse dates, normalise Accepted 1/0 → Yes/No."""
    try:
        name = getattr(uploaded_file, "name", "file")
        raw  = uploaded_file.read()
        is_xlsx = raw[:4] == b'PK\x03\x04'
        is_xls  = raw[:8] == bytes([0xd0,0xcf,0x11,0xe0,0xa1,0xb1,0x1a,0xe1])
        if is_xlsx or is_xls or name.lower().endswith((".xlsx",".xlsm",".xls")):
            probe = pd.read_excel(_io_mod.BytesIO(raw), engine="openpyxl", header=None, nrows=10)
            header_row = 0
            for i, row in probe.iterrows():
                vals = [str(v).strip().lower() for v in row if pd.notna(v)]
                if any(v in ("ticket no.","ticket no","e-token","material","source site","accepted") for v in vals):
                    header_row = i; break
            df = pd.read_excel(_io_mod.BytesIO(raw), engine="openpyxl", header=header_row)
        else:
            df = None
            for enc in ("utf-8-sig","utf-8","cp1252","latin-1"):
                for sep in (",",";","\t","|"):
                    try:
                        c = pd.read_csv(_io_mod.BytesIO(raw), encoding=enc, sep=sep,
                                        on_bad_lines="skip", engine="python")
                        if c.shape[1] >= 4: df = c; break
                    except Exception: pass
                if df is not None: break
            if df is None: raise ValueError("Could not parse file. Try saving as .xlsx.")
    except Exception as e:
        return {"df": None, "error": f"Cannot read file: {e}"}

    df.columns = [str(c).strip() for c in df.columns]
    rename_map = {col: WB_COL_MAP[col.lower().replace("_"," ").strip()]
                  for col in df.columns
                  if col.lower().replace("_"," ").strip() in WB_COL_MAP
                  and col != WB_COL_MAP[col.lower().replace("_"," ").strip()]}
    if rename_map:
        df = df.rename(columns=rename_map)

    missing = [c for c in WB_REQUIRED if c not in df.columns]
    if missing:
        return {"df": None, "error":
                f"Missing WB columns: {', '.join(missing)}\nFound: {', '.join(df.columns.tolist())}"}

    if "Date In" in df.columns and "Time In" in df.columns:
        df["DateTime In"] = pd.to_datetime(
            df["Date In"].astype(str) + " " + df["Time In"].astype(str), dayfirst=True, errors="coerce")
    elif "DateTime In" in df.columns:
        df["DateTime In"] = pd.to_datetime(df["DateTime In"], dayfirst=True, errors="coerce")

    if "Date Out" in df.columns and "Time Out" in df.columns:
        df["DateTime Out"] = pd.to_datetime(
            df["Date Out"].astype(str) + " " + df["Time Out"].astype(str), dayfirst=True, errors="coerce")
    elif "DateTime Out" in df.columns:
        df["DateTime Out"] = pd.to_datetime(df["DateTime Out"], dayfirst=True, errors="coerce")

    for col in ("In Weight","Out Weight","Net Weight"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["Accepted"] = df["Accepted"].apply(_normalize_accepted)

    return {"df": df, "error": None}


# ─────────────────────────────────────────────────────────────
# Date helpers
# ─────────────────────────────────────────────────────────────
def wb_get_date_range(df):
    col = "DateTime In" if "DateTime In" in df.columns else "Date In"
    valid = df[col].dropna()
    if valid.empty: return None, None
    return valid.min().date(), valid.max().date()


def wb_filter_by_date(df, filter_date):
    if "DateTime In" in df.columns:
        mask = df["DateTime In"].dt.date == filter_date
    elif "Date In" in df.columns:
        mask = pd.to_datetime(df["Date In"], dayfirst=True, errors="coerce").dt.date == filter_date
    else:
        return df.copy()
    return df[mask].copy()


# ─────────────────────────────────────────────────────────────
# Core processing — unified logic + Online sync
# ─────────────────────────────────────────────────────────────
def wb_apply_unified_logic(wb_df, online_df, online_corrections=None):
    """
    Unified processing flow:
      A. Accepted=0 → Rejected, remove from pivot.
      A2. Online-rejected / filtered-out tokens → also reject in WB.
      B. Build Online lookup by Token (incorporates user corrections).
      C. Blank Out Weight / Out Time → resolve via E-Token lookup.
      D. Sync Online user corrections → WB rows (non-blank too).
      E. Net Weight pre-validation: In − Out must equal Net Weight.
    """
    online_corrections = online_corrections or {}

    # Normalize Accepted values (handles "No"/"0"/0, "Yes"/"1"/1 equivalences)
    wb_df = wb_df.copy()
    wb_df["Accepted"] = wb_df["Accepted"].apply(_normalize_accepted)

    # A: Reject Accepted=No (WB-native rejections)
    rej_mask     = wb_df["Accepted"].str.strip().str.lower() == "no"
    rejected_idx = wb_df[rej_mask].index
    working      = wb_df[~rej_mask].copy()

    # A2: Sync Online rejections → WB
    # Any token that is Rejected (NO / has reject reason) in Online Data
    # must also be excluded from WB accepted set.
    online_rejected_tokens: set = set()
    if not online_df.empty and "Token" in online_df.columns:
        for idx, row in online_df.iterrows():
            accepted_val = str(row.get("Accepted", "")).strip().lower()
            reject_reason = str(row.get("Reject Reason", "") or "").strip()
            is_online_rejected = (
                accepted_val in ("no", "rejected")
                or reject_reason not in ("", "nan", "none")
            )
            # Also treat correction-marked NO rows as rejected
            if idx in online_corrections:
                if str(online_corrections[idx].get("Accepted","")).strip().upper() == "NO":
                    is_online_rejected = True
            if is_online_rejected:
                tok = str(row.get("Token", "")).strip()
                if tok:
                    online_rejected_tokens.add(tok)

    if online_rejected_tokens and "E-Token" in working.columns:
        online_rej_mask = working["E-Token"].astype(str).str.strip().isin(online_rejected_tokens)
        online_synced_rej_idx = working[online_rej_mask].index
        rejected_idx = rejected_idx.append(online_synced_rej_idx) if len(online_synced_rej_idx) > 0 else rejected_idx
        working = working[~online_rej_mask].copy()
    else:
        online_synced_rej_idx = pd.Index([])

    # B: Build Online lookup {token_str → {Out Weight, Net Weight, In Weight}}
    token_col     = "Token" if "Token" in online_df.columns else None
    online_lookup = {}
    if token_col is not None and not online_df.empty:
        for idx, row in online_df.iterrows():
            tok  = str(row.get(token_col, "")).strip()
            if not tok: continue
            in_w  = row.get("In Weight")
            out_w = row.get("Out Weight")
            net_w = row.get("Net Weight")
            # Apply user correction from Online step 3
            if idx in online_corrections:
                corr = online_corrections[idx]
                if "Out Weight" in corr:
                    out_w = float(corr["Out Weight"])
                    if pd.notna(in_w):
                        net_w = round(float(in_w) - float(out_w), 3)
                if "Net Weight" in corr:
                    net_w = float(corr["Net Weight"])
            online_lookup[tok] = {
                "Out Weight": float(out_w) if pd.notna(out_w) else None,
                "Net Weight": float(net_w) if pd.notna(net_w) else None,
                "In Weight":  float(in_w)  if pd.notna(in_w)  else None,
            }

    # C: Blank Out Weight / Out Time
    out_w_blank = working["Out Weight"].isna()
    out_t_blank = (working["DateTime Out"].isna()
                   if "DateTime Out" in working.columns
                   else pd.Series(False, index=working.index))
    blank_idx      = working[out_w_blank | out_t_blank].index
    blank_resolved = []
    blank_excluded = []

    for idx in blank_idx:
        etoken = str(working.at[idx, "E-Token"]).strip()
        if etoken in online_lookup:
            entry = online_lookup[etoken]
            if entry["Out Weight"] is not None:
                working.at[idx, "Out Weight"] = entry["Out Weight"]
            if entry["Net Weight"] is not None:
                working.at[idx, "Net Weight"] = entry["Net Weight"]
            blank_resolved.append(idx)
        else:
            blank_excluded.append(idx)

    working = working.drop(index=blank_excluded)

    # D: Sync Online corrections → non-blank WB rows via E-Token
    sync_count = 0
    corrected_rows_info = []   # Fix 6: per-record correction details for UI indicator
    if online_corrections and online_lookup:
        # Build set of tokens that actually had a user correction
        corrected_tokens = set()
        if token_col is not None and not online_df.empty:
            for idx, row in online_df.iterrows():
                if idx in online_corrections:
                    tok = str(row.get(token_col,"")).strip()
                    if tok:
                        corrected_tokens.add(tok)

        for idx in working.index:
            etoken = str(working.at[idx, "E-Token"]).strip()
            if etoken in corrected_tokens and etoken in online_lookup:
                entry = online_lookup[etoken]
                if entry["Out Weight"] is not None:
                    old_out_w = working.at[idx, "Out Weight"]
                    in_w      = working.at[idx, "In Weight"]
                    new_out_w = entry["Out Weight"]
                    working.at[idx, "Out Weight"] = new_out_w
                    new_net_w = None
                    if pd.notna(in_w):
                        new_net_w = round(float(in_w) - new_out_w, 3)
                        working.at[idx, "Net Weight"] = new_net_w
                    sync_count += 1
                    corrected_rows_info.append({
                        "etoken":    etoken,
                        "vehicle":   str(working.at[idx, "Vehicle No"])
                                     if "Vehicle No" in working.columns else "—",
                        "old_out_w": float(old_out_w) if pd.notna(old_out_w) else 0.0,
                        "new_out_w": float(new_out_w),
                        "new_net_w": float(new_net_w) if new_net_w is not None else 0.0,
                    })

    # E: Net Weight pre-validation
    nw_errors = []
    for _, row in working.iterrows():
        in_w  = row.get("In Weight")
        out_w = row.get("Out Weight")
        net_w = row.get("Net Weight")
        if any(v is None or (isinstance(v, float) and pd.isna(v)) for v in (in_w, out_w, net_w)):
            continue
        try:
            calc = round(float(in_w) - float(out_w), 3)
            if abs(float(net_w) - calc) > 0.001:
                nw_errors.append({
                    "etoken":          str(row.get("E-Token","—")),
                    "in_weight":       float(in_w),
                    "out_weight":      float(out_w),
                    "net_weight_file": float(net_w),
                    "net_weight_calc": calc,
                })
        except (TypeError, ValueError):
            continue

    return {
        "wb_accepted_df":          working,
        "rejected_idx":            rejected_idx,
        "blank_idx":               blank_idx,
        "blank_resolved":          pd.Index(blank_resolved),
        "blank_excluded":          pd.Index(blank_excluded),
        "nw_errors":               nw_errors,
        "rejected_count":          len(rejected_idx),
        "accepted_count":          len(working),
        "sync_count":              sync_count,
        "online_synced_rej_count": len(online_synced_rej_idx),
        "synced_tokens":           online_rejected_tokens,
        "corrected_rows_info":     corrected_rows_info,
    }


# ─────────────────────────────────────────────────────────────
# Comparison — EXACT match only (no tolerance)
# ─────────────────────────────────────────────────────────────
def wb_compare_with_online(wb_result, online_df, online_corrections=None):
    """
    Compare WB accepted totals vs Online accepted totals.
    Count match  : exact integer equality.
    NW match     : exact equality to 3 decimal places (< 0.001 tolerance).
    online_df is already date-filtered.

    FIX: Records marked NO via corrections OR with a Reject Reason are treated
    as Rejected on the Online side — they are excluded from both counts.
    """
    online_corrections = online_corrections or {}

    online_accepted = online_df[
        online_df["Accepted"].astype(str).str.strip().str.lower() == "yes"
    ].copy()

    # Exclude rows that have a non-empty Reject Reason (filter-based rejections)
    if "Reject Reason" in online_accepted.columns:
        has_reject_reason = _rr_has_value(online_accepted["Reject Reason"])
        online_accepted = online_accepted[~has_reject_reason]

    # Exclude rows where user correction sets Accepted=NO
    no_correction_idx = [
        idx for idx, corr in online_corrections.items()
        if str(corr.get("Accepted","")).strip().upper() == "NO"
    ]
    online_accepted = online_accepted.drop(
        index=[i for i in no_correction_idx if i in online_accepted.index]
    )

    # Apply user corrections to Online net weight before summing
    for idx, corr in online_corrections.items():
        if idx in online_accepted.index:
            in_w = online_accepted.at[idx, "In Weight"]
            if "Out Weight" in corr and pd.notna(in_w):
                corrected_nw = round(float(in_w) - float(corr["Out Weight"]), 3)
                online_accepted.at[idx, "Net Weight"] = corrected_nw
            elif "Net Weight" in corr:
                online_accepted.at[idx, "Net Weight"] = float(corr["Net Weight"])

    wb_acc   = wb_result["wb_accepted_df"]
    wb_count = len(wb_acc)
    on_count = len(online_accepted)

    # Sum only numeric, non-NaN Net Weight values
    wb_nw = round(float(wb_acc["Net Weight"].dropna().sum()), 3) if not wb_acc.empty else 0.0
    on_nw = round(float(online_accepted["Net Weight"].dropna().sum()), 3) if not online_accepted.empty else 0.0

    return {
        "wb_count":     wb_count,
        "wb_net_wt":    wb_nw,
        "online_count": on_count,
        "online_net_wt": on_nw,
        "count_match":  wb_count == on_count,
        "nw_match":     abs(wb_nw - on_nw) < 0.001,
    }


# ─────────────────────────────────────────────────────────────
# E-Token validation — compare WB tokens against Online tokens
# ─────────────────────────────────────────────────────────────
def wb_validate_etoken_match(wb_accepted_df, online_df, online_corrections=None,
                             online_rejected_tokens=None):
    """
    Compare E-Token values in WB accepted data against Online data (Token column).
    Excludes Online records that are rejected (NO / Reject Reason / correction-marked NO).

    online_rejected_tokens: set of tokens already excluded from WB via Online-rejection sync.
      These are NOT flagged as mismatches — they were intentionally removed.

    Returns:
        {
            "wb_only":          list[str]  — E-Tokens in WB but NOT in Online (genuine mismatches)
            "online_only":      list[str]  — Tokens in Online but NOT in WB
            "wb_total":         int        — total unique WB E-Tokens checked
            "online_total":     int        — total unique Online Tokens checked
            "all_match":        bool       — True when both wb_only and online_only are empty
        }
    """
    online_corrections      = online_corrections      or {}
    online_rejected_tokens  = online_rejected_tokens  or set()

    # WB E-Token set (accepted rows only — already cleaned by wb_apply_unified_logic)
    wb_tokens = set(
        wb_accepted_df["E-Token"].dropna().astype(str).str.strip().tolist()
    ) if not wb_accepted_df.empty else set()

    # Online Token set — accepted rows only, excluding rejected/filtered records
    online_tokens = set()
    if not online_df.empty and "Token" in online_df.columns:
        online_accepted = online_df[
            online_df["Accepted"].astype(str).str.strip().str.lower() == "yes"
        ].copy()
        # Exclude rows with a Reject Reason (filter-based rejections)
        if "Reject Reason" in online_accepted.columns:
            has_reject = _rr_has_value(online_accepted["Reject Reason"])
            online_accepted = online_accepted[~has_reject]
        # Exclude rows where user correction sets Accepted=NO
        no_idx = [
            i for i, c in online_corrections.items()
            if str(c.get("Accepted","")).strip().upper() == "NO"
        ]
        online_accepted = online_accepted.drop(
            index=[i for i in no_idx if i in online_accepted.index]
        )
        online_tokens = set(
            online_accepted["Token"].dropna().astype(str).str.strip().tolist()
        )

    # Tokens in WB not found in Online accepted set
    wb_only_raw = wb_tokens - online_tokens

    # Remove tokens that were already excluded from WB because they matched
    # Online-rejected records — these are handled, not genuine mismatches.
    wb_only = sorted(wb_only_raw - online_rejected_tokens)

    online_only = sorted(online_tokens - wb_tokens)

    return {
        "wb_only":      wb_only,
        "online_only":  online_only,
        "wb_total":     len(wb_tokens),
        "online_total": len(online_tokens),
        "all_match":    len(wb_only) == 0 and len(online_only) == 0,
    }


# ─────────────────────────────────────────────────────────────
# Pivot Table — Material header rows + data rows + Grand Total
# NO subtotal rows (matches reference Image 2 exactly)
# ─────────────────────────────────────────────────────────────
def wb_build_pivot(df):
    """
    Layout (per reference Excel sample):
        MATERIAL_NAME    ← bold group header, no numeric values
          source_site_1  |  loads  |  weight_in  |  weight_out  |  net_weight
          source_site_2  |  ...
        (next material)
        ...
        Grand Total      |  total loads  |  total wi  |  total wo  |  total nw

    No subtotal rows.
    Columns (exact names): Row Labels | Sum of Loads |
        Sum of Weight In (T) | Sum of Weight Out (T) | Sum of Net Weight (T)
    """
    COLS = ["Row Labels","Sum of Loads","Sum of Weight In (T)",
            "Sum of Weight Out (T)","Sum of Net Weight (T)","_row_type","_material"]
    if df.empty:
        return pd.DataFrame(columns=COLS)

    grp = (
        df.groupby(["Material","Source Site"], dropna=False)
        .agg(Loads=("E-Token","count"), WI=("In Weight","sum"),
             WO=("Out Weight","sum"),   NW=("Net Weight","sum"))
        .reset_index()
    )
    # Replace any NaN sums (all-null groups) with 0.0 so JSON stays valid
    grp[["WI","WO","NW"]] = grp[["WI","WO","NW"]].fillna(0.0)

    rows = []
    for material, gm in grp.groupby("Material", sort=True):
        # Material group-header row — no numeric values
        rows.append({
            "Row Labels":            material,
            "_material":             material,
            "Sum of Loads":          None,
            "Sum of Weight In (T)":  None,
            "Sum of Weight Out (T)": None,
            "Sum of Net Weight (T)": None,
            "_row_type":             "mat_header",
        })
        # Source site data rows — sorted alphabetically
        for _, r in gm.sort_values("Source Site").iterrows():
            rows.append({
                "Row Labels":            r["Source Site"],
                "_material":             material,
                "Sum of Loads":          int(r["Loads"]) if pd.notna(r["Loads"]) else 0,
                "Sum of Weight In (T)":  round(safe_float(r["WI"]), 3),
                "Sum of Weight Out (T)": round(safe_float(r["WO"]), 3),
                "Sum of Net Weight (T)": round(safe_float(r["NW"]), 3),
                "_row_type":             "data",
            })
        # NO subtotal row

    # Grand Total only
    rows.append({
        "Row Labels":            "Grand Total",
        "_material":             "",
        "Sum of Loads":          int(grp["Loads"].sum()),
        "Sum of Weight In (T)":  round(safe_float(grp["WI"].sum()),  3),
        "Sum of Weight Out (T)": round(safe_float(grp["WO"].sum()),  3),
        "Sum of Net Weight (T)": round(safe_float(grp["NW"].sum()),  3),
        "_row_type":             "grand_total",
    })
    return pd.DataFrame(rows)


def wb_pivot_summary(pivot_df):
    gt = pivot_df[pivot_df["_row_type"] == "grand_total"]
    if gt.empty:
        return {"total_loads":0,"total_wi":0.0,"total_wo":0.0,"total_nw":0.0}
    r = gt.iloc[0]
    return {"total_loads": int(r["Sum of Loads"]),
            "total_wi":    safe_float(r["Sum of Weight In (T)"]),
            "total_wo":    safe_float(r["Sum of Weight Out (T)"]),
            "total_nw":    safe_float(r["Sum of Net Weight (T)"])}


# ─────────────────────────────────────────────────────────────
# Excel download  (Pivot_Table_<Date>.xlsx)
# ─────────────────────────────────────────────────────────────
def wb_pivot_to_excel(pivot_df, filter_date=None) -> tuple:
    """
    Generate formatted Excel for WB pivot table.

    Layout (per latest spec):
      Row 1  — headers row: col A blank, col B blank, col C = "WB DATA" (bold 13pt centred),
               remaining cols blank.  ALL cells have thin borders.
      Row 2  — column headers: light blue fill, bold 11pt black, centred, thin borders.
      mat_header rows: white bg, bold 11pt, left label + material totals.
      data rows:       white bg, regular 11pt, left label + values.
      grand total row: light blue fill, bold 11pt, all cols.

    Column widths (exact, per spec):
      A (Row Labels)           → 13.22
      B (Sum of Loads)         → 11.56
      C (Sum of Weight In (T)) → 17.87
      D (next)                 → 17
      E (next)                 → 17
      F (Sum of Feed Out (T))  → 19.22   [if present]
      G (final col)            → 19.13   [if present]

    Since we have exactly 5 data columns, mapping is:
      A=13.22, B=11.56, C=17.87, D=17, E=17
    All cells in every row have borders.
    """
    _B         = Side(style="thin", color="FF000000")
    ALL_BORDER = Border(left=_B, right=_B, top=_B, bottom=_B)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT       = Alignment(horizontal="left",   vertical="center")

    BLUE_FILL  = PatternFill("solid", fgColor="FFBDD7EE")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")

    TITLE_FONT = Font(name="Calibri", bold=True,  size=13, color="FF000000")
    HDR_FONT   = Font(name="Calibri", bold=True,  size=11, color="FF000000")
    MAT_FONT   = Font(name="Calibri", bold=True,  size=11, color="FF000000")
    DATA_FONT  = Font(name="Calibri", bold=False, size=11, color="FF000000")
    GT_FONT    = Font(name="Calibri", bold=True,  size=11, color="FF000000")

    COL_HEADERS = ["Row Labels", "Sum of Loads",
                   "Sum of Weight In (T)", "Sum of Weight Out (T)", "Sum of Net Weight (T)"]
    N_COLS = len(COL_HEADERS)

    # ── Exact column widths per spec ──
    # A=Row Labels, B=Sum of Loads, C=Sum of Weight In, D=Sum of Weight Out, E=Sum of Net Weight
    COL_WIDTHS = [13.22, 11.56, 17.87, 17, 17]

    # ── Pre-compute material totals ──
    mat_totals = {}
    for _, r in pivot_df.iterrows():
        if r["_row_type"] == "data":
            mat = r["_material"]
            if mat not in mat_totals:
                mat_totals[mat] = {"Loads": 0, "WI": 0.0, "WO": 0.0, "NW": 0.0}
            mat_totals[mat]["Loads"] += int(r["Sum of Loads"])
            mat_totals[mat]["WI"]    += safe_float(r["Sum of Weight In (T)"])
            mat_totals[mat]["WO"]    += safe_float(r["Sum of Weight Out (T)"])
            mat_totals[mat]["NW"]    += safe_float(r["Sum of Net Weight (T)"])

    wb_out = Workbook()
    ws     = wb_out.active
    ws.title = "WB Data"

    # ── Row 1: Header row — "WB DATA" in cell C1 only, all cells bordered ──
    # Per spec: remove merged row, put "WB DATA" in C1 (col 3), bold 13pt centred.
    # All other cells in row 1 are blank but still have borders.
    for ci in range(1, N_COLS + 1):
        c = ws.cell(1, ci, value="WB DATA" if ci == 3 else None)
        c.font      = TITLE_FONT
        c.fill      = WHITE_FILL
        c.alignment = CENTER
        c.border    = ALL_BORDER
    ws.row_dimensions[1].height = 22

    # ── Row 2: Column headers — light blue, bold black, all bordered ──
    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(2, ci, value=h)
        c.font      = HDR_FONT
        c.fill      = BLUE_FILL
        c.alignment = CENTER
        c.border    = ALL_BORDER
    ws.row_dimensions[2].height = 18

    # ── Data rows from row 3 ──
    current_row = 3

    for _, r in pivot_df.iterrows():
        rt  = r["_row_type"]
        lbl = r["Row Labels"]

        if rt == "mat_header":
            mt   = mat_totals.get(lbl, {"Loads": 0, "WI": 0.0, "WO": 0.0, "NW": 0.0})
            vals = [lbl, mt["Loads"], round(mt["WI"], 2), round(mt["WO"], 2), round(mt["NW"], 2)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(current_row, ci, value=v)
                c.font      = MAT_FONT
                c.fill      = WHITE_FILL
                c.border    = ALL_BORDER
                c.alignment = LEFT if ci == 1 else CENTER
            ws.row_dimensions[current_row].height = 17
            current_row += 1
            continue

        if rt == "grand_total":
            vals = ["Grand Total",
                    int(r["Sum of Loads"]),
                    round(safe_float(r["Sum of Weight In (T)"]),  2),
                    round(safe_float(r["Sum of Weight Out (T)"]), 2),
                    round(safe_float(r["Sum of Net Weight (T)"]), 2)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(current_row, ci, value=v)
                c.font      = GT_FONT
                c.fill      = BLUE_FILL
                c.border    = ALL_BORDER
                c.alignment = LEFT if ci == 1 else CENTER
            ws.row_dimensions[current_row].height = 17
            current_row += 1
            continue

        # Data row — white, regular black
        vals = [lbl,
                int(r["Sum of Loads"]),
                round(safe_float(r["Sum of Weight In (T)"]),  2),
                round(safe_float(r["Sum of Weight Out (T)"]), 2),
                round(safe_float(r["Sum of Net Weight (T)"]), 2)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(current_row, ci, value=v)
            c.font      = DATA_FONT
            c.fill      = WHITE_FILL
            c.border    = ALL_BORDER
            c.alignment = LEFT if ci == 1 else CENTER
        ws.row_dimensions[current_row].height = 16
        current_row += 1

    # ── Auto-fit column widths based on actual content ──
    # Measure max content length per column across all rows
    col_content_maxes = [len(h) for h in COL_HEADERS]  # start with header lengths
    for _, r in pivot_df.iterrows():
        rt  = r["_row_type"]
        lbl = r["Row Labels"]
        vals = [str(lbl)]
        if rt == "mat_header":
            mt = mat_totals.get(lbl, {"Loads": 0, "WI": 0.0, "WO": 0.0, "NW": 0.0})
            vals += [str(mt["Loads"]), f"{mt['WI']:.2f}", f"{mt['WO']:.2f}", f"{mt['NW']:.2f}"]
        elif rt == "grand_total":
            vals += [str(int(r["Sum of Loads"])),
                     f"{safe_float(r['Sum of Weight In (T)']):.2f}",
                     f"{safe_float(r['Sum of Weight Out (T)']):.2f}",
                     f"{safe_float(r['Sum of Net Weight (T)']):.2f}"]
        else:
            vals += [str(int(r["Sum of Loads"])),
                     f"{safe_float(r['Sum of Weight In (T)']):.2f}",
                     f"{safe_float(r['Sum of Weight Out (T)']):.2f}",
                     f"{safe_float(r['Sum of Net Weight (T)']):.2f}"]
        for i, v in enumerate(vals):
            if i < len(col_content_maxes) and len(v) > col_content_maxes[i]:
                col_content_maxes[i] = len(v)

    for ci, max_len in enumerate(col_content_maxes, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 2  # +2 padding

    buf = _io_mod.BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    date_str = filter_date.strftime("%d%m%Y") if filter_date else "export"
    fname    = f"Pivot_Table_{date_str}.xlsx"
    return buf.getvalue(), fname



# ─────────────────────────────────────────────────────────────
# Copy-paste text builder (center-aligned, tab-separated)
# ─────────────────────────────────────────────────────────────
def wb_pivot_copy_text(pivot_df):
    """Full tab-separated copy of pivot, center-padded, paste-ready."""
    COL_W   = [30, 14, 22, 22, 22]
    HEADERS = ["Row Labels","Sum of Loads","Sum of Weight In (T)",
               "Sum of Weight Out (T)","Sum of Net Weight (T)"]

    def _pad(val, width): return str(val).center(width)

    lines = []
    lines.append("WB DATA".center(sum(COL_W) + len(COL_W) * 3))
    lines.append("")
    lines.append("\t".join(_pad(h, COL_W[i]) for i, h in enumerate(HEADERS)))
    lines.append("\t".join("-" * w for w in COL_W))

    for _, r in pivot_df.iterrows():
        rt  = r["_row_type"]
        lbl = r["Row Labels"]

        if rt == "mat_header":
            lines.append(_pad(lbl, COL_W[0]) + "\t" +
                         "\t".join(_pad("", COL_W[i]) for i in range(1, 5)))
        elif rt == "grand_total":
            lines.append("\t".join([
                _pad("Grand Total",                         COL_W[0]),
                _pad(f"{int(r['Sum of Loads'])}",         COL_W[1]),
                _pad(f"{r['Sum of Weight In (T)']:.3f}",   COL_W[2]),
                _pad(f"{r['Sum of Weight Out (T)']:.3f}",  COL_W[3]),
                _pad(f"{r['Sum of Net Weight (T)']:.3f}",  COL_W[4]),
            ]))
        else:
            def _n(v, dec=3):
                return f"{float(v):.{dec}f}" if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
            lines.append("\t".join([
                _pad(lbl,                                    COL_W[0]),
                _pad(f"{int(r['Sum of Loads'])}" if r["Sum of Loads"] is not None else "", COL_W[1]),
                _pad(_n(r["Sum of Weight In (T)"]),          COL_W[2]),
                _pad(_n(r["Sum of Weight Out (T)"]),         COL_W[3]),
                _pad(_n(r["Sum of Net Weight (T)"]),         COL_W[4]),
            ]))

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────
# NEW: Row-level WB validation helpers (mandatory field check)
# ─────────────────────────────────────────────────────────────

def wb_find_incomplete_rows(df):
    """
    Return a DataFrame of rows where Date Out or Time Out is blank.
    These rows need user decision before processing.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    date_out_blank = df.get("Date Out", pd.Series(dtype=object)).isna()
    time_out_blank = df.get("Time Out", pd.Series(dtype=object)).isna()

    # Also catch string "nan", "none", empty strings
    def _is_blank(series):
        return series.isna() | series.astype(str).str.strip().str.lower().isin(
            ["", "nan", "none", "nat"]
        )

    if "Date Out" in df.columns:
        date_out_blank = _is_blank(df["Date Out"])
    else:
        date_out_blank = pd.Series(True, index=df.index)

    if "Time Out" in df.columns:
        time_out_blank = _is_blank(df["Time Out"])
    else:
        time_out_blank = pd.Series(True, index=df.index)

    incomplete = df[date_out_blank | time_out_blank].copy()
    return incomplete


def wb_apply_row_decisions(df, row_decisions):
    """
    Apply user decisions for each incomplete row.

    row_decisions: dict of {original_index: {"decision": 0|1, "out_weight": float|None}}
      decision=0 → remove row
      decision=1 → keep row, fill Out Weight, compute Net Weight

    Returns:
        processed_df  — df with decisions applied (rows removed or updated)
        applied_info  — list of dicts describing what happened to each row
    """
    processed = df.copy()
    applied_info = []

    rows_to_drop = []

    for idx, dec in row_decisions.items():
        if idx not in processed.index:
            continue

        decision = dec.get("decision")
        etoken   = str(processed.at[idx, "E-Token"]) if "E-Token" in processed.columns else str(idx)

        if decision == 0:
            # Remove the row
            rows_to_drop.append(idx)
            applied_info.append({"etoken": etoken, "action": "removed", "idx": idx})

        elif decision == 1:
            out_w = dec.get("out_weight")
            in_w  = processed.at[idx, "In Weight"] if "In Weight" in processed.columns else None

            if out_w is not None:
                processed.at[idx, "Out Weight"] = float(out_w)
                if in_w is not None and pd.notna(in_w):
                    net_w = round(float(in_w) - float(out_w), 3)
                    processed.at[idx, "Net Weight"] = net_w
                else:
                    net_w = None

                # Mark Accepted as Yes
                processed.at[idx, "Accepted"] = "Yes"

                applied_info.append({
                    "etoken":   etoken,
                    "action":   "accepted",
                    "out_w":    float(out_w),
                    "net_w":    net_w,
                    "in_w":     float(in_w) if in_w is not None and pd.notna(in_w) else None,
                    "idx":      idx,
                })

    processed = processed.drop(index=rows_to_drop)
    return processed, applied_info


def wb_validate_counts(wb_df_filtered, online_df, wb_row_decisions=None,
                       online_corrections=None, online_accepted_total=None,
                       online_rejected_total=None):
    """
    Validation 2 & 3:
      - Compare total row count: WB (after decisions) vs Online
      - Compare rejected count: WB vs Online

    Returns dict with validation results.
    """
    wb_row_decisions   = wb_row_decisions   or {}
    online_corrections = online_corrections or {}

    # WB: rows removed by decision=0 are excluded
    removed_indices = {idx for idx, d in wb_row_decisions.items() if d.get("decision") == 0}
    wb_working = wb_df_filtered.drop(
        index=[i for i in removed_indices if i in wb_df_filtered.index]
    )

    wb_total    = len(wb_working)
    wb_rejected = int(
        (wb_working["Accepted"].astype(str).str.strip().str.lower() == "no").sum()
    ) if not wb_working.empty else 0
    wb_accepted = wb_total - wb_rejected

    # Online totals (if provided directly, use them; else compute)
    if online_accepted_total is not None:
        on_total    = (online_accepted_total or 0) + (online_rejected_total or 0)
        on_rejected = online_rejected_total or 0
        on_accepted = online_accepted_total or 0
    elif online_df is not None and not online_df.empty:
        on_total    = len(online_df)
        on_rejected = int(
            (online_df["Accepted"].astype(str).str.strip().str.lower() == "no").sum()
        )
        on_accepted = on_total - on_rejected
    else:
        on_total = on_rejected = on_accepted = None

    count_match    = (wb_total    == on_total)    if on_total    is not None else None
    rejected_match = (wb_rejected == on_rejected) if on_rejected is not None else None

    return {
        "wb_total":      wb_total,
        "wb_accepted":   wb_accepted,
        "wb_rejected":   wb_rejected,
        "on_total":      on_total,
        "on_accepted":   on_accepted,
        "on_rejected":   on_rejected,
        "count_match":   count_match,
        "rejected_match": rejected_match,
    }


def wb_net_weight_validation(df):
    """
    Validation 1: For each row, verify Net Weight == In Weight - Out Weight.
    Returns list of dicts for mismatched rows.
    """
    errors = []
    if df is None or df.empty:
        return errors

    for idx, row in df.iterrows():
        in_w  = row.get("In Weight")
        out_w = row.get("Out Weight")
        net_w = row.get("Net Weight")

        if any(v is None or (isinstance(v, float) and pd.isna(v))
               for v in (in_w, out_w, net_w)):
            continue
        try:
            calc = round(float(in_w) - float(out_w), 3)
            if abs(float(net_w) - calc) > 0.001:
                errors.append({
                    "idx":             idx,
                    "etoken":          str(row.get("E-Token", idx)),
                    "in_weight":       float(in_w),
                    "out_weight":      float(out_w),
                    "net_weight_file": float(net_w),
                    "net_weight_calc": calc,
                    "diff":            round(float(net_w) - calc, 3),
                })
        except (TypeError, ValueError):
            continue
    return errors
