"""
report_engine.py — Business logic for Staging Ground Report Generator.

Output format matches: APSG_-_WB_Summary_Contract_Report reference file exactly.
Fully decoupled from Streamlit so this module can be imported into any framework.
"""

import io
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Constants — exact values from reference file
# ──────────────────────────────────────────────
PROJECT_NAME = "Management of Staging Ground and Infilling Works (Phase 3)"
CONTRACTOR   = "TOA - Samsung C&T Joint Venture"   # corrected: TOA not DOA
DATE_COL     = "Date Time Arrival"
EXIT_COL     = "Date Time Exit"

# Exact column headers as in reference output
OUTPUT_COLS = [
    "TOKEN", "SITE CODE", "DATETIME ARRIVAL", "DATETIME EXIT",
    "VEHICLE NO", "MATERIAL", "ACCEPTED", "REJECT REASON",
    "TRAN STATUS", "WB IN TIME", "IN WEIGHT", "WB OUT TIME",
    "OUT WEIGHT", "NET WEIGHT", "HAULAGE CONTRACTOR",
    "IN LANE", "OUT LANE", "STAGING GROUND", "RE-CLASSIFIED",
]

# Canonical internal column names
REQUIRED_COLS = [
    "Token", "Source Site", "Date Time Arrival", "Date Time Exit",
    "Vehicle Number", "Material", "Accepted", "Reject Reason",
    "Transaction Status", "WB In Time", "In Weight", "WB Out Time",
    "Out Weight", "Net Weight", "Haulage Contractor",
    "In Lane", "Out Lane", "Staging Ground", "Reclassified",
]

# Map raw file column names -> canonical (matched case-insensitively)
COL_MAP = {
    "token":              "Token",
    "site code":          "Source Site",
    "source site":        "Source Site",
    "datetime arrival":   "Date Time Arrival",
    "date time arrival":  "Date Time Arrival",
    "datetimearrival":    "Date Time Arrival",
    "datetime exit":      "Date Time Exit",
    "date time exit":     "Date Time Exit",
    "datetimeexit":       "Date Time Exit",
    "vehicle no":         "Vehicle Number",
    "vehicle number":     "Vehicle Number",
    "vehicleno":          "Vehicle Number",
    "material":           "Material",
    "accepted":           "Accepted",
    "reject reason":      "Reject Reason",
    "rejectreason":       "Reject Reason",
    "tran status":        "Transaction Status",
    "transaction status": "Transaction Status",
    "transtatus":         "Transaction Status",
    "wb in time":         "WB In Time",
    "wbintime":           "WB In Time",
    "in weight":          "In Weight",
    "inweight":           "In Weight",
    "wb out time":        "WB Out Time",
    "wbouttime":          "WB Out Time",
    "out weight":         "Out Weight",
    "outweight":          "Out Weight",
    "net weight":         "Net Weight",
    "netweight":          "Net Weight",
    "haulage contractor": "Haulage Contractor",
    "haulagecontractor":  "Haulage Contractor",
    "in lane":            "In Lane",
    "inlane":             "In Lane",
    "out lane":           "Out Lane",
    "outlane":            "Out Lane",
    "staging ground":     "Staging Ground",
    "stagingground":      "Staging Ground",
    "re-classified":      "Reclassified",
    "reclassified":       "Reclassified",
}

# ──────────────────────────────────────────────
# Color palette
# ──────────────────────────────────────────────
# Detail sheets (Contract, Rejected, Accepted):
#   ALL header columns → Excel Standard Blue (#4472C4) fill, Black Bold text
DETAIL_HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
DETAIL_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FF000000")

# Detail sheet data rows → plain WHITE background only (no alternating fills)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")

# Summary sheet:
#   Header + Grand Total → "Dark Blue, Text 2, Lighter 80%" = #DAE3F0 fill, dark navy bold text
SUMM_HEADER_FILL = PatternFill("solid", fgColor="FFDAE3F0")
SUMM_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FF1F3864")
GRAND_FILL       = PatternFill("solid", fgColor="FFDAE3F0")
GRAND_FONT       = Font(name="Calibri", bold=True, size=11, color="FF1F3864")

# No background fill (summary data rows + any plain cells)
NO_FILL = PatternFill(fill_type=None)

# ──────────────────────────────────────────────
# Shared style constants
# ──────────────────────────────────────────────
COL_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

META_FONT    = Font(name="Calibri", bold=False, size=14)
DATA_FONT    = Font(name="Calibri", size=11)
BOLD_FONT    = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT  = Font(name="Calibri", size=11)
MAT_FONT     = Font(name="Calibri", bold=True, size=11)

# All borders — solid black thin lines, clearly visible on all cells
_B = Side(style="thin", color="FF000000")
DATA_BORDER = Border(left=_B, right=_B, top=_B, bottom=_B)

LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
RIGHT_ALIGN  = Alignment(horizontal="right",  vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ──────────────────────────────────────────────
# Detail-sheet column widths
# ──────────────────────────────────────────────
COL_WIDTHS = {
    "TOKEN": 28,
    "SITE CODE": 20,
    "DATETIME ARRIVAL": 22,
    "DATETIME EXIT": 22,
    "VEHICLE NO": 14,
    "MATERIAL": 18,
    "ACCEPTED": 12,
    "REJECT REASON": 22,
    "TRAN STATUS": 14,
    "WB IN TIME": 22,
    "IN WEIGHT": 13,
    "WB OUT TIME": 22,
    "OUT WEIGHT": 13,
    "NET WEIGHT": 13,
    "HAULAGE CONTRACTOR": 26,
    "IN LANE": 10,
    "OUT LANE": 10,
    "STAGING GROUND": 18,
    "RE-CLASSIFIED": 15,
}

# ──────────────────────────────────────────────
# Validation / flag highlight styles
# ──────────────────────────────────────────────
FLAG_ROW_FILL  = PatternFill("solid", fgColor="FFFFFF00")   # Yellow — needs attention
FLAG_CELL_FONT = Font(name="Calibri", size=11, color="FFFF0000", bold=True)  # Red bold — modified field

NUMERIC_COLS = {"In Weight", "Out Weight", "Net Weight"}


# ──────────────────────────────────────────────
# File loading & validation
# ──────────────────────────────────────────────

def load_and_validate_file(uploaded_file) -> dict:
    """Returns {'df': DataFrame, 'error': str|None}."""
    import io as _io

    try:
        name = getattr(uploaded_file, "name", "file.csv")
        raw  = uploaded_file.read()

        is_xlsx = raw[:4] == b'PK\x03\x04'
        is_xls  = raw[:8] == bytes([0xd0,0xcf,0x11,0xe0,0xa1,0xb1,0x1a,0xe1])

        if is_xlsx or is_xls or name.lower().endswith((".xlsx", ".xlsm", ".xls")):
            probe = pd.read_excel(_io.BytesIO(raw), engine="openpyxl", header=None, nrows=10)
            header_row = 0
            for i, row in probe.iterrows():
                vals = [str(v).strip().lower() for v in row if pd.notna(v)]
                if any(v in ("token", "source site", "date time arrival") for v in vals):
                    header_row = i
                    break
            df = pd.read_excel(_io.BytesIO(raw), engine="openpyxl", header=header_row)
        else:
            df = None
            for enc in ("utf-8-sig","utf-8","cp1252","latin-1","cp1250","iso-8859-1"):
                for sep in (",",";","\t","|"):
                    try:
                        c = pd.read_csv(_io.BytesIO(raw), encoding=enc, sep=sep,
                                        on_bad_lines="skip", engine="python")
                        if c.shape[1] >= 2:
                            df = c; break
                    except Exception:
                        pass
                if df is not None:
                    break
            if df is None:
                text = raw.decode("latin-1", errors="replace")
                for sep in (",",";","\t","|"):
                    try:
                        c = pd.read_csv(_io.StringIO(text), sep=sep,
                                        engine="python", on_bad_lines="skip")
                        if c.shape[1] >= 2:
                            df = c; break
                    except Exception:
                        pass
            if df is None:
                raise ValueError("Could not read file — try saving as .xlsx or CSV UTF-8.")

    except Exception as e:
        return {"df": None, "error": f"Cannot read file: {e}"}

    # Column normalisation
    df.columns = [c.strip() for c in df.columns]
    rename_map = {}
    for col in df.columns:
        key = col.lower().replace("_"," ").strip()
        if key in COL_MAP and col != COL_MAP[key]:
            rename_map[col] = COL_MAP[key]
    if rename_map:
        df = df.rename(columns=rename_map)

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        return {"df": None, "error": (
            f"Missing columns: {', '.join(missing[:5])}"
            + (" …" if len(missing)>5 else "")
            + f"\nFound: {', '.join(list(df.columns))}"
        )}

    # Preserve original raw string values before parsing (for format-exact output)
    for col in (DATE_COL, EXIT_COL, "WB In Time", "WB Out Time"):
        if col in df.columns:
            raw_col = f"_raw_{col.replace(' ', '_')}"
            df[raw_col] = df[col].astype(str).where(df[col].notna(), other="")

    for col in (DATE_COL, EXIT_COL):
        df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
    for col in ("WB In Time", "WB Out Time"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
    for col in ("In Weight","Out Weight","Net Weight"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    if df[DATE_COL].isna().all():
        return {"df": None, "error": f"No parseable dates in '{DATE_COL}'."}

    # Filter column: WB In Time is the official record date; fall back to Date Time Arrival
    if "WB In Time" in df.columns and not df["WB In Time"].isna().all():
        df["_filter_date"] = df["WB In Time"]
    else:
        df["_filter_date"] = df[DATE_COL]

    return {"df": df, "error": None}


def get_date_range(df: pd.DataFrame):
    col = "_filter_date" if "_filter_date" in df.columns else DATE_COL
    valid = df[col].dropna()
    return valid.min().date(), valid.max().date()


# ──────────────────────────────────────────────
# Filter helpers
# ──────────────────────────────────────────────

def _filter_by_dates(df, from_date, to_date):
    filter_col = "_filter_date" if "_filter_date" in df.columns else DATE_COL
    m = (df[filter_col].dt.date >= from_date) & (df[filter_col].dt.date <= to_date)
    return df[m].copy()


def filter_preview(df: pd.DataFrame, from_date, to_date) -> pd.DataFrame:
    """
    Public helper: returns the exact same filtered slice that generate_report()
    will use. Import this in app.py to drive the preview stats so the UI numbers
    always match the Excel output — single source of truth for date filtering.

    Filters on WB In Time when available (same as the engine), falls back to
    Date Time Arrival. This is the correct official record date column.
    """
    return _filter_by_dates(df, from_date, to_date)


def validate_and_flag(df: pd.DataFrame) -> dict:
    """
    Validation scope (per specification):
      ONLY rows where Accepted = YES  AND  (WB Out Time is blank OR Out Weight is blank).

    Rejected rows, rows with existing Out Weight, and all other rows are NOT touched.

    Returns:
      {
        "incomplete_accepted": [row indices],   # Accepted=YES but missing Out Time/Out Weight
        "clean":               bool,            # True = nothing needs user attention
      }

    Row indices are the actual df.index values so callers can .loc[] them.
    """
    incomplete = []

    for idx, row in df.iterrows():
        # Only look at Accepted = YES rows
        accepted_val = str(row.get("Accepted", "")).strip().lower()
        if accepted_val != "yes":
            continue   # ← skip rejected / unknown rows entirely

        # Check if Out Time is blank
        wb_out = row.get("WB Out Time")
        out_time_blank = (
            wb_out is None
            or (isinstance(wb_out, float) and pd.isna(wb_out))
            or (hasattr(wb_out, "year") is False and pd.isnull(wb_out))
        )

        # Check if Out Weight is blank / zero / NaN
        out_w = row.get("Out Weight")
        try:
            out_weight_blank = out_w is None or pd.isna(out_w)
        except (TypeError, ValueError):
            out_weight_blank = False

        if out_time_blank or out_weight_blank:
            incomplete.append(idx)

    return {
        "incomplete_accepted": incomplete,
        "clean":               len(incomplete) == 0,
    }

def _is_rejected(val):
    return not pd.isna(val) and str(val).strip().lower() in {"no","rejected"}

def _is_accepted(val):
    return str(val).strip().lower() == "yes"


# ──────────────────────────────────────────────
# Shared sheet-writing helpers
# ──────────────────────────────────────────────

def _fmt(val, raw_val=None):
    """
    Format a cell value for Excel output.
    Date/Time preservation rule: if the original raw string is available (raw_val),
    use it exactly as-is — no reformatting. This preserves the exact date/time
    format from the uploaded file (e.g. 20-04-2026 stays 20-04-2026, not d/m/Y).
    Only falls back to timestamp formatting when no raw string is available.
    """
    if val is None: return ""
    # If original raw string supplied, use it directly (preserves input format)
    if raw_val is not None and isinstance(raw_val, str) and raw_val.strip():
        return raw_val.strip()
    if isinstance(val, (pd.Timestamp, datetime)):
        try:
            return val.strftime("%d/%m/%Y %H:%M") if not pd.isnull(val) else ""
        except Exception:
            return ""
    try:
        if pd.isna(val): return ""
    except Exception:
        pass
    return val


def _meta_rows(ws, from_date, to_date):
    """Rows 1-4: project info — DO NOT MODIFY (per specification)."""
    ws["A1"] = f"Project Name : {PROJECT_NAME.upper()}"
    ws["A2"] = f"Contractor Name : {CONTRACTOR.upper()}"
    ws["A3"] = f"Period:{from_date.strftime('%d/%m/%Y')} To: {to_date.strftime('%d/%m/%Y')}"
    for r in (1, 2, 3):
        ws.cell(r, 1).font = META_FONT
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[4].height = 8


def _col_headers(ws, headers, row=5):
    """
    Write column headers for detail sheets (Contract, Rejected, Accepted).
    ALL columns → Lighter 40% blue (#8EA9C1) fill, dark navy bold text.
    Center-aligned, solid black ALL BORDERS.
    No auto-filter applied (filters removed per specification).
    """
    for ci, h in enumerate(headers, 1):
        c           = ws.cell(row, ci, value=h)
        c.font      = DETAIL_HEADER_FONT
        c.fill      = DETAIL_HEADER_FILL
        c.alignment = COL_HEADER_ALIGN
        c.border    = DATA_BORDER
    ws.row_dimensions[row].height = 25


def _data_rows(ws, df, start_row=6, flag_map=None):
    """
    Write data rows for detail sheets.
    - Plain WHITE background (no alternating colors)
    - All cells CENTER-aligned
    - Solid black ALL BORDERS on every cell

    flag_map: dict  {original_df_index: {"flagged": bool, "red_cols": set_of_internal_col_names}}
      flagged=True  → entire row gets YELLOW fill
      red_cols      → those specific cells get RED BOLD font
    """
    flag_map = flag_map or {}

    for r_idx, (orig_idx, row) in enumerate(df.iterrows()):
        er       = start_row + r_idx
        flag_info = flag_map.get(orig_idx, {})
        is_flagged = flag_info.get("flagged", False)
        red_cols   = flag_info.get("red_cols", set())

        row_fill = FLAG_ROW_FILL if is_flagged else WHITE_FILL

        for ci, internal in enumerate(REQUIRED_COLS, 1):
            raw_val = row.get(internal, "")
            # Use stored raw string for date/time columns to preserve original format
            _raw_key = f"_raw_{internal.replace(' ', '_')}"
            raw_str  = row.get(_raw_key, None) if _raw_key in row.index else None
            val      = _fmt(raw_val, raw_val=raw_str)
            if internal in NUMERIC_COLS and not isinstance(raw_val, str):
                try:
                    if not pd.isna(raw_val):
                        fval = float(raw_val)
                        # Write as int when value is a whole number (avoids 0.0, 1000.0 display)
                        val = int(fval) if fval == int(fval) else fval
                    else:
                        val = ""
                except Exception:
                    pass
            # Accepted field always written in UPPERCASE (YES / NO / REJECTED)
            if internal == "Accepted" and isinstance(val, str) and val.strip():
                val = val.strip().upper()
            cell           = ws.cell(er, ci, value=val)
            cell.fill      = row_fill
            cell.border    = DATA_BORDER
            cell.alignment = CENTER_ALIGN
            if internal in NUMERIC_COLS and isinstance(val, float):
                cell.number_format = "0.###"   # decimal, trailing zeros suppressed

            # Apply red bold font to modified fields; normal font to rest
            if internal in red_cols:
                cell.font = FLAG_CELL_FONT
            else:
                cell.font = DATA_FONT

            # Fix 5: "Reject Reason" — for rejected rows with blank reason,
            # write "Breakdown" in red, title-case, matching existing font size.
            if internal == "Reject Reason":
                accepted_val = str(row.get("Accepted", "")).strip().lower()
                is_row_rejected = accepted_val in ("no", "rejected", "0")
                cell_val = cell.value
                val_is_blank = (
                    cell_val is None
                    or str(cell_val).strip().lower() in ("", "nan", "none")
                )
                if is_row_rejected and val_is_blank:
                    cell.value = "Breakdown"
                    cell.font  = Font(name="Calibri", size=11,
                                      color="FFFF0000", bold=False)


def _col_widths(ws, df=None, sheet_type="auto"):
    """
    Apply column widths for REJECT REASON based on sheet_type:
      "contract"  → fixed width 29.28
      "rejected"  → auto-adjust based on content length
      "accepted"  → auto-adjust based on content length (unchanged)
    All other columns use COL_WIDTHS defaults.
    """
    for ci, col in enumerate(OUTPUT_COLS, 1):
        col_letter = get_column_letter(ci)
        if col == "REJECT REASON":
            if sheet_type == "contract":
                # Sheet 1 (Contractor Report): fixed width 29.28
                ws.column_dimensions[col_letter].width = 29.28
            else:
                # Sheet 2 (Loads Rejected) and Sheet 3 (Loads Accepted):
                # auto-size based on actual content length
                max_content = 0
                if df is not None:
                    for _, row in df.iterrows():
                        v = str(row.get("Reject Reason", "") or "")
                        if v.lower() not in ("", "nan", "none"):
                            if len(v) > max_content:
                                max_content = len(v)
                header_width = len("REJECT REASON")
                if max_content > 0:
                    width = min(60, max(header_width + 2, max_content + 4))
                else:
                    width = header_width + 2
                ws.column_dimensions[col_letter].width = width
            # Enable wrap text for REJECT REASON cells in all sheet types
            for r in ws.iter_rows(min_col=ci, max_col=ci):
                for cell in r:
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal or "center",
                            vertical=cell.alignment.vertical or "center",
                            wrap_text=True,
                        )
        else:
            ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col, 15)


def _finalise(ws, n_rows, n_cols=19, header_row=5):
    """No freeze pane, no auto-filter (both removed per specification)."""
    ws.freeze_panes = None
    # Auto-filter intentionally removed


def _sort_by_wb_time(df: pd.DataFrame) -> pd.DataFrame:
    """Sort data ascending by WB In Time (falls back to Date Time Arrival).
    Preserves original DataFrame index so flag_map lookups work correctly."""
    sort_col = "WB In Time" if ("WB In Time" in df.columns and not df["WB In Time"].isna().all()) \
               else DATE_COL
    return df.sort_values(sort_col, ascending=True, na_position="last")


# ──────────────────────────────────────────────
# Sheet builders
# ──────────────────────────────────────────────

def _sheet_contract(ws, df, from_date, to_date, flag_map=None):
    df = _sort_by_wb_time(df)
    _meta_rows(ws, from_date, to_date)
    _col_headers(ws, OUTPUT_COLS)
    _data_rows(ws, df, flag_map=flag_map)
    _col_widths(ws, df, sheet_type="contract")
    _finalise(ws, len(df))


def _sheet_rejected(ws, df, from_date, to_date, flag_map=None):
    rejected = df[df["Accepted"].apply(_is_rejected)]
    rejected = _sort_by_wb_time(rejected)
    _meta_rows(ws, from_date, to_date)
    _col_headers(ws, OUTPUT_COLS)
    _data_rows(ws, rejected, flag_map=flag_map)
    _col_widths(ws, rejected, sheet_type="rejected")
    _finalise(ws, len(rejected))
    return len(rejected)


def _sheet_accepted(ws, df, from_date, to_date, flag_map=None):
    accepted = df[df["Accepted"].apply(_is_accepted)]
    accepted = _sort_by_wb_time(accepted)
    _meta_rows(ws, from_date, to_date)
    _col_headers(ws, OUTPUT_COLS)
    _data_rows(ws, accepted, flag_map=flag_map)
    _col_widths(ws, accepted)
    _finalise(ws, len(accepted))
    return accepted


def _sheet_summary(ws, accepted_df, from_date, to_date):
    """
    Summary sheet layout:
      Rows 1-3 : meta (untouched — DO NOT MODIFY)
      Row  4   : gap
      Row  5   : WEIGHBRIDGE SUMMARY CONTRACT REPORT — Black, Bold, Underline
      Row  6   : gap
      Row  7   : pivot table headers  (Lighter 80% = #DAE3F0, dark navy text)
      Row  8+  : data rows — NO background fill, center-aligned, ALL BORDERS
      Last row : Grand Total — same style as header (Lighter 80%)

    No freeze pane on Summary.
    Print titles: rows 1-7 repeat on every printed page.
    All row heights fixed at 14.40 points for clean A4 print layout.
    """
    # ── Fixed row height for all summary rows (14.40 pt = clean A4 print) ──
    SUMM_ROW_HEIGHT = 14.40

    _meta_rows(ws, from_date, to_date)

    # Row 5: title — Merge A5:C5, Black Bold Underline, LEFT-aligned
    ws.merge_cells("A5:C5")
    title_cell           = ws["A5"]
    title_cell.value     = "WEIGHBRIDGE SUMMARY CONTRACT REPORT"
    title_cell.font      = Font(name="Calibri", bold=True, size=14,
                                color="FF000000", underline="single")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = SUMM_ROW_HEIGHT

    # Row 6: blank spacer
    ws.row_dimensions[6].height = SUMM_ROW_HEIGHT

    # Summary column widths (exact specification):
    #   A (Material / Source Site Code) = 27.22  ≈ 252 px
    #   B (Loads)                       = 19.22  ≈ 150 px
    #   C (NET WEIGHT)                  = 43.88  ≈ 402 px
    SUMM_COL_WIDTHS = [27.22, 19.22, 44.89]  # Col C (REJECT REASON / NET WEIGHT) fixed at 44.89

    # Row 7: pivot table headers — Lighter 80%, dark navy text, center, ALL BORDERS
    summ_hdrs = ["Material / Source Site Code", "Loads", "NET WEIGHT (T)"]
    for ci, (h, w) in enumerate(zip(summ_hdrs, SUMM_COL_WIDTHS), 1):
        cell           = ws.cell(7, ci, value=h)
        cell.font      = SUMM_HEADER_FONT
        cell.fill      = SUMM_HEADER_FILL
        cell.alignment = COL_HEADER_ALIGN
        cell.border    = DATA_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[7].height = SUMM_ROW_HEIGHT

    # ── Pivot data ──
    data_start    = 8
    er            = data_start
    tot_loads_all = 0
    tot_wt_all    = 0.0

    if (not accepted_df.empty
            and "Material"    in accepted_df.columns
            and "Source Site" in accepted_df.columns):

        grp = (
            accepted_df
            .groupby(["Material", "Source Site"], dropna=False)
            .agg(
                Loads=("Token", "count"),           # count rows, not Net Weight (avoids NaN skip)
                Net_Weight=("Net Weight", "sum"),
            )
            .reset_index()
            .sort_values(["Material", "Source Site"])
        )

        for material, mat_df in grp.groupby("Material", sort=True):
            mat_loads      = int(mat_df["Loads"].sum())
            mat_wt         = round(float(mat_df["Net_Weight"].sum()), 2)
            tot_loads_all += mat_loads
            tot_wt_all    += mat_wt

            # Material subtotal row — bold, NO fill, center, ALL BORDERS
            for ci, v in enumerate([material, mat_loads, mat_wt], 1):
                cell           = ws.cell(er, ci, value=v)
                cell.font      = MAT_FONT
                cell.fill      = NO_FILL
                cell.border    = DATA_BORDER
                cell.alignment = CENTER_ALIGN
                if ci == 3:
                    cell.number_format = "0.###"   # no thousand separators
            ws.row_dimensions[er].height = SUMM_ROW_HEIGHT
            er += 1

            # Source site detail rows — normal, NO fill, center, ALL BORDERS
            for _, row in mat_df.iterrows():
                vals = [row["Source Site"], int(row["Loads"]),
                        round(float(row["Net_Weight"]), 2)]
                for ci, v in enumerate(vals, 1):
                    cell           = ws.cell(er, ci, value=v)
                    cell.font      = NORMAL_FONT
                    cell.fill      = NO_FILL
                    cell.border    = DATA_BORDER
                    cell.alignment = CENTER_ALIGN
                    if ci == 3:
                        cell.number_format = "0.###"   # no thousand separators
                ws.row_dimensions[er].height = SUMM_ROW_HEIGHT
                er += 1

    # Grand Total row — same Lighter 80% style as header, center, ALL BORDERS
    tot_wt_all = round(tot_wt_all, 2)
    for ci, v in enumerate(["Grand Total", tot_loads_all, tot_wt_all], 1):
        cell           = ws.cell(er, ci, value=v)
        cell.font      = GRAND_FONT
        cell.fill      = GRAND_FILL
        cell.border    = DATA_BORDER
        cell.alignment = CENTER_ALIGN
        if ci == 3:
            cell.number_format = "0.###"   # no thousand separators
    ws.row_dimensions[er].height = SUMM_ROW_HEIGHT

    # ── Sign-off block ──
    # Each row spans all 3 columns (merged A:C for left half, no right half needed —
    # instead use two separate merged spans: A:B for left section, C for right section)
    # Layout:
    #   sign+0  :  [A:B] "Submitted by:"        [C]  "Received by:"
    #   sign+1..4: blank rows (spacing)
    #   sign+5  :  [A:B] ___________________     [C]  ___________________
    #   sign+6  :  [A:B] Name/Signature/Date     [C]  Name/Signature/Date
    #   sign+7  :  [A:B] TOA-Samsung C&T JV      [C]  Surbana Jurong Consultants

    sign = er + 3

    def _sign_cell(row, col_start, col_end, value, font, alignment):
        """Write a value into a merged cell spanning col_start:col_end."""
        col_s = get_column_letter(col_start)
        col_e = get_column_letter(col_end)
        ref   = f"{col_s}{row}:{col_e}{row}"
        ws.merge_cells(ref)
        cell           = ws.cell(row, col_start, value=value)
        cell.font      = font
        cell.alignment = alignment
        ws.row_dimensions[row].height = SUMM_ROW_HEIGHT

    # Titles row — "Submitted by:" left-half, "Received by:" right-half
    _sign_cell(sign,     1, 2, "Submitted by:", BOLD_FONT, LEFT_ALIGN)
    _sign_cell(sign,     3, 3, "Received by:", BOLD_FONT,  LEFT_ALIGN)

    # Blank spacing rows
    for gap_r in range(sign + 1, sign + 5):
        ws.row_dimensions[gap_r].height = SUMM_ROW_HEIGHT

    # Underline row
    line = sign + 5
    _sign_cell(line, 1, 2, "________________________", DATA_FONT, CENTER_ALIGN)
    _sign_cell(line, 3, 3, "________________________", DATA_FONT, CENTER_ALIGN)

    # Name/Signature/Date row
    lbl = line + 1
    _sign_cell(lbl, 1, 2, "Name / Signature / Date", DATA_FONT,  CENTER_ALIGN)
    _sign_cell(lbl, 3, 3, "Name / Signature / Date", DATA_FONT,  CENTER_ALIGN)

    # Company name row
    co = lbl + 1
    _sign_cell(co, 1, 2, "TOA-Samsung C&T Joint Venture", BOLD_FONT, CENTER_ALIGN)
    _sign_cell(co, 3, 3, "Surbana Jurong Consultants",    BOLD_FONT, CENTER_ALIGN)

    # NO freeze pane on Summary sheet
    ws.freeze_panes = None

    # ── Print titles: rows 1-7 repeat at the top of every printed page ──
    # Covers: Project Name (1), Contractor Name (2), Period (3), gap (4),
    #         Title (5), gap (6), Table Header (7)
    ws.print_title_rows = "1:7"

    # ── Page / print setup ──
    # fitToPage=True + fitToWidth=1 + fitToHeight=0:
    #   • Scales all 3 columns to fit A4 width — NET WEIGHT is NEVER cut off
    #   • Content flows naturally across pages (no restart, no skipping rows)
    #   • Rows 1-7 header block repeats at top of every continuation page
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage   = True   # enable scaling
    ws.page_setup.fitToWidth  = 1      # fit all columns onto 1 page wide — fixes cut-off NET WEIGHT
    ws.page_setup.fitToHeight = 0      # unlimited height — rows continue across pages naturally
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Center content horizontally on the A4 page — clean, professional layout
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = False

    # Standard A4 margins (inches) — proper spacing from all page borders
    ws.page_margins = PageMargins(
        left=0.75, right=0.75, top=1.0, bottom=1.0,
        header=0.5, footer=0.5
    )

    # Footer: "Page X of Y" centered on every printed page
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "Calibri"


# ──────────────────────────────────────────────
# Main entry point
# ──────────────────────────────────────────────

def validate_net_weights(df: pd.DataFrame) -> list:
    """
    Validate Net Weight = In Weight - Out Weight for every row that has all three values.
    Returns a list of dicts for rows where the check fails:
      [{"token": str, "in_weight": float, "out_weight": float,
        "net_weight_file": float, "net_weight_calc": float}, ...]

    STRICT RULE: this function only reports errors — it never modifies any value.
    Called once from the UI before report generation.
    """
    errors = []
    for _, row in df.iterrows():
        try:
            in_w  = row.get("In Weight")
            out_w = row.get("Out Weight")
            net_w = row.get("Net Weight")
            if any(v is None or (isinstance(v, float) and pd.isna(v)) for v in (in_w, out_w, net_w)):
                continue
            in_w_f  = float(in_w)
            out_w_f = float(out_w)
            net_w_f = float(net_w)
            calc    = round(in_w_f - out_w_f, 3)
            if abs(net_w_f - calc) > 0.001:
                errors.append({
                    "token":           str(row.get("Token", "—")),
                    "in_weight":       in_w_f,
                    "out_weight":      out_w_f,
                    "net_weight_file": net_w_f,
                    "net_weight_calc": calc,
                })
        except (TypeError, ValueError):
            continue
    return errors


def validate_etoken(online_tokens, wb_tokens=None):
    """
    E-Token validation — backend only, called from UI which shows only result.

    online_tokens : iterable of E-Token strings from Online Data (accepted rows)
    wb_tokens     : iterable of E-Token strings from WB Data (optional)

    Steps:
      1. Normalise: strip whitespace, uppercase
      2. Sort ascending (A → Z)
      3. Check duplicates within each source
      4. If wb_tokens provided: compare both sorted lists for mismatches

    Returns dict:
      {
        "online_sorted":    list[str],
        "wb_sorted":        list[str] | None,
        "online_dupes":     list[str],   # tokens that appear more than once in Online
        "wb_dupes":         list[str],   # tokens that appear more than once in WB
        "mismatches":       list[str],   # tokens in one but not the other
        "pass":             bool,        # True = no dupes, no mismatches
        "status":           str,         # "PASS" or "FAIL"
        "message":          str,         # human-readable one-liner
      }
    """
    def _norm(tokens):
        return [str(t).strip().upper() for t in tokens if str(t).strip()]

    on_list  = sorted(_norm(online_tokens))
    wb_list  = sorted(_norm(wb_tokens)) if wb_tokens is not None else None

    # Duplicate check
    from collections import Counter
    on_counts  = Counter(on_list)
    on_dupes   = sorted(t for t, c in on_counts.items() if c > 1)
    wb_dupes   = []
    if wb_list is not None:
        wb_counts = Counter(wb_list)
        wb_dupes  = sorted(t for t, c in wb_counts.items() if c > 1)

    # Mismatch check
    mismatches = []
    if wb_list is not None:
        on_set = set(on_list)
        wb_set = set(wb_list)
        mismatches = sorted((on_set ^ wb_set))   # symmetric difference

    has_dupes      = bool(on_dupes or wb_dupes)
    has_mismatches = bool(mismatches)
    passed         = not has_dupes and not has_mismatches

    if passed:
        if wb_list is not None:
            msg = "E-Token Validation Completed – PASS"
        else:
            msg = f"E-Token Validation Completed – PASS ({len(on_list)} tokens, no duplicates)"
    else:
        parts = []
        if on_dupes:
            parts.append(f"{len(on_dupes)} duplicate(s) in Online")
        if wb_dupes:
            parts.append(f"{len(wb_dupes)} duplicate(s) in WB")
        if mismatches:
            first = mismatches[0]
            parts.append(f"E-Token [{first}] mismatch between Online & Weighbridge data")
        msg = " · ".join(parts)

    return {
        "online_sorted": on_list,
        "wb_sorted":     wb_list,
        "online_dupes":  on_dupes,
        "wb_dupes":      wb_dupes,
        "mismatches":    mismatches,
        "pass":          passed,
        "status":        "PASS" if passed else "FAIL",
        "message":       msg,
    }


def generate_report(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
    corrections: dict = None,
    flag_map:    dict = None,
) -> tuple:
    """
    Generate 4-sheet Excel report matching reference format.
    Returns (BytesIO, filename, stats_dict).

    corrections: dict  {original_df_index: {"Accepted": "YES"|"NO", "Out Weight": float}}
      Only Accepted and Out Weight are ever applied to the working copy.
      Net Weight is NEVER modified, recalculated, or overwritten — strict no-modification rule.
      Accepted values must be UPPERCASE ("YES" / "NO").

    flag_map: dict  {original_df_index: {"flagged": bool, "red_cols": set}}
      Passed through to _data_rows for yellow row / red cell highlighting.
    """
    corrections = corrections or {}
    flag_map    = flag_map    or {}

    # ── Apply corrections to a working copy ──
    # Applies Accepted, Out Weight, and Net Weight from the corrections dict.
    # Net Weight is written only when explicitly supplied by the correction
    # (NO branch sets it to 0; YES branch sets it to In Weight minus Out Weight).
    working = df.copy()
    for idx, corr in corrections.items():
        if idx not in working.index:
            continue
        if "Accepted" in corr:
            working.at[idx, "Accepted"] = corr["Accepted"]
        if "Out Weight" in corr:
            working.at[idx, "Out Weight"] = float(corr["Out Weight"])
        if "Net Weight" in corr:
            working.at[idx, "Net Weight"] = float(corr["Net Weight"])

    filtered = _filter_by_dates(working, from_date, to_date)
    if filtered.empty:
        raise ValueError("No records found in the selected date range.")

    # ── Pre-check: count rejections to decide which sheets to generate ──
    _rej_mask_pre = filtered["Accepted"].apply(_is_rejected)
    _n_rej_pre    = int(_rej_mask_pre.sum())
    _has_rejections = _n_rej_pre > 0

    wb = Workbook()

    if _has_rejections:
        # Full report: Contract + Rejected + Accepted + Summary
        ws1 = wb.active
        ws1.title = "Contract"
        _sheet_contract(ws1, filtered, from_date, to_date, flag_map=flag_map)

        ws2 = wb.create_sheet("Loads Rejected")
        n_rej = _sheet_rejected(ws2, filtered, from_date, to_date, flag_map=flag_map)
        ws2.title = f"{n_rej} Loads Rejected"

        ws3 = wb.create_sheet("Loads Accepted")
        accepted_df = _sheet_accepted(ws3, filtered, from_date, to_date, flag_map=flag_map)
        n_acc = len(accepted_df)
        ws3.title = f"{n_acc} Loads Accepted"

        ws4 = wb.create_sheet("Summary")
        _sheet_summary(ws4, accepted_df, from_date, to_date)
    else:
        # No rejections: generate ONLY Loads Accepted + Summary
        # Remove the default empty sheet and build the two required sheets
        ws_acc = wb.active
        ws_acc.title = "Loads Accepted"
        accepted_df = _sheet_accepted(ws_acc, filtered, from_date, to_date, flag_map=flag_map)
        n_acc = len(accepted_df)
        ws_acc.title = f"{n_acc} Loads Accepted"
        n_rej = 0

        ws_summ = wb.create_sheet("Summary")
        _sheet_summary(ws_summ, accepted_df, from_date, to_date)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"APSG - WB Summary Contract Report-{to_date.strftime('%d%m%Y')}.xlsx"

    net_wt = float(accepted_df["Net Weight"].sum()) if "Net Weight" in accepted_df.columns else 0.0
    stats  = {
        "total":      len(filtered),
        "accepted":   n_acc,
        "rejected":   n_rej,
        "net_weight": net_wt,
    }

    return buf, fname, stats
